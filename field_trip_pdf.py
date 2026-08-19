"""
field_trip_pdf.py - The BSD field trip application, filled in from the planner.

Two forms, because BSD runs two procedures:

  * "Day Field Trip Application" (2320P, rev. 9/25/2024)
  * "Out of State, Overnight or International Field Trip Planning" (2320P)

Laid out to match the district's own forms field for field and in their order,
the same way the instrument loan form matches the Charms original -- an office
manager should be able to read it without hunting for anything.  What Roka
adds is that the answers are already in it: the roster count, the cost per
student, the dates, the objective, the four narrative answers.

Blank stays blank.  A blank on a form is a question still to be answered, and
filling it with "(none)" or a guess hides that from the person signing it.
"""

import os
from datetime import datetime

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table, TableStyle,
                                Spacer, KeepTogether, HRFlowable, PageBreak)

import field_trip_tools as ft
from concert_tools import fmt_date

CW = 7.0 * inch
_LINE = colors.HexColor("#000000")
_GREY = colors.HexColor("#555555")
_REVISED = "Revised 9/25/2024"


def _ps(name, size=9, bold=False, align=TA_LEFT, leading=None, italic=False,
        color=None):
    font = ("Helvetica-BoldOblique" if (bold and italic) else
            "Helvetica-Bold" if bold else
            "Helvetica-Oblique" if italic else "Helvetica")
    return ParagraphStyle(name, fontName=font, fontSize=size, alignment=align,
                          leading=leading or (size + 2), spaceAfter=0,
                          spaceBefore=0, textColor=color or colors.black)


def _styles():
    return {
        "district": _ps("district", 11, bold=True, align=TA_CENTER),
        "title": _ps("title", 17, bold=True, align=TA_CENTER, leading=20),
        "subtitle": _ps("subtitle", 8.5, bold=True, align=TA_CENTER, leading=11),
        "lbl": _ps("lbl", 8.5, leading=10.5),
        "lblb": _ps("lblb", 8.5, bold=True, leading=10.5),
        "val": _ps("val", 9, leading=11),
        "tiny": _ps("tiny", 6.5, italic=True, leading=8),
        "tinyr": _ps("tinyr", 6.5, italic=True, leading=8,
                     color=colors.HexColor("#B00000")),
        "q": _ps("q", 8.5, bold=True, leading=11),
        "a": _ps("a", 9, leading=12),
        "foot": _ps("foot", 7.5, color=_GREY),
        "footr": _ps("foot_r", 7.5, align=TA_CENTER, color=_GREY),
    }


def _money(v, blank=""):
    try:
        f = float(v or 0)
    except (TypeError, ValueError):
        return blank
    if not f:
        return blank
    return f"{f:,.2f}"


def _cost_cell(v):
    """What goes in a Trip Costs box.

    Three different things were all printing as an empty box, and only one of
    them meant "empty":

      nothing typed  -> "$"          the question is still open
      0              -> "$ 0.00"     answered, and the answer is nothing
      "TBD" / "N/A"  -> "TBD"        answered, and the answer is not a number

    A zero that prints blank is the planner quietly editing the teacher's
    answer, and on a form somebody signs that is not a small thing.
    """
    raw = ("" if v is None else str(v)).strip()
    if not raw:
        return "$"
    try:
        return f"$ {float(raw.replace('$', '').replace(',', '')):,.2f}"
    except ValueError:
        return raw


def _is_number(v):
    try:
        float(str(v).replace("$", "").replace(",", "").strip() or 0)
        return True
    except ValueError:
        return False


def _unknowns(trip):
    """The cost boxes holding words rather than figures, so a total can say
    what it does not include instead of pretending to be complete."""
    out = []
    for key in ("entry_fee", "transport_cost", "sub_cost", "food_cost",
                "other_cost"):
        raw = ("" if trip.get(key) is None else str(trip.get(key))).strip()
        if raw and not _is_number(raw):
            out.append(raw)
    return out


def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _checkbox(checked=False):
    """A small square, ticked or not.  Drawn rather than typed: the standard
    PDF fonts have no ballot-box glyph, so a literal box character prints as
    nothing or as a black blob depending on the reader."""
    mark = Paragraph('<font size=7><b>X</b></font>',
                     _ps("cbx", 7, bold=True, align=TA_CENTER, leading=8)) \
        if checked else ""
    t = Table([[mark]], colWidths=[9], rowHeights=[9])
    t.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, _LINE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t


def _grid(rows, s, label_w=1.55, value_w=1.95):
    """The district header block: label, boxed value, label, boxed value.

    ``rows`` is a list of ((label, value), (label, value) | None).  A row whose
    right half is None spans its value across the rest of the line, which is
    how the overnight form does Destination and Educational Objective.
    """
    data, spans, boxes = [], [], []
    for i, (left, right) in enumerate(rows):
        if right is None:
            data.append([Paragraph(left[0], s["lbl"]),
                         Paragraph(left[1] or "", s["val"]), "", ""])
            spans.append(("SPAN", (1, i), (3, i)))
            boxes.append(("BOX", (1, i), (3, i), 0.7, _LINE))
        else:
            data.append([Paragraph(left[0], s["lbl"]),
                         Paragraph(left[1] or "", s["val"]),
                         Paragraph(right[0], s["lbl"]),
                         Paragraph(right[1] or "", s["val"])])
            boxes.append(("BOX", (1, i), (1, i), 0.7, _LINE))
            boxes.append(("BOX", (3, i), (3, i), 0.7, _LINE))
    t = Table(data, colWidths=[label_w * inch, value_w * inch,
                               label_w * inch, value_w * inch])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ] + spans + boxes))
    return t


def _sub_row(trip, s, note=True, note_w=1.95):
    """Substitute teacher – check one: the district's three rates, with the
    one recorded in Roka ticked.

    The day form carries the rate footnote on the same line; the overnight
    form puts it on the line below, beside the Total.  Same row builder, two
    callers, matching each form as printed.
    """
    chosen = (trip.get("sub_rate") or "").strip()
    cells, widths = [], []
    for key, label, _amt in ft.SUB_RATES:
        # Box BEFORE its rate, not after.  The district's own form puts it
        # after -- "$212/4 hrs ☐ $266/5 hrs ☐" -- and at this width that reads
        # as though the tick belongs to the NEXT rate: a $212 substitute
        # printed with the box beside $266/5 hrs, on a form somebody signs.
        # Same three rates, same order, unambiguous.
        cells.append(_checkbox(chosen == key))
        widths.append(0.17 * inch)
        cells.append(Paragraph(label.replace(" ", "&nbsp;"), s["lbl"]))
        widths.append(0.80 * inch)
    if note:
        cells.append(Paragraph(f"* {ft.SUB_RATE_NOTE}", s["tinyr"]))
        widths.append(note_w * inch)
    inner = Table([cells], colWidths=widths)
    inner.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 1),
        # Tight after a box, roomy after its label: that spacing is what makes
        # each tick visibly belong to the rate it sits beside.
        ("RIGHTPADDING", (0, 0), (0, -1), 3),
        ("RIGHTPADDING", (1, 0), (1, -1), 10),
        ("RIGHTPADDING", (3, 0), (3, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    return inner


def _costs_day(trip, s):
    """The day form's cost table: one column of rows, total boxed at the foot."""
    rows = [
        ("Entry fee/participation:", _cost_cell(trip.get("entry_fee"))),
        ("Transportation:", _cost_cell(trip.get("transport_cost"))),
        None,                                   # substitute row, built below
        ("Food:", _cost_cell(trip.get("food_cost"))),
        ("Other:", _cost_cell(trip.get("other_cost"))),
    ]
    data, styles = [], []
    for i, r in enumerate(rows):
        if r is None:
            data.append([Paragraph("Substitute teacher – check one:", s["lbl"]),
                         _sub_row(trip, s)])
        else:
            data.append([Paragraph(r[0], s["lbl"]), Paragraph(r[1], s["val"])])
    data.append([Paragraph("<b>Total Trip Cost:</b>", s["lblb"]),
                 Paragraph("<b>" + _total_cell(trip) + "</b>", s["val"])])
    t = Table(data, colWidths=[2.10 * inch, 4.90 * inch])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.7, _LINE),
        ("LINEBELOW", (0, len(data) - 2), (-1, len(data) - 2), 1.2, _LINE),
        ("BOX", (0, len(data) - 1), (-1, len(data) - 1), 1.2, _LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ] + styles))
    return t


def _costs_overnight(trip, s):
    """The overnight form pairs the costs two across, with Total beside the
    substitute rates."""
    data = [
        [Paragraph("Entry fee/participation:", s["lbl"]),
         Paragraph(_cost_cell(trip.get("entry_fee")), s["val"]),
         Paragraph("Food:", s["lbl"]),
         Paragraph(_cost_cell(trip.get("food_cost")), s["val"])],
        [Paragraph("Transportation:", s["lbl"]),
         Paragraph(_cost_cell(trip.get("transport_cost")), s["val"]),
         Paragraph("Other:", s["lbl"]),
         Paragraph(_cost_cell(trip.get("other_cost")), s["val"])],
    ]
    t = Table(data, colWidths=[1.55 * inch, 1.95 * inch, 1.55 * inch, 1.95 * inch])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.7, _LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    sub = Table(
        [[Paragraph("Substitute teacher – check one:", s["lbl"]),
          _sub_row(trip, s, note=False), "", ""],
         [Paragraph(f"* {ft.SUB_RATE_NOTE}", s["tinyr"]), "",
          Paragraph("<b>Total:</b>", s["lblb"]),
          Paragraph("<b>" + _total_cell(trip) + "</b>", s["val"])]],
        colWidths=[1.90 * inch, 3.00 * inch, 0.60 * inch, 1.50 * inch])
    sub.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("SPAN", (1, 0), (3, 0)),
        ("SPAN", (0, 1), (1, 1)),
        ("GRID", (0, 0), (-1, -1), 0.7, _LINE),
        ("BOX", (2, 1), (3, 1), 1.2, _LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return [t, sub]


def _total(trip):
    return sum(_num(trip.get(k)) for k in
               ("entry_fee", "transport_cost", "sub_cost", "food_cost",
                "other_cost"))


def _total_cell(trip):
    """The Total box.  Adds up what is a number and names what is not."""
    unknown = _unknowns(trip)
    filled = any(("" if trip.get(k) is None else str(trip.get(k))).strip()
                 for k in ("entry_fee", "transport_cost", "sub_cost",
                           "food_cost", "other_cost"))
    if not filled:
        return "$"
    text = f"$ {_total(trip):,.2f}"
    if unknown:
        text += "  + " + " + ".join(sorted(set(unknown)))
    return text


def _question(question, answer, s, lines=2):
    """A district question with its answer boxed underneath.

    ``lines`` is how much room to leave when there is no answer yet -- roughly
    that many handwritten lines.  An answer that is already filled in gets no
    blank room, which is what keeps the day form on one page as the district's
    own does.
    """
    answer = (answer or "").strip()
    body = Paragraph(answer.replace("\n", "<br/>") or "&nbsp;", s["a"])
    box = Table([[body]], colWidths=[CW])
    box.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, _LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4 if answer else lines * 11),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]))
    return KeepTogether([Paragraph(question, s["q"]), Spacer(1, 1), box,
                         Spacer(1, 5)])


def _approval(s):
    line = Table([[Paragraph("", s["val"])]], colWidths=[2.6 * inch])
    line.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.7, _LINE)]))
    date_line = Table([[Paragraph("", s["val"])]], colWidths=[1.7 * inch])
    date_line.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.7, _LINE)]))
    boxes = Table([[_checkbox(False), Paragraph("Trip Approved", s["val"])],
                   [_checkbox(False), Paragraph("Trip Denied", s["val"])]],
                  colWidths=[0.22 * inch, 1.4 * inch])
    boxes.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    t = Table([[boxes, line, date_line],
               ["", Paragraph("Administrator's Signature", s["tiny"]),
                Paragraph("Date", s["tiny"])]],
              colWidths=[1.9 * inch, 3.1 * inch, 2.0 * inch])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("ALIGN", (1, 1), (2, 1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t


def _supervision_answer(chaperones, teacher_name, overnight, written=None):
    """What the teacher wrote, or a count assembled from the chaperone list.

    The assembled version used to be the only version, which meant her own name
    appeared on a form she had never typed it into.  The trip dialog shows this
    field now and says what happens if it is left blank.
    """
    written = (written or "").strip()
    if written:
        return written
    if not chaperones and not teacher_name:
        return ""
    bits = []
    if chaperones:
        bits.append(f"{chaperones} adult chaperone(s)")
    if teacher_name:
        bits.append(f"plus {teacher_name}")
    text = ", ".join(bits) + "."
    if overnight:
        text += "  Non-BSD personnel VIBES cleared:  ______"
    return text


def build_application(trip, path, students=0, chaperones=0, teacher_name="",
                      school_name="", overnight=None):
    """Write the filled application to ``path`` and return it."""
    s = _styles()
    is_overnight = (ft.trip_type(trip) == ft.TRIP_OVERNIGHT
                    if overnight is None else bool(overnight))

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica-Oblique", 7.5)
        canvas.setFillColor(_GREY)
        canvas.drawString(0.75 * inch, 0.45 * inch, _REVISED)
        canvas.drawRightString(letter[0] - 0.75 * inch, 0.45 * inch,
                               f"{doc.page} | Page")
        canvas.drawCentredString(letter[0] / 2, 0.45 * inch,
                                 "Prepared in Roka's Resonance")
        canvas.restoreState()

    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.62 * inch,
                            title=trip.get("name") or "Field Trip Application")
    flow = []

    # ── Header ──
    flow.append(Paragraph("Bellevue School District", s["district"]))
    flow.append(Paragraph(
        "Out of State, Overnight or International Field Trip Planning"
        if is_overnight else "Day Field Trip Application", s["title"]))
    flow.append(Spacer(1, 2))
    flow.append(Paragraph(
        "(Not for Field Trips scheduled during the school day)" if is_overnight
        else "For School Day or Outside of School Hours. "
             "(Not for Out of State, Overnight or International Field Trips)",
        s["subtitle"]))
    flow.append(Spacer(1, 4))
    flow.append(HRFlowable(width="100%", thickness=1, color=_LINE,
                           spaceBefore=0, spaceAfter=6))

    when_req = trip.get("date_of_request") or datetime.today().strftime("%Y-%m-%d")
    per = ""
    if students:
        if trip.get("covered"):
            per = "$0.00 (covered)"
        else:
            per = f"${_total(trip) / students:,.2f}"
            if _unknowns(trip):
                per += "  + " + " + ".join(sorted(set(_unknowns(trip))))
    n_students = str(students) if students else ""
    n_chaps = str(chaperones) if chaperones else ""

    if is_overnight:
        rows = [
            (("Date of Request:", fmt_date(when_req) or when_req),
             ("Teacher/Advisor Name:", teacher_name)),
            (("Class or Group:", trip.get("groups_list")),
             ("Teacher/Advisor Cell Phone:", trip.get("advisor_phone"))),
            (("Departure Date:", fmt_date(trip.get("depart_date"))),
             ("Return Date:", fmt_date(ft.effective_return_date(trip)))),
            (("Departure Time from School:", trip.get("depart_time")),
             ("Arrival at Destination:", trip.get("arrive_dest_time"))),
            (("Departure from Destination:", trip.get("depart_dest_time")),
             ("Arrive at School:", trip.get("return_time"))),
            (("Number of Students:", n_students),
             ("Number of Chaperones:", n_chaps)),
            (("Method of Travel:", trip.get("travel_method")),
             ("Charge to Budget Code:", trip.get("budget_code"))),
            (("Anticipated Cost / Student:", per),
             ("Dept. Chair's Signature:", "")),
        ]
    else:
        rows = [
            (("Date of Request:", fmt_date(when_req) or when_req),
             ("Trip Destination:", trip.get("destination"))),
            (("Class or Group:", trip.get("groups_list")),
             ("Educational Objectives:", trip.get("educational_objective"))),
            (("Teacher/Advisor Name:", teacher_name),
             ("Number of Students:", n_students)),
            (("Departure Date:", fmt_date(trip.get("depart_date"))),
             ("Anticipated Cost / Student:", per)),
            (("Departure Time:", trip.get("depart_time")),
             ("Method of Travel:", trip.get("travel_method"))),
            (("Return Date:", fmt_date(ft.effective_return_date(trip))),
             ("Charge to Budget Code:", trip.get("budget_code"))),
            (("Return Time:", trip.get("return_time")),
             ("Dept. Chair's Signature:", "")),
        ]
    flow.append(_grid(rows, s))
    if not is_overnight:
        # Sits under "Anticipated Cost / Student", as on the district form.
        note = Table([["", Paragraph("See cost breakdown section below",
                                     s["tiny"])]],
                     colWidths=[5.05 * inch, 1.95 * inch])
        note.setStyle(TableStyle([
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ]))
        flow.append(note)

    if is_overnight:
        flow.append(Spacer(1, 6))
        flow.append(_grid([
            (("Destination of Trip:", trip.get("destination")), None),
            (("Destination Address/Contact:", trip.get("dest_address")), None),
            (("Educational Objective:", trip.get("educational_objective")), None),
        ], s))

    # ── Trip costs ──
    flow.append(Spacer(1, 7))
    flow.append(Paragraph("<b>Trip Costs:</b>", _ps("tc", 9.5, bold=True)))
    flow.append(Spacer(1, 2))
    if is_overnight:
        for piece in _costs_overnight(trip, s):
            flow.append(piece)
    else:
        flow.append(_costs_day(trip, s))
    flow.append(Spacer(1, 8))

    # ── The narrative questions, in the district's order and wording ──
    flow.append(_question(
        "Describe activities planned while on the trip"
        + ("" if is_overnight else
           " (use back of form or additional paper if necessary)") + ":",
        trip.get("activities"), s, lines=2))
    flow.append(_question(
        "What required assignments will participants have to complete related "
        "to this activity? What alternate assignments will be available to "
        "students who miss the activity?", trip.get("assignments"), s,
        lines=2))
    flow.append(_question(
        "What arrangements have been made for students to complete work "
        "missed in other classes?", trip.get("missed_work"), s, lines=1))
    flow.append(_question(
        "How many adults will provide supervision"
        + (" – and have non-BSD personnel been VIBES cleared?"
           if is_overnight else "?"),
        _supervision_answer(chaperones, teacher_name, is_overnight,
                            trip.get("supervision")), s, lines=1))
    flow.append(_question(
        "What considerations have been made for students who cannot afford "
        "the cost of the trip?", trip.get("affordability"), s, lines=2))
    flow.append(_question(
        "Have you reviewed student health needs and discussed with School "
        "Nurse?", trip.get("health_review"), s, lines=1))

    if is_overnight and (trip.get("itinerary") or "").strip():
        flow.append(_question("Itinerary:", trip.get("itinerary"), s,
                              lines=4))

    flow.append(Spacer(1, 4))
    flow.append(_approval(s))

    doc.build(flow, onFirstPage=_footer, onLaterPages=_footer)
    return path


def suggested_filename(trip, overnight=None):
    is_overnight = (ft.trip_type(trip) == ft.TRIP_OVERNIGHT
                    if overnight is None else bool(overnight))
    name = "".join(c for c in (trip.get("name") or "Field Trip")
                   if c.isalnum() or c in " -_").strip() or "Field Trip"
    kind = "Overnight" if is_overnight else "Day"
    when = (trip.get("depart_date") or datetime.today().strftime("%Y-%m-%d"))
    return f"{name} - {kind} Field Trip Application - {when}.pdf"


# ── Permission forms, one per student ────────────────────────────────────────
# 2320P Exhibit C's own directions: "(2) Complete the school portion (top half)
# of form, (3) Duplicate one form per student, (4) Send a copy home for parent
# and student signatures."  Step 3 is the one that costs an evening: the school
# portion is identical on all of them and the only thing that differs is the
# name at the top.  So Roka fills the top half once and stamps a page per
# student, in one PDF, ready for the copier.
#
# Exhibit A is the elementary day-trip equivalent and takes the same treatment.
#
# The notary page is NOT produced.  Exhibit C carries one, and the form itself
# says it is required for international trips only -- which Roka does not
# cover.  Printing it would be inviting somebody to go and find a notary for a
# trip to Ellensburg.

_TRANSPORT_TICKS = ["Airline", "School Bus", "Commercial Carrier",
                    "Leased Vehicle", "District Vehicle", "Other"]

_TRANSPORT_MATCH = {
    "school bus": "School Bus",
    "charter bus": "Commercial Carrier",
    "public transit": "Commercial Carrier",
    "private vehicles": "Other",
    "walking": "Other",
}


def _transport_tick(travel_method):
    """Which of Exhibit C's boxes the planner's travel method means."""
    return _TRANSPORT_MATCH.get((travel_method or "").strip().lower())


def _pupil_line(label, width, s):
    t = Table([[Paragraph("", s["val"])]], colWidths=[width * inch])
    t.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.7, _LINE),
                           ("TOPPADDING", (0, 0), (-1, -1), 10)]))
    return [t, Paragraph(label, s["tiny"]), Spacer(1, 6)]


def build_permission_forms(trip, students, path, teacher_name="",
                           school_name="", elementary=False):
    """One PDF, one permission form per student, school portion filled in."""
    s = _styles()
    overnight = ft.trip_type(trip) == ft.TRIP_OVERNIGHT
    exhibit = "Exhibit A" if (elementary and not overnight) else "Exhibit C"
    title = ("PARENT AUTHORIZATION AND ACKNOWLEDGEMENT OF RISK "
             "FOR OUT OF STATE OR OVERNIGHT FIELD TRIP" if overnight else
             "PARENT AUTHORIZATION AND ACKNOWLEDGEMENT OF RISK FOR FIELD TRIP")

    when = fmt_date(trip.get("depart_date")) or ""
    back = fmt_date(trip.get("return_date")) or ""
    dates = when if (not back or back == when) else f"{when} to {back}"
    ticked = _transport_tick(trip.get("travel_method"))

    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=0.55 * inch, bottomMargin=0.6 * inch,
                            title=f"{exhibit} — {trip.get('name') or 'Field Trip'}")

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica-Oblique", 7.5)
        canvas.setFillColor(_GREY)
        canvas.drawString(0.75 * inch, 0.42 * inch,
                          f"PROCEDURE 2320P {exhibit.upper()}")
        canvas.drawRightString(letter[0] - 0.75 * inch, 0.42 * inch,
                               "Prepared in Roka's Resonance")
        canvas.restoreState()

    flow = []
    for i, stu in enumerate(students):
        if i:
            flow.append(PageBreak())
        name = _student_name(stu)
        sid = ""
        try:
            sid = (stu.get("student_id") or "").strip()
        except Exception:
            pass

        flow.append(Paragraph(f"PROCEDURE 2320P {exhibit.upper()}", s["district"]))
        flow.append(Paragraph(title, _ps("exh", 11, bold=True, align=TA_CENTER,
                                         leading=14)))
        flow.append(Paragraph(
            "(This form and an attached field trip description are required "
            "for all out-of-state or overnight trips.)" if overnight else
            "(Complete and return to your child's music teacher.)",
            s["subtitle"]))
        flow.append(Spacer(1, 8))

        flow.append(_grid([
            (("Name of Student and Student Id#",
              name + (f"   —   {sid}" if sid else "")), None),
            (("Date(s) of Trip:", dates),
             ("Destination:", trip.get("destination"))),
            (("Purpose:", trip.get("educational_objective")), None),
            (("Name of Employee:", teacher_name),
             ("School:", school_name)),
        ], s))
        flow.append(Spacer(1, 6))
        flow.append(Paragraph(
            "Is the District employee responsible for the trip and may be "
            "accompanied by other District staff and approved volunteer "
            "chaperones. They have my permission to do so.", s["val"]))
        flow.append(Spacer(1, 8))

        # Transportation, ticked from the planner where it can be.
        flow.append(Paragraph("TRANSPORTATION BEING PROVIDED BY "
                              "(Check all that apply)", s["q"]))
        cells, widths = [], []
        for label in _TRANSPORT_TICKS:
            cells.append(_checkbox(ticked == label))
            widths.append(0.20 * inch)
            cells.append(Paragraph(label, s["lbl"]))
            widths.append(1.05 * inch)
        tt = Table([cells], colWidths=widths)
        tt.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                                ("TOPPADDING", (0, 0), (-1, -1), 3),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        flow.append(tt)

        flow.append(Paragraph("DRIVERS OF DISTRICT, PRIVATE OR LEASED VEHICLES "
                              "(Check all that apply.)", s["q"]))
        dcells, dwidths = [], []
        for label in ("Parent", "Teacher or Staff Member", "Other"):
            dcells.append(_checkbox(False))
            dwidths.append(0.20 * inch)
            dcells.append(Paragraph(label, s["lbl"]))
            dwidths.append(1.60 * inch)
        dt = Table([dcells], colWidths=dwidths)
        dt.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                                ("TOPPADDING", (0, 0), (-1, -1), 3),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        flow.append(dt)
        flow.append(Paragraph(
            "If travel by private car is involved, your student will ride "
            "with ____________________________  (name of driver), "
            "____________________  (telephone).", s["val"]))
        flow.append(Spacer(1, 4))
        flow.append(Paragraph(
            "Please Note: School staff ensures that all drivers and vehicles "
            "are approved by the District Transportation Department before "
            "driving students.", s["tiny"]))
        if overnight:
            flow.append(Spacer(1, 4))
            flow.append(Paragraph(
                "An itinerary for the trip (detailing dates, place of lodging, "
                "events, etc.) is attached for your information.", s["val"]))

        flow.append(Spacer(1, 10))
        flow.append(HRFlowable(width="100%", thickness=1, color=_LINE,
                               spaceBefore=0, spaceAfter=8))
        flow.append(Paragraph("TO BE COMPLETED AT HOME", s["q"]))
        flow.append(Spacer(1, 4))
        flow.append(Paragraph("Pupil Agreement", _ps("pa", 9.5, bold=True,
                                                     align=TA_CENTER)))
        flow.append(Paragraph(
            "While participating in this field trip, I will accept "
            "responsibility for abiding by all District and school rules, "
            "regulations, policies and procedures; following the directions of "
            "staff and volunteer chaperones; and the expectations set by "
            "advisors. Any incidents of exceptional misconduct as defined in "
            "District Procedure 3241P may result in my being sent home at the "
            "expense of my family.", s["val"]))
        flow.append(Spacer(1, 6))
        for label, w in [("Signature of Student", 3.2),
                         ("Signature of Parent/Guardian", 3.2)]:
            flow.extend(_pupil_line(label, w, s))

        flow.append(Paragraph("PARENTAL AUTHORIZATION AND ACKNOWLEDGEMENT OF "
                              "RISKS", _ps("par", 9.5, bold=True,
                                           align=TA_CENTER)))
        flow.append(Paragraph(
            "If an emergency situation involving illness and/or injury should "
            "arise, the Bellevue district staff member in charge has my "
            "permission to seek the aid of medical professionals for emergency "
            "care. In the event it becomes necessary to obtain emergency care "
            "for your student, neither s/he nor the Bellevue School District "
            "assumes financial liability for expenses incurred because of "
            "accident, injury, illness and/or unforeseen circumstances.",
            s["val"]))
        flow.append(Spacer(1, 4))
        flow.append(Paragraph(
            "I understand that participation in this field trip is voluntary, "
            "that it is not required, and that it may expose my child to some "
            "risk(s). I have read and understand the description of the field "
            "trip and authorize my child to participate in the planned "
            "components of the field trip. I also understand that "
            "participation will involve activities off school property; "
            "therefore, neither the Bellevue School District, nor its "
            "employees and volunteers, will have any responsibility for the "
            "condition or use of any non-school property.", s["val"]))
        flow.append(Spacer(1, 8))
        flow.append(Paragraph(
            f"I give permission for <b>{name}</b> to participate in this "
            f"field trip.", s["val"]))
        flow.append(Spacer(1, 4))
        for label, w in [("Signature of Parent or Guardian", 3.6),
                         ("Parent or Guardian Phone Number", 3.0)]:
            flow.extend(_pupil_line(label, w, s))

    doc.build(flow, onFirstPage=_footer, onLaterPages=_footer)
    return path


def _student_name(stu):
    try:
        first = (stu.get("preferred_name") or stu.get("first_name") or "").strip()
        last = (stu.get("last_name") or "").strip()
        full = f"{first} {last}".strip()
        return full or (stu.get("name") or "").strip()
    except Exception:
        return ""


def suggested_permission_filename(trip):
    name = "".join(c for c in (trip.get("name") or "Field Trip")
                   if c.isalnum() or c in " -_").strip() or "Field Trip"
    return f"{name} - Permission Forms.pdf"



# ── The student list, as a spreadsheet ───────────────────────────────────────
# A list of names belongs in a grid, not in the body of an email.  The office
# sorts it by grade, the attendance secretary filters it, and a teacher looking
# for their own period stops reading forty-seven other names to find three.

def build_student_list(trip, students, path, teacher_name="", school_name=""):
    """Write the attending students to an .xlsx and return the path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    import field_trip_tools as _ft
    rows = _ft.attending_rows(students)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    name = (trip.get("name") or "Field trip").strip()
    when = fmt_date(trip.get("depart_date")) or ""
    ws.append([name])
    ws.append([f"{when}"
               + (f", leaving {trip.get('depart_time')}"
                  if trip.get("depart_time") else "")
               + (f", back {trip.get('return_time')}"
                  if trip.get("return_time") else "")])
    ws.append([f"{len(rows)} student(s)"
               + (f"   ·   {trip.get('groups_list')}"
                  if trip.get("groups_list") else "")])
    if teacher_name or school_name:
        ws.append([("   ·   ".join(x for x in (teacher_name, school_name) if x))])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=14)

    # Read the row back AFTER writing it.  append([]) advances the write
    # position but leaves max_row where it was, because a row with no cells in
    # it is not a row -- so predicting the header's position put the shading
    # and the filter arrows one row above the words.
    ws.append(["Last Name", "First Name", "Grade", "Student ID", "Class"])
    head_at = ws.max_row
    for c in ws[head_at]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5FA3")
        c.alignment = Alignment(horizontal="left")
    for r in rows:
        ws.append([r["last_name"], r["first_name"], r["grade"],
                   r["student_id"], r["ensembles"]])

    for col, width in zip("ABCDE", (20, 18, 8, 14, 34)):
        ws.column_dimensions[col].width = width
    # Frozen under the header, so the names stay labelled on a long list.
    ws.freeze_panes = ws.cell(row=head_at + 1, column=1)
    ws.auto_filter.ref = (f"A{head_at}:E{head_at + len(rows)}")

    wb.save(path)
    return path


def suggested_student_list_filename(trip):
    name = "".join(c for c in (trip.get("name") or "Field Trip")
                   if c.isalnum() or c in " -_").strip() or "Field Trip"
    when = (trip.get("depart_date") or "").strip()
    return f"{name} - Student List{(' - ' + when) if when else ''}.xlsx"
