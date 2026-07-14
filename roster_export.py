"""
roster_export.py - Export a roster of selected ensembles to an Excel (.xlsx)
spreadsheet: Student Name, Grade, Student ID.

Used for field-trip lists and in-school performance pull-out lists, where a
director hands the office / substitute teachers a simple who-is-out sheet.  No
tkinter here (that's ``ui/roster_export_view.py``); pure data + openpyxl so it
can be tested headless.
"""

from __future__ import annotations


def _ensembles_of(student):
    return [e.strip() for e in (student.get("ensembles") or "").split(",")
            if e.strip()]


def filter_students(students, ensembles):
    """Return roster rows (dicts: name, grade, student_id) for students in any of
    ``ensembles`` (a set/list of class labels).  ``ensembles`` empty/None means
    every student who is in at least one ensemble.  Sorted by last, first."""
    want = {e.strip() for e in (ensembles or []) if e and e.strip()}
    rows = []
    for s in students:
        if not (s.get("is_active", 1) in (1, None) or s.get("is_active") == 1):
            # treat missing is_active as active
            if s.get("is_active") == 0:
                continue
        mine = _ensembles_of(s)
        if want:
            if not (set(mine) & want):
                continue
        elif not mine:
            continue
        first = (s.get("first_name") or "").strip()
        last = (s.get("last_name") or "").strip()
        name = f"{last}, {first}".strip(", ").strip() or (first or last)
        rows.append({"name": name,
                     "grade": (s.get("grade") or "").strip(),
                     "student_id": (s.get("student_id") or "").strip(),
                     "_sort": (last.lower(), first.lower())})
    rows.sort(key=lambda r: r["_sort"])
    for r in rows:
        r.pop("_sort", None)
    return rows


def filter_full(students, ensembles):
    """Full student dicts (all columns) for students in any of ``ensembles``
    (empty/None = every active student in at least one ensemble).  Used for the
    handoff export, which needs instruments + contacts, not just name/grade/id."""
    want = {e.strip() for e in (ensembles or []) if e and e.strip()}
    out = []
    for s in students:
        if s.get("is_active") == 0:
            continue
        mine = _ensembles_of(s)
        if want:
            if not (set(mine) & want):
                continue
        elif not mine:
            continue
        out.append(dict(s))
    return out


def write_roster_xlsx(rows, out_path, title="Roster", subtitle=""):
    """Write rows (from ``filter_students``) to an .xlsx with a header row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Roster")[:31]

    r = 1
    if subtitle:
        ws.cell(row=r, column=1, value=subtitle).font = Font(bold=True, size=12)
        r += 2
    headers = ["Student Name", "Grade", "Student ID"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    r += 1
    for row in rows:
        ws.cell(row=r, column=1, value=row["name"])
        ws.cell(row=r, column=2, value=row["grade"])
        ws.cell(row=r, column=3, value=row["student_id"])
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    wb.save(out_path)
    return out_path
