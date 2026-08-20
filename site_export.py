"""
site_export.py - Hand one school's instruments on to whoever teaches it next.

Elementary assignments change often. A teacher who carried six schools last
year may carry three different ones this year, and the person taking over
Sherwood Forest starts with a room full of instruments and no idea which
ones are out, which are broken, or what has already been repaired twice.

Three exports, all for one school:

  * the handoff       - inventory, checkout history and repair history in one
                        workbook, so a successor gets the whole picture in a
                        single file
  * needs repair now  - what to hand the technician
  * repair history    - everything ever done at that school

No tkinter here; this is data and openpyxl only, so it can be tested and
called from anywhere.
"""

from datetime import date


# The handoff carries the instrument columns worth passing on.  Deliberately
# not every column: locker combinations and internal notes are the outgoing
# teacher's business, and a purchase price from 1994 helps nobody.
INSTRUMENT_FIELDS = [
    ("category", "Category"),
    ("description", "Instrument"),
    ("size", "Size"),
    ("brand", "Brand"),
    ("model", "Model"),
    ("serial_no", "Serial #"),
    ("barcode", "Barcode"),
    ("district_no", "District #"),
    ("condition", "Condition"),
    ("year_manufactured", "Year Made"),
    ("est_value", "Est. Value"),
    ("accessories", "Accessories"),
    ("comments", "Notes"),
]

CHECKOUT_FIELDS = [
    ("instrument_desc", "Instrument"),
    ("size", "Size"),
    ("serial_no", "Serial #"),
    ("barcode", "Barcode"),
    ("student", "Student"),
    ("grade", "Grade"),
    ("date_assigned", "Out"),
    ("date_returned", "Returned"),
    ("due_date", "Due"),
    ("notes", "Notes"),
]

REPAIR_FIELDS = [
    ("instrument_desc", "Instrument"),
    ("serial_no", "Serial #"),
    ("barcode", "Barcode"),
    ("description", "Repair"),
    ("date_added", "Reported"),
    ("date_repaired", "Repaired"),
    ("assigned_to", "Sent To"),
    ("location", "Location"),
    ("est_cost", "Est. Cost"),
    ("act_cost", "Actual Cost"),
    ("invoice_number", "Invoice #"),
    ("priority", "Priority"),
    ("notes", "Notes"),
]


def _val(row, key):
    """One cell's value, with the joined-name convenience the rows don't carry."""
    if key == "student":
        name = f"{row.get('first_name') or ''} {row.get('last_name') or ''}".strip()
        return name or (row.get("student_name") or "")
    v = row.get(key)
    return "" if v is None else v


def _sheet(wb, title, fields, rows, first=False):
    from openpyxl.styles import Font, Alignment

    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]

    for c, (_key, head) in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    for r, row in enumerate(rows, start=2):
        for c, (key, _head) in enumerate(fields, start=1):
            ws.cell(row=r, column=c, value=_val(row, key))

    # Size each column to what is actually in it.  A fixed width is right at
    # exactly one font size and wrong at the rest, and a serial number cut off
    # halfway is worse than no serial number at all.
    for c, (key, head) in enumerate(fields, start=1):
        longest = max([len(str(head))]
                      + [len(str(_val(row, key))) for row in rows] or [0])
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = \
            min(max(longest + 2, 9), 46)
    ws.freeze_panes = "A2"
    return ws


def _stamp(ws, site_name, subtitle):
    """A line at the top saying which school and when, because these files get
    emailed around and renamed."""
    from openpyxl.styles import Font
    ws.insert_rows(1, 2)
    ws.cell(row=1, column=1,
            value=f"{site_name} — {subtitle}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1,
            value=f"Exported {date.today():%d %B %Y}").font = Font(italic=True)
    ws.freeze_panes = "A4"


def _site_name(site):
    return (dict(site).get("name") if site else "") or "School"


def export_handoff(db, site, out_path):
    """Everything the next teacher needs for this school, in one workbook."""
    from openpyxl import Workbook

    site = dict(site)
    sid = site["id"]
    instruments = [dict(r) for r in db.get_instruments_with_status(
        include_inactive=True, site_id=sid)]
    checkouts = [dict(r) for r in db.get_checkouts_for_site(sid)]
    repairs = [dict(r) for r in db.get_all_repairs(site_id=sid)]

    wb = Workbook()
    ws = _sheet(wb, "Instruments", INSTRUMENT_FIELDS, instruments, first=True)
    _stamp(ws, site["name"], "instrument inventory")
    _sheet(wb, "Checkout History", CHECKOUT_FIELDS, checkouts)
    _sheet(wb, "Repair History", REPAIR_FIELDS, repairs)
    wb.save(out_path)
    return {"path": out_path, "instrument": len(instruments),
            "checkout": len(checkouts), "repair": len(repairs)}


def export_needs_repair(db, site, out_path):
    """What is broken right now — the list that goes to the technician."""
    from openpyxl import Workbook

    site = dict(site)
    rows = [dict(r) for r in db.get_pending_repairs(site_id=site["id"])]
    wb = Workbook()
    ws = _sheet(wb, "Needs Repair", REPAIR_FIELDS, rows, first=True)
    _stamp(ws, site["name"], "instruments awaiting repair")
    wb.save(out_path)
    return {"path": out_path, "repair": len(rows)}


def export_repair_history(db, site, out_path):
    """Everything ever done to this school's instruments."""
    from openpyxl import Workbook

    site = dict(site)
    rows = [dict(r) for r in db.get_all_repairs(site_id=site["id"])]
    # What was actually spent, not what was quoted -- an estimate that never
    # became an invoice is not money out of anybody's budget.
    total = 0.0
    for r in rows:
        try:
            total += float(r.get("act_cost") or 0)
        except (TypeError, ValueError):
            pass
    wb = Workbook()
    ws = _sheet(wb, "Repair History", REPAIR_FIELDS, rows, first=True)
    _stamp(ws, site["name"], "full repair history")
    ws.cell(row=len(rows) + 5, column=1, value="Total spent")
    ws.cell(row=len(rows) + 5, column=2, value=round(total, 2))
    wb.save(out_path)
    return {"path": out_path, "repair": len(rows), "total": round(total, 2)}


def suggested_filename(site, kind: str) -> str:
    """A filename that still means something in somebody else's downloads."""
    name = "".join(ch for ch in _site_name(site) if ch.isalnum() or ch in " -_").strip()
    name = name.replace(" ", "_") or "School"
    return f"{name}_{kind}_{date.today():%Y%m%d}.xlsx"


# ── Reading one back in ──────────────────────────────────────────────────────
# The export is only half a handoff. Teacher A leaving Clyde Hill is no use
# unless Teacher B can pick the file up in September and carry on, so the same
# workbook reads back in.

def _header_map(ws, fields):
    """{column index: field name} from the header row of an exported sheet.

    Matched on the printed heading rather than position, so a file somebody has
    reordered or added a column to still reads.
    """
    by_head = {head.lower(): key for key, head in fields}
    for r in range(1, min(ws.max_row, 12) + 1):
        found = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            key = by_head.get(str(v).strip().lower())
            if key:
                found[c] = key
        if len(found) >= 3:          # a real header row, not the title stamp
            return r, found
    return None, {}


def _read_sheet(ws, fields):
    header_row, cols = _header_map(ws, fields)
    if not header_row:
        return []
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for c, key in cols.items():
            v = ws.cell(row=r, column=c).value
            row[key] = "" if v is None else v
        if any(str(v).strip() for v in row.values()):
            out.append(row)
    return out


def read_handoff(path):
    """{"instruments": [...], "repairs": [...]} from a handoff workbook.

    Checkout history is deliberately not read back. Those rows name children
    who have moved on to middle school by the time anybody imports this; they
    are in the file so the incoming teacher can see which horn has been through
    four players, not so they can be recreated as live loans.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    inst = _read_sheet(wb["Instruments"], INSTRUMENT_FIELDS)         if "Instruments" in wb.sheetnames else []
    reps = _read_sheet(wb["Repair History"], REPAIR_FIELDS)         if "Repair History" in wb.sheetnames else []
    return {"instruments": inst, "repairs": reps}


def _fingerprints(row):
    """The ways one instrument can be recognized as one already here."""
    out = []
    for key in ("serial_no", "barcode", "district_no"):
        v = str(row.get(key) or "").strip().lower()
        if v:
            out.append((key, v))
    return out


def import_handoff(db, site, path):
    """Load a handoff workbook into one school.

    Runs twice safely. An instrument already here -- matched on serial number,
    barcode or district number -- is left alone rather than added again, since
    a teacher who imports the same file twice should end up with one
    inventory, not two.
    """
    site = dict(site)
    sid = site["id"]
    data = read_handoff(path)

    existing = [dict(r) for r in db.get_instruments_with_status(
        include_inactive=True, site_id=sid)]
    seen = {}
    for row in existing:
        for fp in _fingerprints(row):
            seen[fp] = row["id"]

    added, matched, no_id = 0, 0, 0
    id_for_row = {}
    for row in data["instruments"]:
        fps = _fingerprints(row)
        hit = next((seen[f] for f in fps if f in seen), None)
        if hit:
            matched += 1
            id_for_row[id(row)] = hit
            continue
        payload = {k: row.get(k) for k, _h in INSTRUMENT_FIELDS}
        payload["site_id"] = sid
        new_id = db.add_instrument(payload)
        id_for_row[id(row)] = new_id
        added += 1
        if not fps:
            # Nothing to recognize it by, so a second import would add it
            # again.  Worth telling the teacher rather than silently risking it.
            no_id += 1
        for f in fps:
            seen[f] = new_id

    repairs = 0
    if data["repairs"]:
        repairs = _import_repairs(db, sid, data["repairs"], seen)

    return {"added": added, "matched": matched, "repairs": repairs,
            "unidentifiable": no_id,
            "checkouts_skipped": True}


def _import_repairs(db, site_id, rows, seen):
    """Repair history, attached to whichever instrument it names.

    A repair whose instrument is not here is dropped: a repair record floating
    free of the thing repaired is noise in the history, and the export always
    carries the instruments alongside.
    """
    n = 0
    with db._connect() as conn:
        for row in rows:
            hit = next((seen[f] for f in _fingerprints(row) if f in seen), None)
            if not hit:
                continue
            dup = conn.execute(
                "SELECT 1 FROM repairs WHERE instrument_id = ? "
                "AND IFNULL(description,'') = ? AND IFNULL(date_added,'') = ?",
                (hit, str(row.get("description") or ""),
                 str(row.get("date_added") or ""))).fetchone()
            if dup:
                continue
            conn.execute(
                "INSERT INTO repairs (instrument_id, description, date_added, "
                "date_repaired, assigned_to, location, est_cost, act_cost, "
                "invoice_number, notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (hit, row.get("description"), row.get("date_added"),
                 row.get("date_repaired"), row.get("assigned_to"),
                 row.get("location"), _num(row.get("est_cost")),
                 _num(row.get("act_cost")), row.get("invoice_number"),
                 row.get("notes")))
            n += 1
        conn.commit()
    return n


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def export_all_needs_repair(db, sites, out_path):
    """Every school's outstanding repairs, in one file.

    This one goes to the district's instrument coordinator -- one person,
    looking after every school in the district -- so a single list they can
    work down beats a dozen files they have to open in turn.  The school is a
    column, since without it the list is unusable to them.
    """
    from openpyxl import Workbook

    fields = [("school", "School")] + REPAIR_FIELDS
    rows = []
    for site in sites:
        site = dict(site)
        for r in db.get_pending_repairs(site_id=site["id"]):
            row = dict(r)
            row["school"] = site["name"]
            rows.append(row)

    wb = Workbook()
    ws = _sheet(wb, "Needs Repair", fields, rows, first=True)
    _stamp(ws, "All schools", "instruments awaiting repair")
    wb.save(out_path)
    return {"path": out_path, "repair": len(rows), "schools": len(list(sites))}


def export_year_end_pack(db, sites, out_dir):
    """One folder holding every school's year-end paperwork.

    An inventory file PER SCHOOL, because each one is handed to whoever takes
    that school on next, and a file covering six schools is no use to the
    person taking one of them.  Repairs the other way round: one combined list,
    because that goes to the district coordinator rather than to six different
    successors.
    """
    import os

    sites = [dict(x) for x in sites]
    os.makedirs(out_dir, exist_ok=True)
    done, failed = [], []
    for site in sites:
        path = os.path.join(out_dir, suggested_filename(site, "inventory"))
        try:
            res = export_handoff(db, site, path)
            done.append((site["name"], "inventory", path, res))
        except Exception as e:
            failed.append((site["name"], "inventory", str(e)))

    combined = os.path.join(
        out_dir, f"All_Schools_needs_repair_{date.today():%Y%m%d}.xlsx")
    try:
        res = export_all_needs_repair(db, sites, combined)
        done.append(("All schools", "needs_repair", combined, res))
    except Exception as e:
        failed.append(("All schools", "needs_repair", str(e)))

    return {"folder": out_dir, "written": done, "failed": failed,
            "schools": len(sites)}
