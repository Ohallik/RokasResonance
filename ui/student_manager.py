"""
ui/student_manager.py - Student management by school year
"""

import csv
import os
import re
import tkinter as tk
from tkinter import filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from datetime import datetime
from ui.theme import muted_fg, subtle_fg, fs, bind_copy_menu
from ui.names import display_last_first, display_full, display_first_of


def _current_school_year() -> str:
    today = datetime.today()
    if today.month >= 8:
        return f"{today.year}-{today.year + 1}"
    return f"{today.year - 1}-{today.year}"


def _next_school_year() -> str:
    today = datetime.today()
    if today.month >= 8:
        return f"{today.year + 1}-{today.year + 2}"
    return f"{today.year}-{today.year + 1}"


def _year_start(school_year: str) -> int:
    """Starting calendar year of a '2026-2027' label (-1 if unparseable)."""
    try:
        return int(str(school_year).split("-")[0])
    except (ValueError, AttributeError, IndexError):
        return -1


GRADE_OPTIONS = ["5", "6", "7", "8", "9", "10", "11", "12", "Other"]
GENDER_OPTIONS = ["", "Male", "Female", "Non-binary", "Prefer not to say"]
RELATION_OPTIONS = ["", "Parent", "Guardian", "Grandparent", "Step-parent", "Other"]

# ── Ensembles / class periods / instruments (middle-school level) ──────────────
# Ensemble / instrument vocabulary lives in ui/ensembles.py so the student
# manager, performance dialog, and program importer all share one source.
from ui.ensembles import (
    BAND_ENSEMBLES, ORCHESTRA_ENSEMBLES, CHOIR_ENSEMBLES, PERIOD_OPTIONS,
    BAND_INSTRUMENTS, ORCHESTRA_INSTRUMENTS, CHOIR_PARTS,
    ensembles_for, instruments_for, instrument_sort_key, selectable_ensembles,
    fifth_grade_instruments,
)


def _csv_to_list(val: str):
    return [p.strip() for p in (val or "").split(",") if p.strip()]


# Named-header layout for the middle-school → high-school transfer CSV.  Because
# it uses real headers (not positional columns) the same file both exports from
# one program and re-imports into another to update instrumentation.
HS_TRANSFER_FIELDS = [
    ("Student ID", "student_id"),
    ("Last Name", "last_name"),
    ("First Name", "first_name"),
    ("Grade", "grade"),
    ("Primary Instrument", "primary_instrument"),
    ("Secondary Instrument", "secondary_instrument"),
    ("Ensembles", "ensembles"),
    ("Student Email", "student_email"),
    ("Phone", "phone"),
    ("Parent 1 Name", "parent1_name"),
    ("Parent 1 Email", "parent1_email"),
    ("Parent 1 Phone", "parent1_phone"),
    ("Parent 2 Name", "parent2_name"),
    ("Parent 2 Email", "parent2_email"),
    ("Parent 2 Phone", "parent2_phone"),
    ("Notes", "notes"),
]


class _ClassOptionsMixin:
    """Supplies ``_class_options()`` — the class names a picker should offer.

    Always the teacher's configured classes PLUS whatever classes the roster
    actually uses.  Anything that filters, assigns or counts students has to
    work off names that exist in the data; offering only the configured names
    is how a filter ends up matching nothing and a Save ends up erasing a
    student's class.  See ui.ensembles.selectable_ensembles.
    """

    def _class_year(self):
        for attr in ("_year_var", "school_year", "default_year"):
            val = getattr(self, attr, None)
            if val is None:
                continue
            val = val.get() if hasattr(val, "get") else val
            if val:
                return val
        return None

    # Dialogs that ASSIGN a class (edit a student, bulk assign, import) set this
    # so configured-but-empty classes stay on offer — you have to be able to put
    # the first student into a new class.  Filters leave it False.
    _class_options_include_empty = False

    def _class_options(self):
        """The classes a picker should offer.

        Scoped to the school when there is one.  Every screen that offers a
        class list goes through here, which is why an elementary roster was
        being offered "Advanced" in three different places at once: the list
        was the configured SECONDARY one, and nothing ever said otherwise.
        """
        site_id = getattr(self, "site_id", None)
        try:
            return selectable_ensembles(
                self.db, self._class_year(),
                getattr(self, "program_type", "band"),
                include_empty=self._class_options_include_empty,
                site_id=site_id)
        except Exception:
            if site_id:
                return []
            return list(ensembles_for(getattr(self, "program_type", "band")))

    def _class_display_options(self):
        """Short display labels for FILTER widgets only ("Entry", "Jazz 1").

        Safe there because every filter matches by class identity, so the
        short label finds the class no matter how the records spell it.  Never
        use these for a widget that WRITES the value to a record — assignment
        widgets store the full canonical name."""
        from ui.ensembles import class_display_map
        opts = self._class_options()
        dmap = class_display_map(opts)
        return [dmap[o] for o in opts]


class StudentManager(_ClassOptionsMixin, ttk.Frame):
    def __init__(self, parent, db, program_type: str = "band", site_id=None):
        super().__init__(parent)
        self.db = db
        self.program_type = program_type or "band"
        # Set: this is one school's roster.  Unset: the secondary program.
        self.site_id = site_id
        self._year_var = tk.StringVar()
        self._search_var = tk.StringVar()
        self._show_inactive_var = tk.BooleanVar(value=False)
        self._filter_ensemble_var = tk.StringVar(value="All")
        self._filter_period_var = tk.StringVar(value="All")
        self._selected_id = None
        self._checked = set()   # student ids ticked for bulk actions
        self._repair_prompted = False
        self._dup_prompted = False

        self._build()
        self._populate_year_options()
        self.refresh()
        # A roster of ID-number "students" means a class-list import went
        # wrong (its name column was really the Student ID column).  Offer
        # the one-click repair as soon as the window is up.
        self.after(400, self._offer_botched_repair)
        self.after(900, self._offer_duplicate_merge)

    def _build(self):
        # ── Toolbar ───────────────────────────────────────────────────────────
        toolbar = ttk.Frame(self, bootstyle=LIGHT)
        toolbar.pack(fill=X)

        from ui.help_system import add_help_button
        add_help_button(toolbar, "students", dark=False, pady=6)

        ttk.Button(toolbar, text="➕ Add Student", bootstyle=SUCCESS,
                   command=self._add_student).pack(side=LEFT, padx=6, pady=6)
        ttk.Button(toolbar, text="✏️ Edit", bootstyle=PRIMARY,
                   command=self._edit_student).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="🗑️ Delete", bootstyle=DANGER,
                   command=self._delete_student).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="🗂️ Inactive Students", bootstyle=(SECONDARY, OUTLINE),
                   command=self._show_inactive_window).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="📥 Import ▾", bootstyle=INFO,
                   command=self._open_import_menu).pack(side=LEFT, padx=6, pady=6)

        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)

        ttk.Button(toolbar, text="🏷️ Assign", bootstyle=SUCCESS,
                   command=self._bulk_assign).pack(side=LEFT, padx=6, pady=6)
        ttk.Button(toolbar, text="✉️ Email List", bootstyle=INFO,
                   command=self._email_list).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="📤 Export ▾", bootstyle=INFO,
                   command=self._open_export_menu).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="🔄 Refresh", bootstyle=(SECONDARY, OUTLINE),
                   command=self.refresh).pack(side=LEFT, padx=6, pady=6)

        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)

        ttk.Label(toolbar, text="School Year:").pack(side=LEFT, padx=(0, 4))
        years_combo = ttk.Combobox(toolbar, textvariable=self._year_var,
                                    state="readonly", width=14)
        years_combo.pack(side=LEFT, padx=(0, 10))
        years_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh())
        self._years_combo = years_combo

        # ── Filter bar (below toolbar) ─────────────────────────────────────────
        filter_bar = ttk.Frame(self)
        filter_bar.pack(fill=X, padx=6, pady=(2, 0))

        ttk.Label(filter_bar, text="Search:").pack(side=LEFT, padx=(0, 4))
        ttk.Entry(filter_bar, textvariable=self._search_var, width=26).pack(side=LEFT)
        self._search_var.trace_add("write", lambda *_: self._apply_filter())

        # Neither filter belongs on a 5th grade roster.  The ensemble list is
        # the configured SECONDARY classes, so it was offering to filter an
        # elementary school by "Advanced"; and elementary has no class periods
        # at all -- the group is a pullout on the specialist rotation.  Two
        # sections and a search box is the whole of what is needed here.
        if self.site_id:
            # Section, not "Ensemble" -- and never the secondary class list that
            # was offering to filter an elementary school by "Advanced".  Both
            # sections are always here, even before the roster lands; choir
            # appears once a single child is in it.
            ttk.Label(filter_bar, text="Section:").pack(side=LEFT, padx=(12, 4))
            # The box shows "Section 1"; the record says "Clyde Hill Elementary
            # School: Section 1".  Keep the map, or the filter matches nothing
            # and the roster silently empties.
            self._section_full = {_short_group(g): g for g in self._roster_groups()}
            self._section_filter_combo = ttk.Combobox(
                filter_bar, textvariable=self._filter_ensemble_var,
                state="readonly", width=14,
                values=["All"] + list(self._section_full),
            )
            self._section_filter_combo.pack(side=LEFT)
            self._filter_ensemble_var.trace_add(
                "write", lambda *_: self._apply_filter())
        else:
            ttk.Label(filter_bar, text="Ensemble:").pack(side=LEFT, padx=(12, 4))
            self._ensemble_filter_combo = ttk.Combobox(
                filter_bar, textvariable=self._filter_ensemble_var,
                state="readonly", width=18,
                values=["All"] + self._class_display_options(),
            )
            self._ensemble_filter_combo.pack(side=LEFT)
            self._filter_ensemble_var.trace_add("write", lambda *_: self._apply_filter())

            ttk.Label(filter_bar, text="Period:").pack(side=LEFT, padx=(8, 4))
            ttk.Combobox(
                filter_bar, textvariable=self._filter_period_var,
                state="readonly", width=5, values=["All"] + PERIOD_OPTIONS,
            ).pack(side=LEFT)
            self._filter_period_var.trace_add("write", lambda *_: self._apply_filter())

        ttk.Checkbutton(
            filter_bar,
            text="Show Inactive",
            variable=self._show_inactive_var,
            bootstyle=SECONDARY,
            command=self.refresh,
        ).pack(side=LEFT, padx=(12, 0))

        self._count_lbl = ttk.Label(filter_bar, text="", foreground=muted_fg())
        self._count_lbl.pack(side=RIGHT, padx=6)

        # ── Selection bar (tick boxes for bulk actions) ─────────────────────────
        sel_bar = ttk.Frame(self)
        sel_bar.pack(fill=X, padx=6, pady=(2, 0))
        ttk.Label(sel_bar, text="☑ Tick students to bulk-assign:",
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(side=LEFT, padx=(2, 6))
        ttk.Button(sel_bar, text="Select All Shown", bootstyle=(SECONDARY, OUTLINE),
                   command=self._check_all_shown).pack(side=LEFT, padx=2)
        ttk.Button(sel_bar, text="Clear Selection", bootstyle=(SECONDARY, OUTLINE),
                   command=self._clear_checks).pack(side=LEFT, padx=2)
        self._sel_count_lbl = ttk.Label(sel_bar, text="0 selected",
                                        font=("Segoe UI", 8, "bold"))
        self._sel_count_lbl.pack(side=LEFT, padx=10)

        # ── Content: Tree + Detail ─────────────────────────────────────────────
        paned = ttk.Panedwindow(self, orient=HORIZONTAL)
        paned.pack(fill=BOTH, expand=True, padx=6, pady=6)

        # Left: student list
        list_frame = ttk.Frame(paned)
        paned.add(list_frame, weight=2)

        # An elementary roster has no class periods, and its one grouping is the
        # section -- so it says Class Section rather than Ensembles, and the
        # Period column is not there to be sorted by something that cannot vary.
        if self.site_id:
            cols = ("check", "Name", "Grade", "Class Section", "Instrument",
                    "Parent", "Active Instruments")
            widths = [34, 190, 60, 150, 130, 160, 90]
            _stretch = {"Name", "Class Section", "Parent"}
        else:
            cols = ("check", "Name", "Grade", "Ensembles", "Period", "Instrument",
                    "Parent", "Active Instruments")
            widths = [34, 170, 60, 150, 70, 110, 150, 90]
            _stretch = {"Name", "Ensembles", "Parent"}
        # Scaled, because a width that fits "Grade" at Normal shows "Grad" at
        # Large -- the font grows and a raw pixel column does not.
        from ui.theme import px
        widths = [px(w) for w in widths]

        sb = ttk.Scrollbar(list_frame, orient=VERTICAL)
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings",
                                  yscrollcommand=sb.set, selectmode="extended",
                                  bootstyle=PRIMARY)
        sb.config(command=self.tree.yview)
        sb.pack(side=RIGHT, fill=Y)
        self.tree.pack(fill=BOTH, expand=True)

        for col, w in zip(cols, widths):
            heading = "" if col == "check" else col
            self.tree.heading(col, text=heading, anchor=W,
                              command=(self._toggle_all_header if col == "check"
                                       else (lambda c=col: self._sort_by(c))))
            self.tree.column(col, width=w, anchor=(CENTER if col == "check" else W),
                             stretch=col in _stretch,
                             minwidth=(px(34) if col == "check" else px(50)))

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Button-1>", self._on_tree_click, add="+")
        self.tree.bind("<Double-1>", lambda e: self._edit_student())

        # Right: detail panel
        detail_frame = ttk.Frame(paned, width=300)
        paned.add(detail_frame, weight=1)
        self._build_detail(detail_frame)

        self._sort_col = "Name"
        self._sort_asc = True
        self._all_students = []

    def _build_detail(self, parent):
        nb = ttk.Notebook(parent, bootstyle=PRIMARY)
        nb.pack(fill=BOTH, expand=True)

        info_tab = ttk.Frame(nb)
        checkout_tab = ttk.Frame(nb)
        nb.add(info_tab, text="Student Info")
        nb.add(checkout_tab, text="Instrument History")

        # ── Info tab ──────────────────────────────────────────────────────────
        outer = ttk.Frame(info_tab)
        outer.pack(fill=BOTH, expand=True, padx=8, pady=8)

        self._detail_labels = {}
        fields = [
            ("Name", "_full_name"),
            ("Grade", "grade"),
            ("School Year", "school_year"),
            ("Ensembles", "ensembles"),
            ("Class Periods", "class_periods"),
            ("Primary Instr.", "primary_instrument"),
            ("Secondary Instr.", "secondary_instrument"),
            ("Gender", "gender"),
            ("Student ID", "student_id"),
            ("Phone", "phone"),
            ("Email", "student_email"),
            ("Address", "_address"),
            ("Parent 1", "parent1_name"),
            ("P1 Phone", "parent1_phone"),
            ("P1 Email", "parent1_email"),
            ("Parent 2", "parent2_name"),
            ("P2 Phone", "parent2_phone"),
            ("P2 Email", "parent2_email"),
            ("Notes", "notes"),
        ]
        if self.site_id:
            # A 5th grader has no class period and no second instrument, so
            # those two rows would sit blank on every child for the whole year.
            fields = [f for f in fields
                      if f[1] not in ("class_periods", "secondary_instrument")]
        for label, key in fields:
            r = ttk.Frame(outer)
            r.pack(fill=X, pady=1)
            ttk.Label(r, text=f"{label}:", font=("Segoe UI", fs(8), "bold"),
                      width=12, anchor=W).pack(side=LEFT)
            lbl = ttk.Label(r, text="", font=("Segoe UI", fs(8)),
                             anchor=W, wraplength=180, justify=LEFT)
            lbl.pack(side=LEFT, fill=X, expand=True)
            bind_copy_menu(lbl)
            self._detail_labels[key] = lbl

        # ── Checkout history tab ───────────────────────────────────────────────
        hist_frame = ttk.Frame(checkout_tab)
        hist_frame.pack(fill=BOTH, expand=True)

        cols_h = ("Instrument", "Date Out", "Date In")
        sb_h = ttk.Scrollbar(hist_frame, orient=VERTICAL)
        self._hist_tree = ttk.Treeview(hist_frame, columns=cols_h, show="headings",
                                        yscrollcommand=sb_h.set, bootstyle=INFO)
        sb_h.config(command=self._hist_tree.yview)
        sb_h.pack(side=RIGHT, fill=Y)
        self._hist_tree.pack(fill=BOTH, expand=True)

        _stretch_h = {"Instrument"}
        for col in cols_h:
            self._hist_tree.heading(col, text=col, anchor=W)
            self._hist_tree.column(col, width=100, anchor=W,
                                   minwidth=40, stretch=col in _stretch_h)

    # ─────────────────────────────────────────────────────────── Data Loading ─

    def _populate_year_options(self):
        """Only years that have started.  Previous years stay available for
        reference, but a teacher must not be able to select next year and start
        filing this year's students under it — rolling forward is the New School
        Year wizard's job, and it moves everything together."""
        cur = _current_school_year()
        cur_start = _year_start(cur)
        years = [y for y in self.db.get_school_years() if _year_start(y) <= cur_start]
        if cur not in years:
            years.insert(0, cur)
        years = sorted(set(years), key=_year_start, reverse=True)
        self._years_combo["values"] = years
        # Default to current school year (not necessarily years[0])
        if not self._year_var.get() or self._year_var.get() not in years:
            self._year_var.set(cur)

    def _offer_botched_repair(self):
        """Detect and offer to undo a botched class-list import (students whose
        'names' are ID numbers).  Runs once per window."""
        if self._repair_prompted:
            return
        try:
            bad = self.db.find_botched_import_students()
        except Exception:
            return
        if len(bad) < 2:
            return                       # one odd record isn't a botched import
        self._repair_prompted = True
        if Messagebox.yesno(
                f"{len(bad)} student records look like a failed class-list "
                "import — their names are ID numbers and they have no other "
                "information.\n\n"
                "Fix this now? The ID-number records are removed and, if the "
                "import had archived your real roster, it is restored exactly "
                "as it was.",
                title="Fix broken import?",
                parent=self.winfo_toplevel()) != "Yes":
            return
        deleted, restored_year, restored = self.db.undo_botched_import()
        msg = f"Removed {deleted} ID-number record(s)."
        if restored_year:
            msg += (f"\nRestored your {restored_year} roster "
                    f"({restored} students).")
            self._year_var.set(restored_year)
        Messagebox.show_info(msg, title="Import repaired",
                             parent=self.winfo_toplevel())
        self._populate_year_options()
        self.refresh()

    def _offer_duplicate_merge(self):
        """Offer to fold duplicate student records together.  Older class-list
        imports made a record per roster row, so a student listed in two
        sections became two students."""
        if self._dup_prompted:
            return
        try:
            groups = self.db.find_duplicate_students(self._year_var.get() or None)
        except Exception:
            return
        if not groups:
            return
        self._dup_prompted = True
        extra = sum(len(g) - 1 for g in groups)
        names = [f"  • {(g[0]['first_name'] or '').strip()} "
                 f"{(g[0]['last_name'] or '').strip()}  (×{len(g)})"
                 for g in groups]
        shown = "\n".join(names[:15])
        if len(names) > 15:
            shown += f"\n  … and {len(names) - 15} more"
        if Messagebox.yesno(
                f"{len(groups)} student(s) appear more than once, for a total "
                f"of {extra} duplicate record(s):\n\n{shown}\n\n"
                "Merge them? Each becomes one record that keeps their contact "
                "information and all of their classes; instruments, fees and "
                "check-outs move with them.\n\n"
                "This cannot be undone, so back up first if any of these names "
                "are actually two different students.",
                title="Merge duplicate students?",
                parent=self.winfo_toplevel()) != "Yes":
            return
        merged, removed = self.db.merge_duplicate_students(
            self._year_var.get() or None)
        Messagebox.show_info(
            f"Merged {merged} student(s) and removed {removed} duplicate "
            "record(s).", title="Duplicates merged",
            parent=self.winfo_toplevel())
        self.refresh()

    def refresh(self):
        year = self._year_var.get() or None
        include_inactive = self._show_inactive_var.get()
        # 5th graders are managed in the 5th Grade window, not here -- keeping
        # them out of this roster is what keeps them off secondary mailings.
        self._all_students = list(self.db.get_all_students(
            school_year=year, include_inactive=include_inactive,
            site_id=self.site_id,
            level=None if self.site_id else "secondary"))
        # Keep only ticks for students still in the loaded roster
        loaded_ids = {s["id"] for s in self._all_students}
        self._checked &= loaded_ids
        self._apply_filter()
        self._populate_year_options()

    # ─────────────────────────────────────────────── Tick-box multi-select ─────

    def _on_tree_click(self, event):
        """Toggle a row's tick box when the checkbox column is clicked."""
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        if self.tree.identify_column(event.x) != "#1":
            return
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        sid = int(iid)
        if sid in self._checked:
            self._checked.discard(sid)
        else:
            self._checked.add(sid)
        self.tree.set(iid, "check", "☑" if sid in self._checked else "☐")
        self._update_sel_count()

    def _check_all_shown(self):
        for iid in self.tree.get_children():
            sid = int(iid)
            self._checked.add(sid)
            self.tree.set(iid, "check", "☑")
        self._update_sel_count()

    def _clear_checks(self):
        self._checked.clear()
        for iid in self.tree.get_children():
            self.tree.set(iid, "check", "☐")
        self._update_sel_count()

    def _toggle_all_header(self):
        """Clicking the checkbox column header toggles all shown rows."""
        shown = [int(iid) for iid in self.tree.get_children()]
        if shown and all(sid in self._checked for sid in shown):
            self._clear_checks()
        else:
            self._check_all_shown()

    def _update_sel_count(self):
        if hasattr(self, "_sel_count_lbl"):
            self._sel_count_lbl.config(text=f"{len(self._checked)} selected")

    def _apply_filter(self):
        search = self._search_var.get().lower()
        ens_filter = self._filter_ensemble_var.get()
        per_filter = self._filter_period_var.get()
        visible = []
        for s in self._all_students:
            if search:
                haystack = " ".join([
                    str(s["first_name"] or ""),
                    str(s["last_name"] or ""),
                    str(s["phone"] or ""),
                    str(s["parent1_name"] or ""),
                    str(s["student_id"] or ""),
                    str(self._sval(s, "primary_instrument")),
                    str(self._sval(s, "secondary_instrument")),
                ]).lower()
                if search not in haystack:
                    continue
            if ens_filter and ens_filter != "All":
                held = self._sval(s, "ensembles")
                if self.site_id:
                    # Elementary groups are exact strings, not class identities
                    # -- "Section 1" is not a level the class parser knows.
                    want = self._section_full.get(ens_filter, ens_filter).lower()
                    if not any(p.strip().lower() == want
                               for p in (held or "").split(",")):
                        continue
                else:
                    # Identity match ("Entry" ≡ "Entry Band" ≡ "MS Band (Entry)")
                    import class_registry as cr
                    if not cr.csv_has_class(held, ens_filter):
                        continue
            if per_filter and per_filter != "All":
                if per_filter not in _csv_to_list(self._sval(s, "class_periods")):
                    continue
            visible.append(s)

        self._populate_tree(visible)
        self._count_lbl.config(text=f"{len(visible)} student(s)")

    @staticmethod
    def _sval(row, key):
        """Safe column access for rows that may predate a migration."""
        try:
            return row[key] if key in row.keys() else ""
        except Exception:
            return ""

    def _populate_tree(self, students):
        col_key_map = {
            "Name": "_sort_name",
            "Grade": "_sort_grade",
            "Ensembles": "ensembles",
            "Period": "class_periods",
            "Instrument": "_sort_instrument",
            "Parent": "parent1_name",
            "Active Instruments": "_active_count",
        }
        key = col_key_map.get(self._sort_col, "_sort_name")

        # Get active checkout counts
        active_checkouts = {}
        for s in students:
            with self.db._connect() as conn:
                count = conn.execute(
                    "SELECT COUNT(*) FROM checkouts WHERE student_id=? AND date_returned IS NULL",
                    (s["id"],)
                ).fetchone()[0]
            active_checkouts[s["id"]] = count

        def sort_key(s):
            """Always returns a comparable tuple.  Sorting has to stay total:
            a blank instrument/parent/grade must not make the comparison mix
            numbers with strings (which used to raise TypeError mid-sort and
            leave the list unsorted)."""
            if key == "_active_count":
                return (0, "", active_checkouts.get(s["id"], 0))
            if key == "_sort_name":
                return (0, f"{(s['last_name'] or '').lower()}"
                           f" {(s['first_name'] or '').lower()}", 0)
            if key == "_sort_instrument":
                # Score order (flute → tuba → percussion), not alphabetical:
                # that's how a director reads a roster.  Blanks sort last.
                inst = self._sval(s, "primary_instrument") or ""
                rank, name = instrument_sort_key(inst)
                return (rank, name, 0)
            if key == "_sort_grade":
                raw = str(self._sval(s, "grade") or "").strip()
                # Numeric grades sort numerically; "Other"/blank sort after.
                return ((int(raw), "", 0) if raw.isdigit()
                        else (9999, raw.lower(), 0))
            val = self._sval(s, key)
            if val is None or val == "":
                return (1, "", 0)              # blanks last, ascending
            return (0, str(val).lower(), 0)

        students = sorted(students, key=sort_key, reverse=not self._sort_asc)
        self._update_sort_indicator()

        # Configure inactive tag with current theme color
        self.tree.tag_configure("inactive", foreground=subtle_fg())
        self.tree.tag_configure("provisional", foreground=subtle_fg())

        self.tree.delete(*self.tree.get_children())
        for s in students:
            full_name = display_last_first(s)
            active = active_checkouts.get(s["id"], 0)
            active_str = f"✓ {active}" if active > 0 else ""
            iid = str(s["id"])
            is_inactive = not s["is_active"]
            # "Incoming" students (pre-loaded from a feeder handoff, not yet on the
            # official roster) are grayed and tagged, but stay fully contactable.
            is_prov = bool(s.get("provisional")) and not is_inactive
            if is_inactive:
                tags = ("inactive",)
                name_display = f"{full_name} (inactive)"
            elif is_prov:
                tags = ("provisional",)
                name_display = f"{full_name}  — incoming"
            else:
                tags = ()
                name_display = full_name
            # Ensembles show their short display form ("Entry, Jazz 2") — the
            # full canonical names stay on the record and in the detail pane.
            import class_registry as cr
            if self.site_id:
                # "Section 2", not "Jing Mei Elementary School: Section 2".  The
                # window is one school; repeating its name in every cell of
                # every row pushes the part that differs off the edge.
                ensembles = ", ".join(
                    _short_group(p.strip())
                    for p in (self._sval(s, "ensembles") or "").split(",")
                    if p.strip())
            else:
                ensembles = cr.display_csv(self._sval(s, "ensembles"))
            periods = self._sval(s, "class_periods") or ""
            prim = self._sval(s, "primary_instrument") or ""
            sec = self._sval(s, "secondary_instrument") or ""
            instr_display = prim + (f" / {sec}" if sec else "")
            row = ["☑" if s["id"] in self._checked else "☐",
                   name_display,
                   s["grade"] or "",
                   ensembles]
            if not self.site_id:
                row.append(periods)
            row += [instr_display, s["parent1_name"] or "", active_str]
            self.tree.insert("", "end", iid=iid, tags=tags, values=tuple(row))
        self._update_sel_count()

    def _update_sort_indicator(self):
        """Mark the sorted column with ▲/▼ so it's obvious the header is a
        sort button (and which way it's pointing)."""
        for col in self.tree["columns"]:
            if col == "check":
                continue
            arrow = ""
            if col == self._sort_col:
                arrow = "  ▲" if self._sort_asc else "  ▼"
            self.tree.heading(col, text=f"{col}{arrow}")

    def _sort_by(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._apply_filter()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        self._selected_id = int(sel[0])
        self._load_detail(self._selected_id)

    def _load_detail(self, student_id: int):
        student = self.db.get_student(student_id)
        if not student:
            return

        row = dict(student)
        row["_full_name"] = display_full(student)
        addr_parts = [row.get("address", ""), row.get("city", ""),
                      row.get("state", ""), row.get("zip_code", "")]
        row["_address"] = ", ".join(p for p in addr_parts if p)

        for key, lbl in self._detail_labels.items():
            val = row.get(key, "")
            lbl.config(text=str(val) if val else "")

        # Load checkout history
        self._hist_tree.delete(*self._hist_tree.get_children())
        with self.db._connect() as conn:
            history = conn.execute(
                """SELECT c.date_assigned, c.date_returned, i.description
                   FROM checkouts c
                   JOIN instruments i ON i.id=c.instrument_id
                   WHERE c.student_id=?
                   ORDER BY c.date_assigned DESC""",
                (student_id,)
            ).fetchall()
        for h in history:
            self._hist_tree.insert("", "end", values=(
                h["description"] or "",
                h["date_assigned"] or "",
                h["date_returned"] or "Active",
            ))

    # ─────────────────────────────────────────────────────────── Actions ──────

    def _get_selected(self):
        sel = self.tree.selection()
        if not sel:
            Messagebox.show_warning("Please select a student first.", title="No Selection",
    parent=self.winfo_toplevel())
            return None
        return int(sel[0])

    def _selected_ids(self):
        """Ticked students take priority; otherwise fall back to the highlighted row."""
        if self._checked:
            return list(self._checked)
        return [int(iid) for iid in self.tree.selection()]

    def _bulk_assign(self):
        ids = self._selected_ids()
        if not ids:
            Messagebox.show_warning(
                "Tick the ☐ boxes next to the students you want (or use 'Select All Shown'),\n"
                "then assign an ensemble, class period, and/or instrument to all of them at once.",
                title="No Students Selected",
                parent=self.winfo_toplevel())
            return
        dlg = _BulkAssignDialog(self.winfo_toplevel(), self.db, ids,
                                self.program_type, site_id=self.site_id)
        self.wait_window(dlg)
        # Ticks are cleared once an assignment lands.  Leaving them on invites
        # the next assignment to silently hit the same students again — that is
        # how a whole roster ends up playing one instrument.
        if getattr(dlg, "applied", False):
            self._clear_checks()
        self.refresh()

    def _email_list(self):
        dlg = _EmailListDialog(self.winfo_toplevel(), self.db, self.program_type,
                               self._year_var.get() or None,
                               site_id=self.site_id)
        self.wait_window(dlg)

    def _open_import_menu(self):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="📂  Class roster (district CSV)…",
                         command=self._import_csv)
        menu.add_command(label="🎼  Update instruments from another school's CSV…",
                         command=self._import_instruments_from_csv)
        try:
            menu.tk_popup(self.winfo_pointerx(), self.winfo_pointery())
        finally:
            menu.grab_release()

    def _open_numbers_per_part(self):
        from ui.instrumentation_view import open_instrumentation
        import os
        base_dir = os.path.dirname(os.path.abspath(self.db.db_path))
        open_instrumentation(self, self.db, base_dir,
                             self._year_var.get() or None,
                             site_id=self.site_id)

    def _open_export_menu(self):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="🔢  Numbers per part — how many copies to print…",
                         command=self._open_numbers_per_part)
        menu.add_separator()
        menu.add_command(label="📊  Student list (Excel)…", command=self._export_students)
        menu.add_command(label="🎓  Outgoing students for HS directors (CSV)…",
                         command=self._export_for_hs)
        menu.add_separator()
        menu.add_command(label="🏷️  Barcode Sheet — scan Student IDs (PDF)…",
                         command=self._export_barcodes)
        try:
            menu.tk_popup(self.winfo_pointerx(), self.winfo_pointery())
        finally:
            menu.grab_release()

    def _export_barcodes(self):
        """A printable Code128 sheet for the handheld scanner — one label per
        student (their Student ID), sorted by last then first name, 3 columns."""
        try:
            import barcode_labels as bl
        except Exception as e:
            Messagebox.show_error(f"Could not load the barcode tool:\n{e}",
                                  title="Error", parent=self.winfo_toplevel())
            return
        year = self._year_var.get().strip() or None
        students = self.db.get_all_students(school_year=year)
        if not students:
            Messagebox.show_info(f"No students found for {year or 'this year'}.",
                                 title="Nothing to Print", parent=self.winfo_toplevel())
            return
        path = filedialog.asksaveasfilename(
            title="Save Barcode Sheet", parent=self.winfo_toplevel(),
            defaultextension=".pdf",
            initialfile=f"Student_Barcodes_{(year or 'all')}.pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if not path:
            return
        try:
            printed, skipped = bl.export_student_barcodes(students, path)
        except Exception as e:
            Messagebox.show_error(f"Could not create the PDF:\n{e}",
                                  title="Export Error", parent=self.winfo_toplevel())
            return
        msg = f"Created a barcode sheet with {printed} student(s)."
        if skipped:
            msg += (f"\n\n{skipped} student(s) were left off because they have no "
                    "Student ID to scan. Add their Student ID to include them.")
        if Messagebox.yesno(msg + "\n\nOpen it now?", title="Barcode Sheet",
                            parent=self.winfo_toplevel()) == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)

    def _import_csv(self):
        paths = filedialog.askopenfilenames(
            title="Select Student Roster CSV Files (up to 10)",
            filetypes=[("CSV files", "*.csv *.CSV"), ("All files", "*.*")],
            parent=self.winfo_toplevel(),
        )
        if not paths:
            return
        if len(paths) > 10:
            Messagebox.show_warning(
                "Please select at most 10 files at a time.",
                title="Too Many Files"
            ,
            parent=self.winfo_toplevel())
            return
        school_year = self._year_var.get() or _current_school_year()
        dlg = _StudentImportDialog(
            self.winfo_toplevel(), self.db, list(paths), school_year,
            program_type=self.program_type,
        )
        self.wait_window(dlg)
        self.refresh()

    def _add_student(self):
        high_water = self._max_student_id()
        dlg = StudentDialog(self.winfo_toplevel(), self.db,
                             student_id=None,
                             default_year=self._year_var.get(),
                             program_type=self.program_type,
                             site=self._site())
        self.wait_window(dlg)
        self._claim_new_students(high_water)
        self.refresh()

    def _roster_groups(self):
        """This school's sections (and its choir, if anybody is in it)."""
        if not self.site_id:
            return []
        try:
            from ui.ensembles import site_groups
            site = self.db.get_site(self.site_id)
            return site_groups(self.db, site, self._year_var.get() or None) if site else []
        except Exception:
            return []

    def _site(self):
        """The school this roster belongs to, or None for the secondary one."""
        if not self.site_id:
            return None
        try:
            row = self.db.get_site(self.site_id)
            return dict(row) if row else None
        except Exception:
            return None

    def _max_student_id(self):
        try:
            with self.db._connect() as conn:
                row = conn.execute("SELECT MAX(id) AS m FROM students").fetchone()
            return (row["m"] or 0) if row else 0
        except Exception:
            return 0

    def _claim_new_students(self, since_id):
        """A child added here belongs to this school.

        Rows created before the dialog opened are left alone: a profile with
        several schools also has older unassigned rows, and claiming those for
        whichever school is on screen is the mix-up all of this exists to stop.
        """
        if self.site_id is None:
            return
        try:
            with self.db._connect() as conn:
                conn.execute(
                    "UPDATE students SET site_id = ? "
                    "WHERE site_id IS NULL AND id > ?", (self.site_id, since_id))
                conn.commit()
        except Exception:
            pass

    def _edit_student(self):
        sid = self._get_selected()
        if sid is None:
            return
        dlg = StudentDialog(self.winfo_toplevel(), self.db,
                             student_id=sid,
                             default_year=self._year_var.get(),
                             program_type=self.program_type,
                             site=self._site())
        self.wait_window(dlg)
        self.refresh()
        self._load_detail(sid)

    def _name_of(self, sid):
        st = self.db.get_student(sid)
        if not st:
            return ""
        return f"{st['first_name']} {st['last_name']}".strip()

    def _delete_student(self):
        """Remove the ticked students, or the highlighted one if none are ticked.

        This used to read only the highlighted row, so ticking a whole roster
        and pressing Delete did nothing at all, and then deleted exactly the
        one student who happened to be highlighted.  Silence is the worst
        possible answer to a bulk action: it reads as a broken button, and the
        obvious next move (click a name and press it again) is the one that
        does damage.
        """
        ids = [i for i in self._selected_ids() if i]
        if not ids:
            Messagebox.show_warning(
                "Tick the boxes next to the students you want to remove "
                "(or use 'Select All Shown'), then press Delete.",
                title="No Students Selected",
                parent=self.winfo_toplevel())
            return

        # An instrument out on loan is the one thing that has to be settled
        # first; those students are kept rather than the whole action refused,
        # so one unreturned trumpet does not block a roster of thirty.
        held = [i for i in ids if self.db.get_student_active_checkout_count(i)]
        deletable = [i for i in ids if i not in held]
        if not deletable:
            who = (self._name_of(held[0]) if len(held) == 1
                   else f"All {len(held)} of them")
            Messagebox.show_warning(
                f"{who} still {'has' if len(held) == 1 else 'have'} "
                f"instruments checked out.\n\nCheck the instruments in first, "
                f"then delete.",
                title="Cannot Delete",
                parent=self.winfo_toplevel())
            return

        if len(deletable) == 1:
            question = f"Remove {self._name_of(deletable[0])} from the roster?"
        else:
            names = [self._name_of(i) for i in deletable[:4]]
            listed = ", ".join(n for n in names if n)
            if len(deletable) > 4:
                listed += f", and {len(deletable) - 4} more"
            question = (f"Remove {len(deletable)} students from the roster?"
                        f"\n\n{listed}")
        kept = ""
        if held:
            kept = (f"\n\n{len(held)} of the ones you ticked "
                    f"{'has' if len(held) == 1 else 'have'} instruments "
                    f"checked out and will be kept.")
        if Messagebox.yesno(
                question + kept
                + "\n\nThey move to the Inactive / Former Students list, where "
                  "you can put any of them back.",
                title="Remove students?",
                parent=self.winfo_toplevel()) != "Yes":
            return

        for sid in deletable:
            self.db.deactivate_student(sid)
        self._checked.clear()
        self._selected_id = None
        self.refresh()
        self._update_sel_count()
        if held:
            Messagebox.show_info(
                f"{len(deletable)} removed.  {len(held)} kept, still holding "
                f"instruments.", title="Done",
                parent=self.winfo_toplevel())

    def _show_inactive_window(self):
        """List former (inactive) students, most-recently-enrolled first, with a
        Reactivate action."""
        inactive = [dict(s) for s in self.db.get_all_students(include_inactive=True)
                    if not s["is_active"]]
        inactive.sort(key=lambda s: ((s.get("school_year") or ""),
                                     (s.get("last_name") or "").lower()),
                      reverse=True)

        win = ttk.Toplevel(self.winfo_toplevel())
        win.title("Inactive / Former Students")
        win.resizable(True, True)
        win.grab_set()
        win.lift()

        hdr = ttk.Frame(win, bootstyle=SECONDARY)
        hdr.pack(fill=X)
        ttk.Label(hdr, text="🗂️  Inactive / Former Students",
                  font=("Segoe UI", 13, "bold"),
                  bootstyle=(INVERSE, SECONDARY)).pack(pady=10, padx=16, anchor=W)
        ttk.Label(win, text="Most recently enrolled at the top. Select a student and "
                            "click Reactivate if they've returned.",
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(anchor=W, padx=14, pady=(8, 2))

        frame = ttk.Frame(win)
        frame.pack(fill=BOTH, expand=True, padx=14, pady=6)
        cols = ("Name", "Grade", "Last Enrolled", "Instrument", "Ensembles")
        sb = ttk.Scrollbar(frame, orient=VERTICAL)
        tree = ttk.Treeview(frame, columns=cols, show="headings",
                            yscrollcommand=sb.set, selectmode="browse", bootstyle=SECONDARY)
        sb.config(command=tree.yview)
        sb.pack(side=RIGHT, fill=Y)
        tree.pack(fill=BOTH, expand=True)
        for c, w in zip(cols, [190, 60, 110, 130, 180]):
            tree.heading(c, text=c, anchor=W)
            tree.column(c, width=w, anchor=W, stretch=c in ("Name", "Ensembles"))

        def _fill():
            tree.delete(*tree.get_children())
            for s in inactive:
                prim = self._sval(s, "primary_instrument") or ""
                sec = self._sval(s, "secondary_instrument") or ""
                tree.insert("", "end", iid=str(s["id"]), values=(
                    f"{s['last_name']}, {s['first_name']}".strip(", "),
                    s["grade"] or "",
                    s["school_year"] or "",
                    prim + (f" / {sec}" if sec else ""),
                    self._sval(s, "ensembles") or "",
                ))
        _fill()

        def _do_reactivate():
            sel = tree.selection()
            if not sel:
                Messagebox.show_warning("Select a former student to reactivate.",
                                        title="No Selection", parent=win)
                return
            student = self.db.get_student(int(sel[0]))
            if not student:
                return

            def _after(new_year):
                # Remove from the local list + refresh both views
                nonlocal inactive
                inactive = [s for s in inactive if s["id"] != student["id"]]
                _fill()
                self._year_var.set(new_year)
                self.refresh()

            self._reactivate_dialog(student, parent=win, on_done=_after)

        btns = ttk.Frame(win)
        btns.pack(fill=X, padx=14, pady=12)
        ttk.Button(btns, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btns, text="♻️ Reactivate Selected", bootstyle=SUCCESS,
                   command=_do_reactivate).pack(side=RIGHT, padx=4)

        from ui.theme import fit_window
        fit_window(win, 780, 520)

    def _reactivate_dialog(self, student, parent=None, on_done=None):
        """Reactivate a student, placing them in a chosen school year (returning
        students usually need a fresh year).  Calls on_done(new_year) on success."""
        parent = parent or self.winfo_toplevel()
        sid = student["id"]
        name = f"{student['first_name']} {student['last_name']}".strip()
        win = ttk.Toplevel(parent)
        win.title("Reactivate Student")
        win.resizable(False, False)
        win.grab_set()
        ttk.Label(win, text=f"Reactivate {name}", font=("Segoe UI", 11, "bold"),
                  bootstyle=SUCCESS).pack(anchor=W, padx=18, pady=(16, 4))
        ttk.Label(win, text="Place them in which school year?",
                  font=("Segoe UI", 9)).pack(anchor=W, padx=18)
        year_var = tk.StringVar(value=_current_school_year())
        years = self.db.get_school_years()
        for y in (_next_school_year(), _current_school_year()):
            if y not in years:
                years.insert(0, y)
        ttk.Combobox(win, textvariable=year_var, values=years, width=16).pack(
            anchor=W, padx=18, pady=(2, 4))
        ttk.Label(win, text="(Keeps their existing contact info; you can update "
                            "ensembles/periods afterward.)",
                  font=("Segoe UI", 8), foreground=muted_fg(),
                  wraplength=320, justify=LEFT).pack(anchor=W, padx=18)

        def _do():
            data = dict(student)
            data["is_active"] = 1
            data["school_year"] = year_var.get().strip() or student["school_year"]
            self.db.update_student(sid, data)
            win.destroy()
            Messagebox.show_info(f"{name} reactivated for {data['school_year']}.",
                                 title="Reactivated",
                                 parent=self.winfo_toplevel())
            if on_done:
                on_done(data["school_year"])

        btns = ttk.Frame(win)
        btns.pack(fill=X, padx=16, pady=14)
        ttk.Button(btns, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btns, text="Reactivate", bootstyle=SUCCESS, command=_do).pack(side=RIGHT, padx=4)
        from ui.theme import fit_window
        fit_window(win, 380, 240)

    def _export_students(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required. Run: pip install openpyxl",
                                  title="Missing Dependency",
                                  parent=self.winfo_toplevel())
            return

        # ── Scope dialog ───────────────────────────────────────────────────
        win = ttk.Toplevel(self.winfo_toplevel())
        win.title("Export Students")
        win.resizable(False, False)
        win.grab_set()
        ttk.Label(win, text="Export students to Excel", font=("Segoe UI", 11, "bold"),
                  bootstyle=PRIMARY).pack(anchor=W, padx=18, pady=(16, 8))
        scope_var = tk.StringVar(value="all")
        ttk.Radiobutton(win, text="All school years (alumni included)", value="all",
                        variable=scope_var).pack(anchor=W, padx=24, pady=2)
        ttk.Radiobutton(win, text=f"This school year only ({self._year_var.get()})",
                        value="year", variable=scope_var).pack(anchor=W, padx=24, pady=2)
        inactive_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(win, text="Include inactive students", variable=inactive_var,
                        bootstyle=SECONDARY).pack(anchor=W, padx=24, pady=(6, 2))
        result = {"ok": False}

        def _ok():
            result["ok"] = True
            win.destroy()
        btns = ttk.Frame(win)
        btns.pack(fill=X, padx=16, pady=14)
        ttk.Button(btns, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btns, text="Export", bootstyle=PRIMARY, command=_ok).pack(side=RIGHT, padx=4)
        from ui.theme import fit_window
        fit_window(win, 360, 230)
        self.wait_window(win)
        if not result["ok"]:
            return

        year = None if scope_var.get() == "all" else (self._year_var.get() or None)
        students = list(self.db.get_all_students(
            school_year=year, include_inactive=inactive_var.get(),
            level="secondary"))
        if not students:
            Messagebox.show_info("No students match that scope.", title="Nothing to Export",
    parent=self.winfo_toplevel())
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.freeze_panes = "A2"
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="2E5E8E")
        alt_fill = PatternFill("solid", fgColor="F5F5F5")
        border = Border(*[Side(style="thin", color="CCCCCC")] * 4)

        headers = ["School Year", "Last Name", "First Name", "Grade", "Status",
                   "Ensembles", "Class Periods", "Primary Instrument", "Secondary Instrument",
                   "Student Email", "Phone",
                   "Parent 1", "Parent 1 Email", "Parent 1 Phone",
                   "Parent 2", "Parent 2 Email", "Parent 2 Phone",
                   "Address", "City", "State", "ZIP", "Notes"]
        keys = ["school_year", "last_name", "first_name", "grade", None,
                "ensembles", "class_periods", "primary_instrument", "secondary_instrument",
                "student_email", "phone",
                "parent1_name", "parent1_email", "parent1_phone",
                "parent2_name", "parent2_email", "parent2_phone",
                "address", "city", "state", "zip_code", "notes"]

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border
        ws.row_dimensions[1].height = 18

        for r, s in enumerate(students, 2):
            fill = alt_fill if r % 2 == 0 else None
            for c, key in enumerate(keys, 1):
                if key is None:
                    val = "Active" if self._sval(s, "is_active") else "Inactive"
                else:
                    val = self._sval(s, key)
                cell = ws.cell(row=r, column=c, value=val or "")
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

        widths = [12, 14, 14, 7, 9, 20, 12, 16, 16, 24, 14, 18, 24, 14, 18, 24, 14, 24, 14, 7, 8, 30]
        for col, w in zip(range(1, len(widths) + 1), widths):
            ws.column_dimensions[get_column_letter(col)].width = w

        from tkinter import filedialog
        import datetime
        scope_tag = "AllYears" if scope_var.get() == "all" else self._year_var.get().replace("-", "_")
        path = filedialog.asksaveasfilename(
            title="Save Student Export", parent=self.winfo_toplevel(),
            defaultextension=".xlsx",
            initialfile=f"Students_{scope_tag}_{datetime.date.today().isoformat()}.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            wb.save(path)
        except Exception as e:
            Messagebox.show_error(f"Could not save file:\n{e}", title="Save Error",
    parent=self.winfo_toplevel())
            return
        if Messagebox.yesno(f"Exported {len(students)} student(s).\n\nOpen the file now?",
                            title="Export Complete",
                            parent=self.winfo_toplevel()) == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)

    # ── Middle-school → high-school transfer ────────────────────────────────────

    def _export_for_hs(self):
        """Export a grade's students as a named-header CSV that a high-school
        director can re-import to auto-fill instrumentation + summer contacts."""
        # Choose grade + year
        win = ttk.Toplevel(self.winfo_toplevel())
        win.title("Export for HS Directors")
        win.resizable(False, False)
        win.grab_set()
        ttk.Label(win, text="Export outgoing students for high-school directors",
                  font=("Segoe UI", 11, "bold"), bootstyle=PRIMARY).pack(
            anchor=W, padx=18, pady=(16, 4))
        ttk.Label(win, text="Includes each student's instrument(s), ensembles, and contact "
                            "info so HS directors can reach them for summer rehearsals.",
                  font=("Segoe UI", 8), foreground=muted_fg(), wraplength=380,
                  justify=LEFT).pack(anchor=W, padx=18, pady=(0, 8))
        r1 = ttk.Frame(win); r1.pack(anchor=W, padx=18, pady=2)
        ttk.Label(r1, text="Grade:", font=("Segoe UI", 9)).pack(side=LEFT)
        grade_var = tk.StringVar(value="8")
        ttk.Combobox(r1, textvariable=grade_var, values=GRADE_OPTIONS, width=8).pack(
            side=LEFT, padx=(4, 12))
        ttk.Label(r1, text="School Year:", font=("Segoe UI", 9)).pack(side=LEFT)
        year_var = tk.StringVar(value=self._year_var.get() or _current_school_year())
        ttk.Combobox(r1, textvariable=year_var, values=self.db.get_school_years(),
                     width=12).pack(side=LEFT, padx=(4, 0))
        res = {"ok": False}

        def _ok():
            res["ok"] = True; win.destroy()
        btns = ttk.Frame(win); btns.pack(fill=X, padx=16, pady=14)
        ttk.Button(btns, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btns, text="Export CSV", bootstyle=PRIMARY, command=_ok).pack(side=RIGHT, padx=4)
        from ui.theme import fit_window
        fit_window(win, 420, 220)
        self.wait_window(win)
        if not res["ok"]:
            return

        grade = grade_var.get().strip()
        year = year_var.get().strip() or None
        students = [s for s in self.db.get_all_students(school_year=year)
                    if str(self._sval(s, "grade")).strip() == grade]
        if not students:
            Messagebox.show_info(f"No grade-{grade} students found for {year}.",
                                 title="Nothing to Export",
                                 parent=self.winfo_toplevel())
            return

        path = filedialog.asksaveasfilename(
            title="Save CSV for HS Directors", parent=self.winfo_toplevel(),
            defaultextension=".csv",
            initialfile=f"Outgoing_Grade{grade}_{(year or '').replace('-', '_')}.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow([h for h, _ in HS_TRANSFER_FIELDS])
                for s in students:
                    writer.writerow([self._sval(s, key) or "" for _, key in HS_TRANSFER_FIELDS])
        except Exception as e:
            Messagebox.show_error(f"Could not save file:\n{e}", title="Save Error",
    parent=self.winfo_toplevel())
            return
        Messagebox.show_info(
            f"Exported {len(students)} grade-{grade} student(s) to:\n{path}\n\n"
            "High-school directors can import this after loading their own roster to "
            "auto-fill instrumentation.", title="Export Complete",
            parent=self.winfo_toplevel())

    def _import_instruments_from_csv(self):
        """Update existing students' instruments (and fill blank contacts) from a
        named-header transfer CSV exported by a feeder school."""
        if Messagebox.yesno(
            "Import instrumentation from another school's CSV?\n\n"
            "Load your own class roster FIRST. This matches students by ID or name "
            "and fills in what they played previously — it won't create new students.",
            title="Update Instruments from CSV",
            parent=self.winfo_toplevel()) != "Yes":
            return
        path = filedialog.askopenfilename(
            title="Select transfer CSV (from feeder school)",
            filetypes=[("CSV files", "*.csv *.CSV"), ("All files", "*.*")],
            parent=self.winfo_toplevel(),
        )
        if not path:
            return
        try:
            with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
                reader = csv.DictReader(f)
                rows = [{(k or "").strip(): (v or "").strip() for k, v in r.items()} for r in reader]
        except Exception as e:
            Messagebox.show_error(f"Could not read CSV:\n{e}", title="Import Error",
    parent=self.winfo_toplevel())
            return
        if not rows:
            Messagebox.show_info("That CSV has no rows.", title="Nothing to Import",
    parent=self.winfo_toplevel())
            return

        matched = updated = 0
        unmatched = []
        cur_year = self._year_var.get() or None
        for row in rows:
            sid = row.get("Student ID", "")
            first = row.get("First Name", "")
            last = row.get("Last Name", "")
            student = None
            if sid:
                student = self.db.find_student_by_student_id(sid)
            if not student and (first or last):
                student = (self.db.find_student_by_name(first, last, cur_year)
                           or self.db.find_student_by_name(first, last))
            if not student:
                unmatched.append(f"{first} {last}".strip() or sid or "(unknown)")
                continue
            matched += 1
            prim = row.get("Primary Instrument", "")
            sec = row.get("Secondary Instrument", "")
            if prim or sec:
                self.db.update_student_instruments(
                    student["id"],
                    primary=prim if prim else None,
                    secondary=sec if sec else None)
                updated += 1
            # Fill blank contact fields (useful for summer outreach)
            data = dict(student)
            changed = False
            for key in ("student_email", "phone", "parent1_name", "parent1_email",
                        "parent1_phone", "parent2_name", "parent2_email", "parent2_phone"):
                hdr = next((h for h, k in HS_TRANSFER_FIELDS if k == key), None)
                incoming = row.get(hdr, "") if hdr else ""
                if incoming and not (self._sval(student, key) or "").strip():
                    data[key] = incoming
                    changed = True
            if changed:
                data["is_active"] = student["is_active"]
                self.db.update_student(student["id"], data)

        self.refresh()
        msg = (f"Matched {matched} of {len(rows)} student(s).\n"
               f"Updated instrumentation for {updated}.")
        if unmatched:
            preview = ", ".join(unmatched[:12]) + ("…" if len(unmatched) > 12 else "")
            msg += f"\n\n{len(unmatched)} could not be matched (not on your roster):\n{preview}"
        Messagebox.show_info(msg, title="Instrument Import Complete",
    parent=self.winfo_toplevel())


# ── CSV roster parser ─────────────────────────────────────────────────────────

# Column indices (positional — avoids issues with duplicate header names in
# the district CSV export, which has two "Phone" and two "ParentEmail" cols).
_COL_SID      = 0   # Student ID
_COL_NAME     = 1   # Student Name  (Last, First)
_COL_GRADE    = 2   # Grd
_COL_GENDER   = 3   # Gen
_COL_BDATE    = 4   # Birth Date
_COL_PNAME    = 5   # Parent Name
_COL_PEMAIL   = 8   # ParentEmail   (first occurrence)
_COL_PHONE    = 9   # Phone         (first occurrence — main contact phone)
_COL_RELATION = 11  # Relation
_COL_ORDERBY  = 17  # Orderby (1 = primary contact, 2 = secondary, …)
_COL_SEMAIL   = 21  # Student Email
_COL_ADDRESS  = 22  # Street Address
_COL_CSZ      = 23  # CityStateZip  e.g. "BELLEVUE, WA 98004"

_CSZ_RE = re.compile(r'^(.*?),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$')


def _parse_csz(csz: str):
    """'BELLEVUE, WA 98004' → ('Bellevue', 'WA', '98004')"""
    m = _CSZ_RE.match((csz or "").strip())
    if m:
        return m.group(1).strip().title(), m.group(2), m.group(3)
    return (csz or "").strip(), "", ""


def _split_name(name: str):
    """'Last, First Middle' → (first, last)"""
    name = (name or "").strip()
    if "," in name:
        last, first = name.split(",", 1)
        return first.strip(), last.strip()
    parts = name.split()
    return (parts[0] if parts else ""), (" ".join(parts[1:]) if len(parts) > 1 else "")


def _normalize_date(d: str) -> str:
    if not d:
        return ""
    try:
        return datetime.strptime(d.strip(), "%m/%d/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return d.strip()


def _parse_student_csvs(paths: list) -> dict:
    """
    Parse one or more school district roster CSV files.

    Each file may have multiple rows per student (one per parent/guardian).
    Students are deduplicated by Student ID across all files.

    Returns a dict  {student_id_or_name: student_data_dict}  ready for
    db.add_student().
    """
    students = {}   # keyed by student_id (or raw name as fallback)

    for path in paths:
        with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
            reader = csv.reader(f)
            next(reader, None)   # skip header row
            for row in reader:
                if len(row) <= _COL_CSZ:
                    continue

                sid     = row[_COL_SID].strip()
                name    = row[_COL_NAME].strip()
                grade   = row[_COL_GRADE].strip().lstrip("0") or "0"
                gender  = {"M": "Male", "F": "Female"}.get(
                              row[_COL_GENDER].strip(), "")
                bdate   = _normalize_date(row[_COL_BDATE])
                p_name  = row[_COL_PNAME].strip()
                p_email = row[_COL_PEMAIL].strip()
                p_phone = row[_COL_PHONE].strip()
                p_rel   = row[_COL_RELATION].strip()
                orderby = row[_COL_ORDERBY].strip() or "1"
                s_email = row[_COL_SEMAIL].strip()
                address = row[_COL_ADDRESS].strip()
                csz     = row[_COL_CSZ].strip()

                key = sid if sid else name
                if not key:
                    continue

                # First time we see this student — create the base record
                if key not in students:
                    first, last = _split_name(name)
                    city, state, zip_code = _parse_csz(csz)
                    students[key] = {
                        "student_id":    sid,
                        "first_name":    first,
                        "last_name":     last,
                        "grade":         grade,
                        "gender":        gender,
                        "birth_date":    bdate,
                        "student_email": s_email,
                        "address":       address,
                        "city":          city,
                        "state":         state,
                        "zip_code":      zip_code,
                        "_parents":      {},
                    }

                # Accumulate parent contacts keyed by Orderby value.
                # Only keep the first occurrence of each Orderby number.
                if p_name and orderby not in students[key]["_parents"]:
                    students[key]["_parents"][orderby] = {
                        "name":     p_name,
                        "phone":    p_phone,
                        "email":    p_email,
                        "relation": p_rel,
                    }

    # Flatten parent dicts into parent1 / parent2 fields
    result = {}
    for key, s in students.items():
        parents = s.pop("_parents")
        sorted_keys = sorted(parents.keys())
        p1 = parents.get(sorted_keys[0], {}) if sorted_keys else {}
        p2 = parents.get(sorted_keys[1], {}) if len(sorted_keys) > 1 else {}

        s["parent1_name"]     = p1.get("name", "")
        s["parent1_phone"]    = p1.get("phone", "")
        s["parent1_email"]    = p1.get("email", "")
        s["parent1_relation"] = p1.get("relation", "")
        s["parent2_name"]     = p2.get("name", "")
        s["parent2_phone"]    = p2.get("phone", "")
        s["parent2_email"]    = p2.get("email", "")
        s["parent2_relation"] = p2.get("relation", "")
        # Use primary parent phone as the student's main contact phone
        s["phone"] = p1.get("phone", "") or p2.get("phone", "")

        result[key] = s

    return result


# ── Import progress dialog ────────────────────────────────────────────────────

class _StudentImportDialog(_ClassOptionsMixin, ttk.Toplevel):
    """Shows import progress and results for a CSV roster import."""

    _class_options_include_empty = True     # assigns a class, so offer empties

    def __init__(self, parent, db, paths: list, school_year: str, program_type="band"):
        super().__init__(parent)
        self.db          = db
        self.paths       = paths
        self.school_year = school_year
        self.program_type = program_type
        self._ensemble_var = tk.StringVar(value="— none —")
        self._period_var = tk.StringVar(value="— none —")
        self._instrument_var = tk.StringVar(value="— none —")
        self._carry_over_var = tk.BooleanVar(value=True)

        self.title("Import Student Roster")
        self.resizable(False, False)
        self.grab_set()
        self.lift()

        self._build()

        from ui.theme import fit_window
        fit_window(self, 560, 540)

    def _build(self):
        ttk.Label(
            self,
            text=f"Importing students for {self.school_year}",
            font=("Segoe UI", 11, "bold"),
            bootstyle=PRIMARY,
        ).pack(pady=(16, 4), padx=16, anchor=W)

        # ── Label-the-whole-class controls ────────────────────────────────
        cfg = tk.LabelFrame(self, text=" Label this class (optional) ",
                            font=("Segoe UI", 9, "bold"), padx=10, pady=8)
        cfg.pack(fill=X, padx=16, pady=(2, 6))
        ttk.Label(
            cfg,
            text="This roster has no class column. Pick an ensemble and/or period to "
                 "assign to everyone in this import (existing students are updated too).",
            font=("Segoe UI", 8), foreground=muted_fg(), wraplength=480, justify=LEFT,
        ).pack(anchor=W, pady=(0, 6))
        row = ttk.Frame(cfg)
        row.pack(fill=X)
        ttk.Label(row, text="Ensemble:", font=("Segoe UI", 8)).pack(side=LEFT)
        ttk.Combobox(row, textvariable=self._ensemble_var, state="readonly", width=20,
                     values=["— none —"] + self._class_options()).pack(
            side=LEFT, padx=(4, 12))
        ttk.Label(row, text="Class Period:", font=("Segoe UI", 8)).pack(side=LEFT)
        ttk.Combobox(row, textvariable=self._period_var, state="readonly", width=8,
                     values=["— none —"] + PERIOD_OPTIONS).pack(side=LEFT, padx=(4, 0))

        row2 = ttk.Frame(cfg)
        row2.pack(fill=X, pady=(6, 0))
        ttk.Label(row2, text="Instrument (whole class):", font=("Segoe UI", 8)).pack(side=LEFT)
        ttk.Combobox(row2, textvariable=self._instrument_var, state="readonly", width=22,
                     values=["— none —"] + instruments_for(self.program_type)).pack(
            side=LEFT, padx=(4, 0))

        ttk.Checkbutton(
            cfg, bootstyle=PRIMARY, variable=self._carry_over_var,
            text="Carry over instruments from previous years for returning students",
        ).pack(anchor=W, pady=(8, 0))
        ttk.Label(
            cfg,
            text="Matches this roster against prior years (by ID or name) and fills in "
                 "each returning student's instrument automatically — only where it's blank.",
            font=("Segoe UI", 8), foreground=muted_fg(), wraplength=480, justify=LEFT,
        ).pack(anchor=W, pady=(0, 2))

        self._start_btn = ttk.Button(self, text="Start Import", bootstyle=SUCCESS,
                                     command=self._start)
        self._start_btn.pack(pady=(0, 6))

        self._progress = ttk.Progressbar(self, mode="indeterminate", bootstyle=PRIMARY)
        self._progress.pack(fill=X, padx=20, pady=4)

        log_frame = ttk.Frame(self)
        log_frame.pack(fill=BOTH, expand=True, padx=16, pady=8)

        self._log = tk.Text(
            log_frame, height=14, font=("Consolas", 9),
            state="disabled", relief="flat", bd=1, bg="#F8F8F8",
        )
        sb = ttk.Scrollbar(log_frame, orient=VERTICAL, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        self._log.pack(fill=BOTH, expand=True)

        self._close_btn = ttk.Button(
            self, text="Close", bootstyle=PRIMARY,
            state="disabled", command=self.destroy,
        )
        self._close_btn.pack(pady=(0, 12))

    def _log_msg(self, msg: str):
        self._log.config(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.config(state="disabled")
        self.update()

    def _start(self):
        self._start_btn.config(state="disabled")
        self._progress.start(10)
        self.after(50, self._run_import)

    def _run_import(self):
        # Resolve the batch label chosen in the config controls
        ens = self._ensemble_var.get()
        per = self._period_var.get()
        instr = self._instrument_var.get()
        batch_ensemble = ens if ens and ens != "— none —" else None
        batch_period = per if per and per != "— none —" else None
        batch_instrument = instr if instr and instr != "— none —" else None
        carry_over = self._carry_over_var.get()
        label_ids = []   # student ids to label after import
        try:
            # ── Parse files ───────────────────────────────────────────────
            for i, path in enumerate(self.paths, 1):
                self._log_msg(
                    f"Reading file {i}/{len(self.paths)}: {os.path.basename(path)}"
                )

            students = _parse_student_csvs(self.paths)
            self._log_msg(
                f"\nFound {len(students)} unique student(s) across "
                f"{len(self.paths)} file(s)."
            )
            self._log_msg(f"Importing into school year: {self.school_year}\n")

            # ── Import with deduplication ─────────────────────────────────
            imported = 0
            skipped  = 0
            enriched = 0
            fields_filled = 0

            for data in students.values():
                data["school_year"] = self.school_year

                # Primary dedup key: student_id + school_year
                existing = None
                if data.get("student_id"):
                    with self.db._connect() as conn:
                        existing = conn.execute(
                            "SELECT id FROM students "
                            "WHERE student_id=? AND school_year=?",
                            (data["student_id"], self.school_year),
                        ).fetchone()

                # Fallback dedup: exact first+last name match
                if not existing:
                    existing = self.db.find_student_by_name(
                        data.get("first_name", ""),
                        data.get("last_name", ""),
                        self.school_year,
                    )

                if existing:
                    # The student is already on the roster, but that record may
                    # have come from a class list, which carries names and
                    # nothing else.  Fill in the contact details it never had —
                    # skipping the row outright is why beginners ended up with
                    # no address or parent information.
                    n = self.db.fill_student_blanks(existing["id"], data)
                    if n:
                        enriched += 1
                        fields_filled += n
                    else:
                        skipped += 1
                    label_ids.append(existing["id"])
                else:
                    new_id = self.db.add_student(data)
                    label_ids.append(new_id)
                    imported += 1

            # ── Apply the batch ensemble / period label ───────────────────
            if (batch_ensemble or batch_period) and label_ids:
                if batch_ensemble:
                    self.db.bulk_set_student_multi(
                        label_ids, "ensembles", [batch_ensemble], replace=False)
                if batch_period:
                    self.db.bulk_set_student_multi(
                        label_ids, "class_periods", [batch_period], replace=False)
                bits = ", ".join(filter(None, [batch_ensemble, f"Period {batch_period}"
                                               if batch_period else None]))
                self._log_msg(f"\nLabeled {len(label_ids)} student(s) as: {bits}")

            # ── Carry over instruments from prior years (returning students) ──
            if carry_over and label_ids:
                filled = self.db.carry_over_instruments(label_ids)
                self._log_msg(
                    f"\nCarried over instruments for {filled} returning student(s) "
                    f"from previous years.")

            # ── Apply a whole-class instrument (explicit, overrides above) ────
            if batch_instrument and label_ids:
                self.db.bulk_set_student_field(label_ids, "primary_instrument", batch_instrument)
                self._log_msg(f"Set instrument '{batch_instrument}' for {len(label_ids)} student(s).")

            # ── Summary ───────────────────────────────────────────────────
            self._log_msg("─" * 42)
            self._log_msg(f"Import complete!")
            self._log_msg(f"  Imported : {imported} new student(s)")
            self._log_msg(
                f"  Updated  : {enriched} existing student(s) "
                f"({fields_filled} blank field(s) filled in)"
            )
            self._log_msg(
                f"  Unchanged: {skipped} "
                f"(already on file for {self.school_year})"
            )

        except Exception as e:
            self._log_msg(f"\nError: {e}")
        finally:
            self._progress.stop()
            self._close_btn.config(state="normal")


class StudentDialog(_ClassOptionsMixin, ttk.Toplevel):
    """Add / edit one student."""

    _class_options_include_empty = True     # assigns a class, so offer empties

    def __init__(self, parent, db, student_id=None, default_year=None,
                 program_type="band", site=None):
        super().__init__(parent)
        self.db = db
        self.student_id = student_id
        self.program_type = program_type or "band"
        # The school being edited for, when this is a 5th grade roster.  It
        # decides the wording, the instrument order, and whether choir is
        # offered at all.
        self.site = dict(site) if site else None
        self._elementary = bool(self.site
                                and self.site.get("level") == "elementary")
        self.default_year = default_year or _current_school_year()

        self.title("Edit Student" if student_id else "Add Student")
        self.resizable(True, True)
        self.grab_set()
        self.lift()

        self._vars = {}
        self._multi_vars = {}   # key -> {option: BooleanVar} for checkbox groups
        self._multi_grids = {}  # key -> (grid frame, columns) so options can grow
        self._build()
        if student_id:
            self._load(student_id)

        from ui.theme import fit_window
        fit_window(self, 620, 680)

    def _build(self):
        hdr = ttk.Frame(self, bootstyle=PRIMARY)
        hdr.pack(fill=X)
        title = "Edit Student" if self.student_id else "Add Student"
        ttk.Label(hdr, text=title, font=("Segoe UI", 13, "bold"),
                  bootstyle=(INVERSE, PRIMARY)).pack(pady=12, padx=16, anchor=W)

        # Scrollable body
        canvas = tk.Canvas(self, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        canvas.pack(fill=BOTH, expand=True)

        content = ttk.Frame(canvas)
        cw = canvas.create_window((0, 0), window=content, anchor=NW)

        def _resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(cw, width=canvas.winfo_width())

        content.bind("<Configure>", _resize)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))

        def _wheel(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except tk.TclError:
                canvas.unbind_all("<MouseWheel>")
        canvas.bind_all("<MouseWheel>", _wheel)
        self.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        self._build_form(content)

        btn = ttk.Frame(self)
        btn.pack(fill=X, padx=16, pady=10)
        ttk.Button(btn, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btn, text="Save", bootstyle=SUCCESS,
                   command=self._save).pack(side=RIGHT, padx=4)

        # Mark Inactive / Reactivate — only shown when editing an existing student
        if self.student_id:
            student = self.db.get_student(self.student_id)
            is_active = student["is_active"] if student else 1
            if is_active:
                ttk.Button(btn, text="Mark Inactive", bootstyle=(WARNING, OUTLINE),
                           command=self._mark_inactive).pack(side=LEFT, padx=4)
            else:
                ttk.Button(btn, text="Reactivate", bootstyle=(SUCCESS, OUTLINE),
                           command=self._reactivate).pack(side=LEFT, padx=4)

    def _section(self, parent, text):
        f = ttk.Frame(parent)
        f.pack(fill=X, padx=8, pady=(10, 2))
        ttk.Label(f, text=text, font=("Segoe UI", 10, "bold"), bootstyle=PRIMARY).pack(side=LEFT)
        ttk.Separator(f).pack(side=LEFT, fill=X, expand=True, padx=8)

    def _field(self, parent, label, key, widget="entry", options=None, side=LEFT, width=20):
        f = ttk.Frame(parent)
        f.pack(side=side, padx=6, pady=2)
        ttk.Label(f, text=label, font=("Segoe UI", 8)).pack(anchor=W)
        var = tk.StringVar()
        self._vars[key] = var
        if widget == "combobox":
            ttk.Combobox(f, textvariable=var, values=options or [],
                          width=width).pack(anchor=W)
        else:
            ttk.Entry(f, textvariable=var, width=width).pack(anchor=W)
        return var

    def _checkbox_group(self, parent, key, options, columns=4):
        """A horizontal set of checkboxes; values collected into a comma string.

        The grid is remembered so ``_ensure_option`` can add a box later for a
        value the student already has — see ``_load``."""
        holder = self._multi_vars.setdefault(key, {})
        grid = ttk.Frame(parent)
        grid.pack(fill=X, padx=16, pady=(2, 0))
        self._multi_grids[key] = (grid, columns)
        for opt in options:
            self._ensure_option(key, opt)

    def _ensure_option(self, key, opt, value=False):
        """Add a checkbox for ``opt`` if the group doesn't already have one.

        Save rebuilds these fields from the ticked boxes alone, so an option
        that isn't on screen is an option that gets ERASED the moment the
        teacher opens a student and clicks Save.  Any value a student actually
        holds therefore gets a box, whether or not it matches the configured
        class list — the record decides what's offered, not the other way round.

        For the ensembles group, "already have one" means BY CLASS IDENTITY:
        a student holding "Entry Band" ticks the offered "MS Band (Entry)" box
        rather than growing a duplicate.  The box then saves the canonical
        spelling, so records converge on the configured names as they're
        edited.  Checkbox labels show the short form ("Entry"); the stored
        value is always the full canonical name.
        """
        import class_registry as cr
        holder = self._multi_vars.setdefault(key, {})
        is_class = key == "ensembles"
        if is_class:
            for existing, var in holder.items():
                if cr.same_class(existing, opt):
                    if value:
                        var.set(True)
                    return
        elif opt in holder:
            if value:
                holder[opt].set(True)
            return
        grid, columns = self._multi_grids.get(key, (None, 4))
        if grid is None:
            return
        var = tk.BooleanVar(value=value)
        holder[opt] = var
        i = len(holder) - 1
        text = cr.short_class_label(opt) if is_class else opt
        ttk.Checkbutton(grid, text=text, variable=var, bootstyle=PRIMARY).grid(
            row=i // columns, column=i % columns, sticky=W, padx=6, pady=2)

    def _build_form(self, parent):
        self._section(parent, "Basic Information")
        row0 = ttk.Frame(parent)
        row0.pack(fill=X, padx=16)
        self._field(row0, "School Year *", "school_year", side=LEFT, width=14)
        self._field(row0, "First Name *", "first_name", side=LEFT, width=18)
        self._field(row0, "Last Name *", "last_name", side=LEFT, width=18)

        row0b = ttk.Frame(parent)
        row0b.pack(fill=X, padx=16, pady=(4, 0))
        self._field(row0b, "Preferred Name (shown instead of first)", "preferred_name",
                    side=LEFT, width=24)
        ttk.Label(row0b, text="Middle names / initials are hidden automatically.",
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(side=LEFT, padx=(8, 0))

        row1 = ttk.Frame(parent)
        row1.pack(fill=X, padx=16, pady=(4, 0))
        self._field(row1, "Grade", "grade", widget="combobox",
                    options=GRADE_OPTIONS, side=LEFT, width=8)
        self._field(row1, "Gender", "gender", widget="combobox",
                    options=GENDER_OPTIONS, side=LEFT, width=16)
        self._field(row1, "Student ID", "student_id", side=LEFT, width=16)

        if self._elementary:
            # No class periods in an elementary school -- the group is a pullout
            # from the regular class, twice a week on the specialist rotation --
            # and none of the middle school classes belong here either. They were
            # showing because this dialog offered the configured secondary class
            # list to everybody.
            self._section(parent, "Section")
            self._build_section_row(parent)
        else:
            self._section(parent, "Ensembles & Class Periods")
            ttk.Label(parent, text="Ensemble(s):", font=("Segoe UI", 8),
                      foreground=muted_fg()).pack(anchor=W, padx=16)
            self._checkbox_group(parent, "ensembles", self._class_options(), columns=3)
            ttk.Label(parent, text="Class Period(s):", font=("Segoe UI", 8),
                      foreground=muted_fg()).pack(anchor=W, padx=16, pady=(6, 0))
            self._checkbox_group(parent, "class_periods", PERIOD_OPTIONS, columns=7)

        instr_row = ttk.Frame(parent)
        instr_row.pack(fill=X, padx=16, pady=(6, 0))

        if self._elementary:
            # A 5th grader plays one instrument.  "Primary" only means anything
            # next to a "Secondary", and there is no secondary, no jazz band and
            # no doubling down here -- so those fields are not shown at all
            # rather than sitting empty for the whole year.
            from ui.ensembles import fifth_grade_instruments
            self._field(instr_row, "5th Grade Instrument", "primary_instrument",
                        widget="combobox",
                        options=fifth_grade_instruments(self.program_type),
                        side=LEFT, width=22)
            ttk.Label(instr_row,
                      text="The usual starters come first; the rest of the list "
                           "is underneath for the occasional oboist.",
                      font=("Segoe UI", 8), foreground=muted_fg(),
                      wraplength=250, justify=LEFT).pack(side=LEFT, padx=(8, 0))
            self._build_choir_row(parent)
        else:
            instr_opts = instruments_for(self.program_type)
            self._field(instr_row, "Concert Instrument (Primary)", "primary_instrument",
                        widget="combobox", options=instr_opts, side=LEFT, width=22)
            self._field(instr_row, "Concert Instrument (Secondary)", "secondary_instrument",
                        widget="combobox", options=instr_opts, side=LEFT, width=22)

            jazz_row = ttk.Frame(parent)
            jazz_row.pack(fill=X, padx=16, pady=(4, 0))
            from ui.ensembles import JAZZ_INSTRUMENTS
            self._field(jazz_row, "Jazz Band Instrument", "jazz_instrument",
                        widget="combobox", options=[""] + JAZZ_INSTRUMENTS,
                        side=LEFT, width=22)
            ttk.Label(jazz_row, text="Only if different in jazz band (e.g. Horn "
                                     "player on Guitar) — used for jazz rosters.",
                      font=("Segoe UI", 8), foreground=muted_fg(),
                      wraplength=260, justify=LEFT).pack(side=LEFT, padx=(8, 0))

        self._section(parent, "Contact Information")
        row2 = ttk.Frame(parent)
        row2.pack(fill=X, padx=16)
        self._field(row2, "Phone", "phone", side=LEFT, width=18)
        self._field(row2, "Student Email", "student_email", side=LEFT, width=28)

        row3 = ttk.Frame(parent)
        row3.pack(fill=X, padx=16, pady=(4, 0))
        self._field(row3, "Address", "address", side=LEFT, width=30)

        row4 = ttk.Frame(parent)
        row4.pack(fill=X, padx=16, pady=(4, 0))
        self._field(row4, "City", "city", side=LEFT, width=20)
        self._field(row4, "State", "state", side=LEFT, width=6)
        self._field(row4, "ZIP", "zip_code", side=LEFT, width=10)

        self._section(parent, "Parent / Guardian 1")
        row5 = ttk.Frame(parent)
        row5.pack(fill=X, padx=16)
        self._field(row5, "Name", "parent1_name", side=LEFT, width=24)
        self._field(row5, "Relation", "parent1_relation", widget="combobox",
                    options=RELATION_OPTIONS, side=LEFT, width=14)

        row6 = ttk.Frame(parent)
        row6.pack(fill=X, padx=16, pady=(4, 0))
        self._field(row6, "Phone", "parent1_phone", side=LEFT, width=18)
        self._field(row6, "Email", "parent1_email", side=LEFT, width=28)

        self._section(parent, "Parent / Guardian 2")
        row7 = ttk.Frame(parent)
        row7.pack(fill=X, padx=16)
        self._field(row7, "Name", "parent2_name", side=LEFT, width=24)
        self._field(row7, "Relation", "parent2_relation", widget="combobox",
                    options=RELATION_OPTIONS, side=LEFT, width=14)

        row8 = ttk.Frame(parent)
        row8.pack(fill=X, padx=16, pady=(4, 0))
        self._field(row8, "Phone", "parent2_phone", side=LEFT, width=18)
        self._field(row8, "Email", "parent2_email", side=LEFT, width=28)

        self._section(parent, "Notes")
        notes_frame = ttk.Frame(parent)
        notes_frame.pack(fill=X, padx=20, pady=(4, 0))
        self._notes_text = tk.Text(notes_frame, height=3, font=("Segoe UI", 9),
                                    relief="solid", bd=1, wrap=WORD, width=60)
        self._notes_text.pack(fill=X)

        # Set default year
        if "school_year" in self._vars:
            self._vars["school_year"].set(self.default_year)

    def _load(self, student_id: int):
        student = self.db.get_student(student_id)
        if not student:
            return
        if self._elementary:
            held = student["ensembles"] if "ensembles" in student.keys() else ""
            self._section_load(held)
            self._choir_load(held)
        for key, var in self._vars.items():
            val = student[key] if key in student.keys() else None
            var.set("" if val is None else str(val))
        # Multi-value checkbox groups.  Held values tick their box by class
        # IDENTITY ("Entry Band" ticks the "MS Band (Entry)" box); anything
        # that matches no option at all still gets its own box (ticked) rather
        # than being dropped on the next Save.
        import class_registry as cr
        for key in list(self._multi_vars):
            current = _csv_to_list(student[key] if key in student.keys() else "")
            is_class = key == "ensembles"
            for opt, var in self._multi_vars[key].items():
                var.set(any(cr.same_class(opt, c) for c in current)
                        if is_class else (opt in current))
            for opt in current:
                self._ensure_option(key, opt, value=True)
        notes = student["notes"] or ""
        self._notes_text.delete("1.0", "end")
        self._notes_text.insert("1.0", notes)

    def _collect(self) -> dict:
        data = {k: v.get().strip() for k, v in self._vars.items()}
        data["notes"] = self._notes_text.get("1.0", "end").strip()
        # Checkbox groups → comma-separated strings (order follows the option list)
        for key, holder in self._multi_vars.items():
            data[key] = ",".join(opt for opt, var in holder.items() if var.get())
        if self.student_id:
            data["is_active"] = 1
        return data

    def _validate(self, data: dict) -> bool:
        if not data.get("first_name") and not data.get("last_name"):
            Messagebox.show_warning("Student name is required.", title="Validation",
    parent=self.winfo_toplevel())
            return False
        if not data.get("school_year"):
            Messagebox.show_warning("School year is required.", title="Validation",
    parent=self.winfo_toplevel())
            return False
        return True

    def _mark_inactive(self):
        student = self.db.get_student(self.student_id)
        name = f"{student['first_name']} {student['last_name']}".strip() if student else "this student"
        answer = Messagebox.yesno(
            f"Mark {name} as inactive?\n\nThey will be hidden from the student list "
            f"unless 'Show Inactive' is checked.",
            title="Mark Inactive",
            parent=self,
        )
        if answer == "Yes":
            self.db.deactivate_student(self.student_id)
            self.destroy()

    def _reactivate(self):
        self.db.reactivate_student(self.student_id)
        self.destroy()

    SECTION_NUMBERS = (1, 2)

    def _section_labels(self):
        name = (self.site or {}).get("name") or ""
        return [f"{name}: Section {n}" for n in self.SECTION_NUMBERS]

    def _build_section_row(self, parent):
        """Which of this school's two sections the child is in.

        A radio, not tickboxes: the sections run back to back in the same room
        with the same teacher, so being in both is not a thing that can happen,
        and a pair of checkboxes quietly says it can.
        """
        ttk.Label(parent, text="Sections run back to back, so a child is in one "
                               "or the other.",
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(anchor=W, padx=16)
        row = ttk.Frame(parent)
        row.pack(anchor=W, padx=16, pady=(2, 0))
        self._section_var = tk.StringVar(value="")
        for i, label in enumerate(self._section_labels(), start=1):
            ttk.Radiobutton(row, text=f"Section {i}", value=label,
                            variable=self._section_var,
                            bootstyle=PRIMARY).pack(side=LEFT, padx=(0, 16))
        ttk.Button(row, text="Clear", bootstyle=(SECONDARY, OUTLINE),
                   command=lambda: self._section_var.set("")
                   ).pack(side=LEFT, padx=(6, 0))

    def _section_apply(self, ensembles_csv: str) -> str:
        """Put the chosen section in, and take any other section of this school
        out.  Their choir and any non-elementary class are left alone."""
        chosen = getattr(self, "_section_var", None)
        if chosen is None:
            return ensembles_csv
        mine = {lab.lower() for lab in self._section_labels()}
        parts = [p.strip() for p in (ensembles_csv or "").split(",") if p.strip()]
        parts = [p for p in parts if p.lower() not in mine]
        if chosen.get():
            parts.append(chosen.get())
        return ", ".join(parts)

    def _section_load(self, ensembles_csv: str):
        chosen = getattr(self, "_section_var", None)
        if chosen is None:
            return
        held = [p.strip() for p in (ensembles_csv or "").split(",") if p.strip()]
        held_l = {h.lower() for h in held}
        for label in self._section_labels():
            if label.lower() in held_l:
                chosen.set(label)
                return
        chosen.set("")

    def _build_choir_row(self, parent):
        """Choir sits alongside the instrument, never instead of it.

        It is not a class anywhere in the district -- a few schools run one
        before or after school -- so it is recorded as an ensemble this child
        is in as well, which means the section tickboxes, the class filters and
        the email lists all understand it with nothing added.
        """
        from ui.ensembles import choir_ensemble
        row = ttk.Frame(parent)
        row.pack(fill=X, padx=16, pady=(6, 0))
        self._choir_label = choir_ensemble(self.site.get("name") if self.site else "")
        self._choir_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row, text="Also sings in choir",
                        variable=self._choir_var,
                        bootstyle=PRIMARY).pack(side=LEFT)
        note = ("This school's choir is ticked for every child by default."
                if (self.site or {}).get("choir_default")
                else "Only if this school runs a choir and this child is in it.")
        ttk.Label(row, text=note, font=("Segoe UI", 8), foreground=muted_fg(),
                  wraplength=300, justify=LEFT).pack(side=LEFT, padx=(8, 0))
        if not self.student_id and (self.site or {}).get("choir_default"):
            self._choir_var.set(True)

    def _choir_apply(self, ensembles_csv: str) -> str:
        """Put the choir into (or take it out of) this child's ensembles."""
        label = getattr(self, "_choir_label", "")
        if not label:
            return ensembles_csv
        parts = [p.strip() for p in (ensembles_csv or "").split(",") if p.strip()]
        parts = [p for p in parts if p.lower() != label.lower()]
        if self._choir_var.get():
            parts.append(label)
        return ", ".join(parts)

    def _choir_load(self, ensembles_csv: str):
        label = getattr(self, "_choir_label", "")
        if not label or not hasattr(self, "_choir_var"):
            return
        parts = [p.strip().lower() for p in (ensembles_csv or "").split(",")]
        self._choir_var.set(label.lower() in parts)

    def _save(self):
        data = self._collect()
        if not self._validate(data):
            return
        # Choir rides along in ensembles rather than in a column of its own, so
        # it is folded in after the checkbox groups have had their say.
        if self._elementary:
            data["ensembles"] = self._section_apply(data.get("ensembles", ""))
            if hasattr(self, "_choir_var"):
                data["ensembles"] = self._choir_apply(data["ensembles"])
        if self.student_id:
            self.db.update_student(self.student_id, data)
        else:
            if self.site:
                data["site_id"] = self.site["id"]
            self.db.add_student(data)
        self.destroy()


def _short_group(group: str) -> str:
    """Tickbox text without the school name.  The window is already one school,
    so boxes all starting "Clyde Hill Elementary School:" say nothing."""
    return group.split(":", 1)[1].strip() if ":" in group else group


# ── Bulk ensemble / period assignment dialog ──────────────────────────────────

class _BulkAssignDialog(_ClassOptionsMixin, ttk.Toplevel):
    """Assign ensemble(s), class period(s), and/or instruments to many students
    at once."""

    _class_options_include_empty = True     # assigns a class, so offer empties

    def __init__(self, parent, db, student_ids, program_type, site_id=None):
        super().__init__(parent)
        self.db = db
        self.student_ids = student_ids
        self.program_type = program_type
        # Set when assigning within one school: its own sections, no class
        # periods, no second instrument.
        self.site_id = site_id
        self.applied = False    # caller clears the roster ticks only if we ran
        self._ens_vars = {}
        self._per_vars = {}
        self._mode_var = tk.StringVar(value="add")

        self.title("Assign Ensemble / Period")
        self.resizable(False, True)
        self.grab_set()
        self.lift()
        self._build()

        from ui.theme import fit_window
        fit_window(self, 480, 560)

    def _build(self):
        hdr = ttk.Frame(self, bootstyle=SUCCESS)
        hdr.pack(fill=X)
        ttk.Label(hdr, text="🏷️  Bulk Assign", font=("Segoe UI", 13, "bold"),
                  bootstyle=(INVERSE, SUCCESS)).pack(pady=12, padx=16, anchor=W)

        body = ttk.Frame(self)
        body.pack(fill=BOTH, expand=True, padx=18, pady=12)

        ttk.Label(body, text=f"Applying to {len(self.student_ids)} selected student(s).",
                  font=("Segoe UI", 9, "bold")).pack(anchor=W, pady=(0, 8))

        # Mode — the default (Add) is shown as a filled green toolbutton so it's
        # visually obvious we won't wipe out existing ensembles/periods.
        mode_box = tk.LabelFrame(body, text=" How to apply to ensembles & periods ",
                                 font=("Segoe UI", 9, "bold"), padx=8, pady=6)
        mode_box.pack(fill=X, pady=(0, 10))
        mode_row = ttk.Frame(mode_box)
        mode_row.pack(fill=X)
        ttk.Radiobutton(mode_row, text="➕  Add to existing  (default)", value="add",
                        variable=self._mode_var, bootstyle=(SUCCESS, OUTLINE, TOOLBUTTON),
                        width=24, command=self._update_mode_hint).pack(side=LEFT, padx=(0, 6))
        ttk.Radiobutton(mode_row, text="⚠  Replace existing", value="replace",
                        variable=self._mode_var, bootstyle=(WARNING, OUTLINE, TOOLBUTTON),
                        width=20, command=self._update_mode_hint).pack(side=LEFT)
        self._mode_hint = ttk.Label(mode_box, text="", font=("Segoe UI", 8),
                                    foreground=muted_fg(), wraplength=440, justify=LEFT)
        self._mode_hint.pack(anchor=W, pady=(6, 0))
        self._update_mode_hint()

        ttk.Label(body, text="Ensemble(s):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        ens_grid = ttk.Frame(body)
        ens_grid.pack(fill=X, pady=(2, 8))
        # Short labels on screen; the full canonical name is what gets written
        # to the records (identity-aware merge swaps out old spellings).
        from ui.ensembles import class_display_map
        opts = self._class_options()
        dmap = class_display_map(opts)
        for i, opt in enumerate(opts):
            v = tk.BooleanVar(value=False)
            self._ens_vars[opt] = v
            ttk.Checkbutton(ens_grid, text=dmap[opt], variable=v,
                            bootstyle=PRIMARY).grid(
                row=i // 3, column=i % 3, sticky=W, padx=6, pady=2)

        ttk.Label(body, text="Class Period(s):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        per_grid = ttk.Frame(body)
        per_grid.pack(fill=X, pady=(2, 8))
        for i, opt in enumerate(PERIOD_OPTIONS):
            v = tk.BooleanVar(value=False)
            self._per_vars[opt] = v
            ttk.Checkbutton(per_grid, text=opt, variable=v, bootstyle=PRIMARY).grid(
                row=0, column=i, sticky=W, padx=6, pady=2)

        ttk.Separator(body, orient=HORIZONTAL).pack(fill=X, pady=8)

        ttk.Label(body, text="Set instrument (optional — leave blank to skip):",
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(anchor=W)
        instr_row = ttk.Frame(body)
        instr_row.pack(fill=X, pady=(2, 0))
        instr_opts = [""] + instruments_for(self.program_type)
        self._prim_var = tk.StringVar()
        self._sec_var = tk.StringVar()
        pf = ttk.Frame(instr_row); pf.pack(side=LEFT, padx=(0, 10))
        ttk.Label(pf, text="Primary", font=("Segoe UI", 8)).pack(anchor=W)
        ttk.Combobox(pf, textvariable=self._prim_var, values=instr_opts,
                     state="readonly", width=20).pack()
        sf = ttk.Frame(instr_row); sf.pack(side=LEFT)
        ttk.Label(sf, text="Secondary", font=("Segoe UI", 8)).pack(anchor=W)
        ttk.Combobox(sf, textvariable=self._sec_var, values=instr_opts,
                     state="readonly", width=20).pack()

        btn = ttk.Frame(self)
        btn.pack(fill=X, padx=16, pady=12)
        ttk.Button(btn, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btn, text="Apply", bootstyle=SUCCESS,
                   command=self._apply).pack(side=RIGHT, padx=4)

    def _update_mode_hint(self):
        if self._mode_var.get() == "replace":
            self._mode_hint.config(
                text="⚠ Replace: clears each student's current ensembles/periods and sets "
                     "only what you tick below. Use with care.")
        else:
            self._mode_hint.config(
                text="Add (default): keeps each student's current ensembles/periods and adds "
                     "the ones you tick below — nothing is erased.")

    def _names(self, limit=4):
        """The first few students this will hit, then how many more."""
        out = []
        for sid in self.student_ids[:limit]:
            st = self.db.get_student(sid)
            if st:
                out.append(f"{st['first_name']} {st['last_name']}".strip())
        listed = ", ".join(n for n in out if n)
        rest = len(self.student_ids) - limit
        if rest > 0:
            listed += f", and {rest} more"
        return listed

    def _confirm_apply(self, n, ensembles, periods, prim, sec, replace):
        """Show what is about to happen, to how many, and to whom."""
        lines = []
        if ensembles:
            joined = ", ".join(ensembles)
            lines.append(f"Class: {'replace what they have with' if replace else 'add'} {joined}")
        if periods:
            joined = ", ".join(periods)
            lines.append(f"Period: {'replace with' if replace else 'add'} {joined}")
        if prim:
            lines.append(f"Instrument: {prim}, replacing what they have now")
        if sec:
            lines.append(f"Second instrument: {sec}, replacing what they have now")
        return Messagebox.yesno(
            f"Apply this to {n} students?\n\n"
            + "\n".join(lines)
            + f"\n\n{self._names()}",
            title=f"Apply to {n} students?", parent=self) == "Yes"

    def _apply(self):
        replace = self._mode_var.get() == "replace"
        ensembles = [o for o, v in self._ens_vars.items() if v.get()]
        periods = [o for o, v in self._per_vars.items() if v.get()]
        prim = self._prim_var.get().strip()
        sec = self._sec_var.get().strip()

        if not ensembles and not periods and not prim and not sec:
            Messagebox.show_warning("Nothing selected to assign.", title="Nothing to Apply",
                                    parent=self)
            return

        # Ticks carry over between assignments, and the roster this runs
        # against is invisible from here -- so a teacher who assigns the
        # violins, then ticks the violas without unticking, silently puts
        # every violin on viola too.  That happened.  The count and the first
        # few names are what catch it: six names when you ticked four is
        # obvious, and nothing else on screen would have told you.
        n = len(self.student_ids)
        overwrites = bool(prim or sec or replace)
        if n > 5 or (n > 1 and overwrites):
            if not self._confirm_apply(n, ensembles, periods, prim, sec,
                                       replace):
                return

        if ensembles:
            self.db.bulk_set_student_multi(self.student_ids, "ensembles", ensembles, replace)
        if periods:
            self.db.bulk_set_student_multi(self.student_ids, "class_periods", periods, replace)
        if prim:
            self.db.bulk_set_student_field(self.student_ids, "primary_instrument", prim)
        if sec:
            self.db.bulk_set_student_field(self.student_ids, "secondary_instrument", sec)

        self.applied = True
        Messagebox.show_info(
            f"Updated {len(self.student_ids)} student(s).", title="Done", parent=self)
        self.destroy()


# ── Email-list generator dialog ───────────────────────────────────────────────

class _EmailListDialog(_ClassOptionsMixin, ttk.Toplevel):
    """Build a copy/paste-ready email list filtered by ensemble / period /
    instrument, for students, parents, or everyone."""

    def __init__(self, parent, db, program_type, school_year, site_id=None):
        super().__init__(parent)
        self.db = db
        self.program_type = program_type
        self.school_year = school_year
        self.site_id = site_id
        self._group_vars = {}
        self.base_dir = getattr(parent, "base_dir", "")
        # The whole family by default, everywhere.  Writing to the students
        # alone was the old default and it is the wrong one twice over: a
        # 10-year-old has no school email at all, so a 5th grade send reached
        # nobody, and even at high school the people who need to know about a
        # concert are the ones driving to it.
        self._recip_var = tk.StringVar(value="everyone")
        self._ens_var = tk.StringVar(value="All")
        self._per_var = tk.StringVar(value="All")
        self._instr_var = tk.StringVar(value="All")

        self.title("Generate Email List")
        self.resizable(True, True)
        self.grab_set()
        self.lift()
        self._build()

        from ui.theme import fit_window
        fit_window(self, 560, 560)
        self._generate()

    def _build(self):
        hdr = ttk.Frame(self, bootstyle=INFO)
        hdr.pack(fill=X)
        ttk.Label(hdr, text="✉️  Email List", font=("Segoe UI", 13, "bold"),
                  bootstyle=(INVERSE, INFO)).pack(pady=12, padx=16, anchor=W)

        body = ttk.Frame(self)
        body.pack(fill=X, padx=18, pady=(12, 4))

        # Recipients
        ttk.Label(body, text="Recipients:", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, sticky=W, pady=4)
        rr = ttk.Frame(body); rr.grid(row=0, column=1, columnspan=3, sticky=W)
        for val, lbl in [("everyone", "Students & guardians"),
                         ("students", "Students only"),
                         ("parents", "Guardians only")]:
            ttk.Radiobutton(rr, text=lbl, value=val, variable=self._recip_var,
                            bootstyle=INFO, command=self._generate).pack(side=LEFT, padx=4)

        # Filters.  One school has two sections and maybe a choir -- a short,
        # known list -- so they are tickboxes: "Section 1 and the choir" is one
        # glance and two clicks, which a single-pick dropdown cannot express at
        # all.  No period filter, because elementary has no periods.
        if self.site_id:
            ttk.Label(body, text="Groups:", font=("Segoe UI", 9, "bold")).grid(
                row=1, column=0, sticky=NW, pady=4)
            gbox = ttk.Frame(body)
            gbox.grid(row=1, column=1, columnspan=3, sticky=W, padx=4)
            groups = self._site_groups()
            if not groups:
                ttk.Label(gbox, text="Nobody is on this school's roster yet.",
                          font=("Segoe UI", 8),
                          foreground=muted_fg()).pack(anchor=W)
            for i, g in enumerate(groups):
                v = tk.BooleanVar(value=True)      # everyone, until narrowed
                self._group_vars[g] = v
                ttk.Checkbutton(gbox, text=_short_group(g), variable=v,
                                bootstyle=INFO, command=self._generate
                                ).grid(row=i // 2, column=i % 2, sticky=W,
                                       padx=(0, 16), pady=1)
        else:
            ttk.Label(body, text="Ensemble:", font=("Segoe UI", 9, "bold")).grid(
                row=1, column=0, sticky=W, pady=4)
            ttk.Combobox(body, textvariable=self._ens_var, state="readonly", width=20,
                         values=["All"] + self._class_display_options()).grid(
                row=1, column=1, sticky=W, padx=4)
            self._ens_var.trace_add("write", lambda *_: self._generate())

            ttk.Label(body, text="Period:", font=("Segoe UI", 9, "bold")).grid(
                row=1, column=2, sticky=W, pady=4, padx=(10, 0))
            ttk.Combobox(body, textvariable=self._per_var, state="readonly", width=6,
                         values=["All"] + PERIOD_OPTIONS).grid(row=1, column=3, sticky=W, padx=4)
            self._per_var.trace_add("write", lambda *_: self._generate())

        # Instruments are a long list, so this is a multi-select rather than a
        # grid of eighteen tickboxes.  "The low brass" is trombone, baritone and
        # tuba; sending that message three separate times is how one of the
        # three ends up not being told.
        ttk.Label(body, text="Instrument:", font=("Segoe UI", 9, "bold")).grid(
            row=2, column=0, sticky=NW, pady=4)
        _instr = (fifth_grade_instruments(self.program_type) if self.site_id
                  else instruments_for(self.program_type))
        ibox = ttk.Frame(body)
        ibox.grid(row=2, column=1, columnspan=2, sticky=W, padx=4)
        isb = ttk.Scrollbar(ibox, orient=VERTICAL)
        self._instr_list = tk.Listbox(
            ibox, selectmode="extended", exportselection=False,
            height=min(6, max(3, len(_instr))), width=24,
            yscrollcommand=isb.set, font=("Segoe UI", 9),
            activestyle="none")
        isb.config(command=self._instr_list.yview)
        for name in _instr:
            self._instr_list.insert("end", name)
        self._instr_list.pack(side=LEFT)
        isb.pack(side=LEFT, fill=Y)
        self._instr_list.bind("<<ListboxSelect>>", lambda e: self._generate())

        side = ttk.Frame(body)
        side.grid(row=2, column=3, sticky=NW, padx=(10, 0), pady=4)
        ttk.Label(side, text="Ctrl+click for several.\n"
                             "None selected = every "
                             "instrument.",
                  font=("Segoe UI", 8), foreground=muted_fg(),
                  justify=LEFT).pack(anchor=W)
        ttk.Button(side, text="Clear", bootstyle=(SECONDARY, OUTLINE),
                   command=self._clear_instruments).pack(anchor=W, pady=(4, 0))
        ttk.Label(side, text="School year: " + (self.school_year or "all"),
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(anchor=W,
                                                                    pady=(6, 0))

        # Output
        out_frame = ttk.Frame(self)
        out_frame.pack(fill=BOTH, expand=True, padx=18, pady=(8, 4))
        self._count_lbl = ttk.Label(out_frame, text="", font=("Segoe UI", 9, "bold"))
        self._count_lbl.pack(anchor=W)
        ttk.Label(out_frame,
                  text="Paste into the To/Bcc field in Outlook (semicolon-separated):",
                  font=("Segoe UI", 8), foreground=muted_fg()).pack(anchor=W, pady=(2, 2))
        txt_wrap = ttk.Frame(out_frame)
        txt_wrap.pack(fill=BOTH, expand=True)
        sb = ttk.Scrollbar(txt_wrap, orient=VERTICAL)
        self._out = tk.Text(txt_wrap, height=8, font=("Consolas", 9), wrap=WORD,
                            yscrollcommand=sb.set, relief="solid", bd=1)
        sb.config(command=self._out.yview)
        sb.pack(side=RIGHT, fill=Y)
        self._out.pack(fill=BOTH, expand=True)

        btn = ttk.Frame(self)
        btn.pack(fill=X, padx=16, pady=12)
        ttk.Button(btn, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btn, text="📋 Copy to Clipboard", bootstyle=(INFO, OUTLINE),
                   command=self._copy).pack(side=RIGHT, padx=4)
        ttk.Button(btn, text="✉ Open in Outlook", bootstyle=SUCCESS,
                   command=self._send).pack(side=RIGHT, padx=4)

    def _site_groups(self):
        """The groups this school can offer: both sections, plus choir if used."""
        from ui.ensembles import site_groups
        site = self.db.get_site(self.site_id)
        return site_groups(self.db, site, self.school_year) if site else []

    def _clear_instruments(self):
        self._instr_list.selection_clear(0, "end")
        self._generate()

    def _chosen_instruments(self):
        """The ticked instruments, or None meaning every one of them.

        Selecting nothing is the same as selecting everything, which is what a
        teacher means when they have not narrowed it -- rather than an empty
        list that silently emails nobody.
        """
        picked = [self._instr_list.get(i)
                  for i in self._instr_list.curselection()]
        return picked or None

    def _generate(self):
        instr = self._chosen_instruments()
        if self.site_id:
            students = list(self.db.get_students_for_email(
                school_year=self.school_year, site_id=self.site_id,
                instrument=instr))
            wanted = {g.lower() for g, v in self._group_vars.items() if v.get()}
            picked = []
            for s in students:
                held = (s["ensembles"] if "ensembles" in s.keys() else "") or ""
                mine = {p.strip().lower() for p in held.split(",") if p.strip()}
                if wanted & mine:
                    picked.append(s)
            students = picked
        else:
            ens = self._ens_var.get()
            per = self._per_var.get()
            students = self.db.get_students_for_email(
                school_year=self.school_year,
                ensemble=None if ens == "All" else ens,
                period=None if per == "All" else per,
                instrument=instr,
            )

        recip = self._recip_var.get()
        emails = []
        seen = set()

        def _add(addr):
            addr = (addr or "").strip()
            if addr and "@" in addr and addr.lower() not in seen:
                seen.add(addr.lower())
                emails.append(addr)

        for s in students:
            if recip in ("everyone", "students"):
                _add(s["student_email"] if "student_email" in s.keys() else "")
            if recip in ("everyone", "parents"):
                _add(s["parent1_email"] if "parent1_email" in s.keys() else "")
                _add(s["parent2_email"] if "parent2_email" in s.keys() else "")

        self._out.delete("1.0", "end")
        self._out.insert("1.0", "; ".join(emails))
        self._count_lbl.config(
            text=f"{len(emails)} email address(es) from {len(students)} student(s)")

    def _copy(self):
        text = self._out.get("1.0", "end").strip()
        self.clipboard_clear()
        self.clipboard_append(text)
        Messagebox.show_info("Email list copied to clipboard.", title="Copied", parent=self)

    def _send(self):
        """Start a blank message to this list, everyone in BCC.

        There is no template here, so nothing is saved: this screen builds a
        list, and the words are written in Outlook."""
        from ui.email_compose import open_message
        text = self._out.get("1.0", "end").strip()
        addrs = [a.strip() for a in text.replace(",", ";").split(";") if a.strip()]
        if not addrs:
            Messagebox.show_info("Generate a list first.",
                                 title="Nothing to Send", parent=self)
            return
        open_message(self, self.base_dir, "", "", addrs)
