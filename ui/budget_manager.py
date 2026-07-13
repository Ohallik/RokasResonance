"""
ui/budget_manager.py - Budget & student-fee tracking.

Tracks expenses and income for a school year (Jul 1 – Jun 30), auto-including
instrument-repair costs, split by funding source (Building / ASB / Boosters /
Other) with customizable categories, plus per-student fee tracking (polo shirts,
instrument rentals) with an "owes a fee AND has an instrument out" nudge list.
"""

import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from datetime import datetime

from ui.ensembles import ensembles_for
from ui.names import display_last_first


def _money(v):
    try:
        return f"${float(v or 0):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


# Funding source → curricular vs extracurricular use.  Building/department money
# is curricular (during-the-day classes); ASB & Boosters fund extracurriculars.
FUNDING_CLASS = {
    "Building": "Curricular",
    "ASB": "Extracurricular",
    "Boosters": "Extracurricular",
    "Other": "Other",
}


def funding_class(src):
    return FUNDING_CLASS.get(src or "", "Other")


class BudgetManager(ttk.Frame):
    def __init__(self, parent, db, base_dir, program_type="band"):
        super().__init__(parent)
        self.db = db
        self.base_dir = base_dir
        self.program_type = program_type
        self._year_var = tk.StringVar()
        self._src_var = tk.StringVar(value="All")
        self._kind_var = tk.StringVar(value="All")
        self._build()
        self._populate_years()
        self.refresh()

    # ─────────────────────────────────────────────────────────── Build UI ──────

    def _build(self):
        hdr = ttk.Frame(self, bootstyle=SUCCESS)
        hdr.pack(fill=X)
        ttk.Label(hdr, text="💵  Budget", font=("Segoe UI", 15, "bold"),
                  bootstyle=(INVERSE, SUCCESS)).pack(pady=10, padx=16, anchor=W)

        # Toolbar
        tb = ttk.Frame(self, bootstyle=LIGHT)
        tb.pack(fill=X)
        ttk.Button(tb, text="➕ Add", bootstyle=SUCCESS,
                   command=self._add_txn).pack(side=LEFT, padx=6, pady=6)
        ttk.Button(tb, text="🚌 Field Trip", bootstyle=(SUCCESS, OUTLINE),
                   command=self._add_field_trip).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(tb, text="✏️ Edit", bootstyle=PRIMARY,
                   command=self._edit_txn).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(tb, text="🗑️ Delete", bootstyle=DANGER,
                   command=self._delete_txn).pack(side=LEFT, padx=2, pady=6)
        ttk.Separator(tb, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)
        ttk.Button(tb, text="🏷️ Categories", bootstyle=SECONDARY,
                   command=self._manage_categories).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(tb, text="🎽 Student Fees", bootstyle=INFO,
                   command=self._open_fees).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(tb, text="📊 Export Excel", bootstyle=(SECONDARY, OUTLINE),
                   command=self._export).pack(side=LEFT, padx=2, pady=6)

        ttk.Separator(tb, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)
        ttk.Label(tb, text="School Year:").pack(side=LEFT, padx=(0, 4))
        self._year_combo = ttk.Combobox(tb, textvariable=self._year_var,
                                        state="readonly", width=12)
        self._year_combo.pack(side=LEFT)
        self._year_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        # Filter bar
        fb = ttk.Frame(self)
        fb.pack(fill=X, padx=10, pady=(6, 2))
        ttk.Label(fb, text="Funding source:").pack(side=LEFT, padx=(0, 4))
        ttk.Combobox(fb, textvariable=self._src_var, state="readonly", width=12,
                     values=["All"] + self.db.FUNDING_SOURCES).pack(side=LEFT, padx=(0, 12))
        self._src_var.trace_add("write", lambda *_: self._apply())
        ttk.Label(fb, text="Type:").pack(side=LEFT, padx=(0, 4))
        ttk.Combobox(fb, textvariable=self._kind_var, state="readonly", width=10,
                     values=["All", "Expense", "Income"]).pack(side=LEFT, padx=(0, 12))
        self._kind_var.trace_add("write", lambda *_: self._apply())
        ttk.Label(fb, text="Use:").pack(side=LEFT, padx=(0, 4))
        self._use_var = tk.StringVar(value="All")
        ttk.Combobox(fb, textvariable=self._use_var, state="readonly", width=14,
                     values=["All", "Curricular", "Extracurricular", "Other"]).pack(side=LEFT)
        self._use_var.trace_add("write", lambda *_: self._apply())

        # Content: transactions + summary side panel
        paned = ttk.Panedwindow(self, orient=HORIZONTAL)
        paned.pack(fill=BOTH, expand=True, padx=10, pady=8)

        left = ttk.Frame(paned)
        paned.add(left, weight=3)
        cols = ("date", "type", "category", "desc", "source", "use", "student", "amount")
        sb = ttk.Scrollbar(left, orient=VERTICAL)
        self.tree = ttk.Treeview(left, columns=cols, show="headings",
                                 yscrollcommand=sb.set, selectmode="browse", bootstyle=SUCCESS)
        sb.config(command=self.tree.yview)
        sb.pack(side=RIGHT, fill=Y)
        self.tree.pack(fill=BOTH, expand=True)
        heads = {"date": "Date", "type": "Type", "category": "Category", "desc": "Description",
                 "source": "Funding", "use": "Use", "student": "Student", "amount": "Amount"}
        widths = {"date": 90, "type": 68, "category": 130, "desc": 220, "source": 82,
                  "use": 96, "student": 120, "amount": 96}
        for c in cols:
            self.tree.heading(c, text=heads[c], anchor=W)
            self.tree.column(c, width=widths[c], anchor=(E if c == "amount" else W),
                             stretch=c == "desc")
        self.tree.tag_configure("income", foreground="#1a7a1a")
        self.tree.tag_configure("repair", foreground="#666")
        self.tree.bind("<Double-1>", lambda e: self._edit_txn())

        right = ttk.Frame(paned, width=260)
        paned.add(right, weight=1)
        ttk.Label(right, text="Summary", font=("Segoe UI", 11, "bold"),
                  bootstyle=SUCCESS).pack(anchor=W, padx=8, pady=(4, 6))
        self._summary_frame = ttk.Frame(right)
        self._summary_frame.pack(fill=BOTH, expand=True, padx=8)

    # ─────────────────────────────────────────────────────────── Data ──────────

    def _populate_years(self):
        years = self.db.get_budget_school_years()
        self._year_combo["values"] = years
        if not self._year_var.get():
            self._year_var.set(self.db.get_budget_default_year())

    def refresh(self):
        self._all = self.db.get_budget_transactions(self._year_var.get() or
                                                     self.db.current_school_year())
        self._apply()
        self._populate_years()

    def _apply(self):
        src = self._src_var.get()
        kind = self._kind_var.get().lower()
        use = self._use_var.get()
        rows = self._all
        if src != "All":
            rows = [r for r in rows if (r.get("funding_source") or "") == src]
        if kind != "all":
            rows = [r for r in rows if (r.get("kind") or "") == kind]
        if use != "All":
            rows = [r for r in rows if funding_class(r.get("funding_source")) == use]
        self._rows = rows
        self._fill(rows)
        self._build_summary(rows)

    def _fill(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, r in enumerate(rows):
            tags = []
            if r.get("kind") == "income":
                tags.append("income")
            if r.get("source") == "repair":
                tags.append("repair")
            amt = _money(r.get("amount"))
            if r.get("kind") == "expense":
                amt = "(" + amt + ")"
            self.tree.insert("", "end", iid=str(i), tags=tuple(tags), values=(
                r.get("txn_date") or "",
                (r.get("kind") or "").title(),
                r.get("category") or "",
                (r.get("description") or "") + ("  [auto]" if r.get("source") == "repair" else ""),
                r.get("funding_source") or "",
                funding_class(r.get("funding_source")),
                r.get("student_name") or "",
                amt,
            ))

    def _build_summary(self, rows):
        for w in self._summary_frame.winfo_children():
            w.destroy()
        by_src = {}
        tot_exp = tot_inc = 0.0
        for r in rows:
            src = r.get("funding_source") or "Other"
            d = by_src.setdefault(src, {"expense": 0.0, "income": 0.0})
            amt = float(r.get("amount") or 0)
            d[r.get("kind") or "expense"] += amt
            if r.get("kind") == "income":
                tot_inc += amt
            else:
                tot_exp += amt

        def line(parent, label, val, bold=False, color=None):
            row = ttk.Frame(parent)
            row.pack(fill=X, pady=1)
            f = ("Segoe UI", 9, "bold") if bold else ("Segoe UI", 9)
            ttk.Label(row, text=label, font=f).pack(side=LEFT)
            lbl = ttk.Label(row, text=val, font=f)
            if color:
                lbl.config(foreground=color)
            lbl.pack(side=RIGHT)

        for src in self.db.FUNDING_SOURCES:
            if src not in by_src:
                continue
            d = by_src[src]
            box = ttk.Labelframe(self._summary_frame, text=f" {src} ")
            box.pack(fill=X, pady=(0, 6))
            line(box, "Income", _money(d["income"]), color="#1a7a1a")
            line(box, "Expenses", _money(d["expense"]), color="#b00000")
            line(box, "Net", _money(d["income"] - d["expense"]), bold=True)

        # Curricular vs Extracurricular rollup
        by_use = {}
        for r in rows:
            u = funding_class(r.get("funding_source"))
            d = by_use.setdefault(u, {"expense": 0.0, "income": 0.0})
            d[r.get("kind") or "expense"] += float(r.get("amount") or 0)
        if by_use:
            ubox = ttk.Labelframe(self._summary_frame, text=" By Use ")
            ubox.pack(fill=X, pady=(4, 6))
            for u in ("Curricular", "Extracurricular", "Other"):
                if u in by_use:
                    d = by_use[u]
                    line(ubox, u, f"net {_money(d['income'] - d['expense'])}", bold=True)
                    line(ubox, f"   in / out",
                         f"{_money(d['income'])} / {_money(d['expense'])}")

        ttk.Separator(self._summary_frame).pack(fill=X, pady=4)
        line(self._summary_frame, "TOTAL Income", _money(tot_inc), bold=True, color="#1a7a1a")
        line(self._summary_frame, "TOTAL Expenses", _money(tot_exp), bold=True, color="#b00000")
        line(self._summary_frame, "NET", _money(tot_inc - tot_exp), bold=True)

    def _selected(self):
        sel = self.tree.selection()
        if not sel:
            return None
        return self._rows[int(sel[0])]

    # ─────────────────────────────────────────────────────────── Actions ───────

    def _add_txn(self):
        _TxnDialog(self.winfo_toplevel(), self.db, self._year_var.get(),
                   on_done=self.refresh)

    def _add_field_trip(self):
        _FieldTripDialog(self.winfo_toplevel(), self.db, self.base_dir,
                         self._year_var.get(), program_type=self.program_type,
                         on_done=self.refresh)

    def _edit_txn(self):
        r = self._selected()
        if not r:
            Messagebox.show_warning("Select a transaction.", title="No Selection", parent=self)
            return
        if r.get("source") == "repair":
            Messagebox.show_info(
                "Repair costs are pulled automatically from the Repair Center. "
                "Edit the repair there to change its actual cost.",
                title="Auto-linked", parent=self)
            return
        if r.get("source") == "fee":
            Messagebox.show_info(
                "Collected student fees are pulled automatically from Student "
                "Fees. Open “🎽 Student Fees” to change or un-mark a payment.",
                title="Auto-linked", parent=self)
            return
        _TxnDialog(self.winfo_toplevel(), self.db, self._year_var.get(),
                   txn=r, on_done=self.refresh)

    def _delete_txn(self):
        r = self._selected()
        if not r:
            return
        if r.get("source") == "repair":
            Messagebox.show_info("Delete the repair record in the Repair Center instead.",
                                 title="Auto-linked", parent=self)
            return
        if r.get("source") == "fee":
            Messagebox.show_info("Remove or un-mark this fee in “🎽 Student Fees”.",
                                 title="Auto-linked", parent=self)
            return
        if Messagebox.yesno("Delete this transaction?", title="Confirm", parent=self) == "Yes":
            self.db.delete_budget_transaction(r["id"])
            self.refresh()

    def _manage_categories(self):
        _CategoryDialog(self.winfo_toplevel(), self.db, on_done=self.refresh)

    def _open_fees(self):
        _FeesDialog(self.winfo_toplevel(), self.db, self.program_type, self._year_var.get())

    def _export(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required.", title="Missing Dependency", parent=self)
            return
        rows = self._rows
        if not rows:
            Messagebox.show_info("Nothing to export for this filter.", title="Nothing to Export",
                                 parent=self)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.freeze_panes = "A2"
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="2E7D32")
        border = Border(*[Side(style="thin", color="CCCCCC")] * 4)
        money = '"$"#,##0.00'
        headers = ["Date", "Type", "Category", "Description", "Funding Source",
                   "Use", "Student", "Amount"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        r = 2
        for t in rows:
            for c, val in enumerate([
                t.get("txn_date") or "", (t.get("kind") or "").title(), t.get("category") or "",
                t.get("description") or "", t.get("funding_source") or "",
                funding_class(t.get("funding_source")),
                t.get("student_name") or "", float(t.get("amount") or 0),
            ], 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                if c == 8:
                    cell.number_format = money
            r += 1
        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Funding Source", "Use", "Income", "Expenses", "Net"])
        for cell in ws2[1]:
            cell.font = hdr_font; cell.fill = hdr_fill
        summ = {}
        for t in rows:
            s = summ.setdefault(t.get("funding_source") or "Other", [0.0, 0.0])
            if t.get("kind") == "income":
                s[0] += float(t.get("amount") or 0)
            else:
                s[1] += float(t.get("amount") or 0)
        for src, (inc, exp) in summ.items():
            ws2.append([src, funding_class(src), inc, exp, inc - exp])
        for row_cells in ws2.iter_rows(min_row=2, min_col=3, max_col=5):
            for cell in row_cells:
                cell.number_format = money
        for col, w in zip(range(1, 9), [12, 9, 16, 32, 14, 14, 16, 12]):
            ws.column_dimensions[get_column_letter(col)].width = w

        from tkinter import filedialog
        import datetime as _d
        tag = self._src_var.get() if self._src_var.get() != "All" else "All"
        path = filedialog.asksaveasfilename(
            title="Save Budget Export", parent=self.winfo_toplevel(),
            defaultextension=".xlsx",
            initialfile=f"Budget_{self._year_var.get()}_{tag}_{_d.date.today().isoformat()}.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb.save(path)
        except Exception as e:
            Messagebox.show_error(f"Could not save:\n{e}", title="Save Error", parent=self)
            return
        if Messagebox.yesno("Exported. Open it now?", title="Export Complete", parent=self) == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)


# ── Add / edit transaction ─────────────────────────────────────────────────────

class _TxnDialog(ttk.Toplevel):
    def __init__(self, parent, db, school_year, txn=None, on_done=None):
        super().__init__(parent)
        self.db = db
        self.school_year = school_year
        self.txn = txn
        self.on_done = on_done
        self.title("Edit Transaction" if txn else "Add Transaction")
        self.resizable(False, False)
        self.grab_set()
        self.lift()
        self._build()
        from ui.theme import fit_window
        fit_window(self, 440, 500)

    def _build(self):
        m = ttk.Frame(self)
        m.pack(fill=BOTH, expand=True, padx=20, pady=16)

        self._kind = tk.StringVar(value=(self.txn or {}).get("kind", "expense"))
        krow = ttk.Frame(m); krow.pack(fill=X, pady=(0, 6))
        ttk.Label(krow, text="Type:", font=("Segoe UI", 9, "bold")).pack(side=LEFT, padx=(0, 8))
        ttk.Radiobutton(krow, text="Expense", value="expense", variable=self._kind,
                        bootstyle=(DANGER, OUTLINE, TOOLBUTTON), width=10,
                        command=self._reload_cats).pack(side=LEFT, padx=2)
        ttk.Radiobutton(krow, text="Income", value="income", variable=self._kind,
                        bootstyle=(SUCCESS, OUTLINE, TOOLBUTTON), width=10,
                        command=self._reload_cats).pack(side=LEFT, padx=2)

        def field(label):
            ttk.Label(m, text=label, font=("Segoe UI", 9, "bold")).pack(anchor=W, pady=(6, 0))

        field("Date (YYYY-MM-DD)")
        self._date = tk.StringVar(value=(self.txn or {}).get("txn_date")
                                  or datetime.today().strftime("%Y-%m-%d"))
        ttk.Entry(m, textvariable=self._date, width=16).pack(anchor=W)

        field("Category")
        self._cat = tk.StringVar(value=(self.txn or {}).get("category", ""))
        self._cat_combo = ttk.Combobox(m, textvariable=self._cat, width=34)
        self._cat_combo.pack(anchor=W)
        self._reload_cats()

        field("Description")
        self._desc = tk.StringVar(value=(self.txn or {}).get("description", ""))
        ttk.Entry(m, textvariable=self._desc, width=44).pack(anchor=W)

        field("Amount ($)")
        self._amt = tk.StringVar(value=str((self.txn or {}).get("amount", "") or ""))
        ttk.Entry(m, textvariable=self._amt, width=14).pack(anchor=W)

        field("Funding Source")
        self._src = tk.StringVar(value=(self.txn or {}).get("funding_source", "Building"))
        ttk.Combobox(m, textvariable=self._src, state="readonly", width=16,
                     values=self.db.FUNDING_SOURCES).pack(anchor=W)

        field("Student (optional)")
        self._students = [(display_last_first(s), s["id"])
                          for s in self.db.get_current_roster()]
        self._stu = tk.StringVar()
        cur_sid = (self.txn or {}).get("student_id")
        if cur_sid:
            for nm, sid in self._students:
                if sid == cur_sid:
                    self._stu.set(nm)
        ttk.Combobox(m, textvariable=self._stu, width=34,
                     values=[""] + [n for n, _ in self._students]).pack(anchor=W)

        field("Notes")
        self._notes = tk.Text(m, height=2, width=44, relief="solid", bd=1)
        self._notes.pack(anchor=W, fill=X)
        self._notes.insert("1.0", (self.txn or {}).get("notes", "") or "")

        btn = ttk.Frame(self)
        btn.pack(fill=X, padx=16, pady=12)
        ttk.Button(btn, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btn, text="Save", bootstyle=SUCCESS, command=self._save).pack(side=RIGHT, padx=4)

    def _reload_cats(self):
        cats = [c["name"] for c in self.db.get_budget_categories(self._kind.get())]
        self._cat_combo["values"] = cats

    def _save(self):
        try:
            amount = float(str(self._amt.get()).replace("$", "").replace(",", "") or 0)
        except ValueError:
            Messagebox.show_warning("Amount must be a number.", title="Invalid", parent=self)
            return
        sid = None
        stu = self._stu.get().strip()
        for nm, _id in self._students:
            if nm == stu:
                sid = _id
                break
        data = {
            "txn_date": self._date.get().strip(),
            "description": self._desc.get().strip(),
            "category": self._cat.get().strip(),
            "kind": self._kind.get(),
            "amount": amount,
            "funding_source": self._src.get().strip(),
            "student_id": sid,
            "notes": self._notes.get("1.0", "end").strip(),
        }
        if self.txn and self.txn.get("id"):
            self.db.update_budget_transaction(self.txn["id"], data)
        else:
            self.db.add_budget_transaction(data)
        if self.on_done:
            self.on_done()
        self.destroy()


# ── Field trip cost sheet ───────────────────────────────────────────────────────

DEFAULT_SUB_RATES = {"4hr": 212.0, "5hr": 266.0, "full": 354.0}


class _FieldTripDialog(ttk.Toplevel):
    """Mirror the district field-trip cost form: entry fee, transportation, a
    substitute rate (editable, remembered), food, other → one Field Trip expense."""

    def __init__(self, parent, db, base_dir, school_year, program_type="band", on_done=None):
        super().__init__(parent)
        self.db = db
        self.base_dir = base_dir
        self.school_year = school_year
        self.program_type = program_type
        self.on_done = on_done
        self._rates = self._load_rates()
        self._trip_ens_vars = {}
        self.title("Field Trip Costs")
        self.resizable(True, True)
        self.grab_set()
        self.lift()
        self._build()
        self._recompute()
        from ui.theme import fit_window
        fit_window(self, 520, 760)

    def _load_rates(self):
        try:
            from ui.settings_dialog import load_settings
            r = (load_settings(self.base_dir).get("budget") or {}).get("sub_rates") or {}
        except Exception:
            r = {}
        return {k: float(r.get(k, DEFAULT_SUB_RATES[k])) for k in DEFAULT_SUB_RATES}

    def _save_rates(self):
        try:
            from ui.settings_dialog import load_settings, save_settings
            s = load_settings(self.base_dir)
            s.setdefault("budget", {})["sub_rates"] = {
                k: float(self._rate_vars[k].get() or 0) for k in DEFAULT_SUB_RATES}
            save_settings(self.base_dir, s)
        except Exception:
            pass

    def _build(self):
        hdr = ttk.Frame(self, bootstyle=SUCCESS)
        hdr.pack(fill=X)
        ttk.Label(hdr, text="🚌  Field Trip Costs", font=("Segoe UI", 13, "bold"),
                  bootstyle=(INVERSE, SUCCESS)).pack(pady=10, padx=16, anchor=W)

        # ── Buttons + total pinned at the bottom (always visible) ──────────
        btn = ttk.Frame(self)
        btn.pack(fill=X, padx=16, pady=10, side=BOTTOM)
        ttk.Button(btn, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(btn, text="Save Trip", bootstyle=SUCCESS,
                   command=self._save).pack(side=RIGHT, padx=4)
        trow = ttk.Frame(self)
        trow.pack(fill=X, padx=20, pady=(6, 0), side=BOTTOM)
        ttk.Label(trow, text="Total Trip Cost:", font=("Segoe UI", 11, "bold")).pack(side=LEFT)
        self._total_lbl = ttk.Label(trow, text="$0.00", font=("Segoe UI", 11, "bold"),
                                    foreground="#b00000")
        self._total_lbl.pack(side=RIGHT)
        ttk.Separator(self, orient=HORIZONTAL).pack(fill=X, side=BOTTOM)

        # ── Scrollable body ────────────────────────────────────────────────
        container = ttk.Frame(self)
        container.pack(fill=BOTH, expand=True)
        canvas = tk.Canvas(container, highlightthickness=0)
        sbar = ttk.Scrollbar(container, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        m = ttk.Frame(canvas)
        _win = canvas.create_window((0, 0), window=m, anchor=NW)
        m.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(_win, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))
        self.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>") if e.widget is self else None)
        m = ttk.Frame(m)   # inner padding frame
        m.pack(fill=BOTH, expand=True, padx=20, pady=12)

        def field(label, default=""):
            ttk.Label(m, text=label, font=("Segoe UI", 9, "bold")).pack(anchor=W, pady=(6, 0))
            v = tk.StringVar(value=default)
            ttk.Entry(m, textvariable=v, width=40).pack(anchor=W)
            return v

        self._trip = field("Trip / Event Name *")
        self._date = field("Date (YYYY-MM-DD)", datetime.today().strftime("%Y-%m-%d"))

        ttk.Label(m, text="Funding Source", font=("Segoe UI", 9, "bold")).pack(anchor=W, pady=(6, 0))
        self._src = tk.StringVar(value="ASB")   # trips are usually extracurricular
        ttk.Combobox(m, textvariable=self._src, state="readonly", width=16,
                     values=self.db.FUNDING_SOURCES).pack(anchor=W)

        ttk.Separator(m, orient=HORIZONTAL).pack(fill=X, pady=(10, 6))

        # Money components
        def money_field(label):
            row = ttk.Frame(m); row.pack(fill=X, pady=2)
            ttk.Label(row, text=label, font=("Segoe UI", 9), width=26, anchor=W).pack(side=LEFT)
            ttk.Label(row, text="$").pack(side=LEFT)
            v = tk.StringVar(value="")
            e = ttk.Entry(row, textvariable=v, width=12)
            e.pack(side=LEFT)
            v.trace_add("write", lambda *_: self._recompute())
            return v

        self._entry_fee = money_field("Entry fee / participation:")
        self._transport = money_field("Transportation:")

        # Substitute — choose one editable rate (or none)
        sub_box = ttk.Labelframe(m, text=" Substitute teacher (check one) ")
        sub_box.pack(fill=X, pady=(8, 4))
        self._sub_choice = tk.StringVar(value="none")
        self._rate_vars = {}
        labels = {"4hr": "4 hrs", "5hr": "5 hrs", "full": "full day"}
        ttk.Radiobutton(sub_box, text="None", value="none", variable=self._sub_choice,
                        command=self._recompute).grid(row=0, column=0, sticky=W, padx=6, pady=2)
        for i, key in enumerate(("4hr", "5hr", "full"), 1):
            rv = tk.StringVar(value=f"{self._rates[key]:.0f}")
            self._rate_vars[key] = rv
            rv.trace_add("write", lambda *_: self._recompute())
            ttk.Radiobutton(sub_box, text=labels[key], value=key, variable=self._sub_choice,
                            command=self._recompute).grid(row=i, column=0, sticky=W, padx=6, pady=2)
            ttk.Label(sub_box, text="$").grid(row=i, column=1, sticky=E)
            ttk.Entry(sub_box, textvariable=rv, width=10).grid(row=i, column=2, sticky=W, padx=(0, 8))
        ttk.Label(sub_box, text="Rates are editable and remembered for next time.",
                  font=("Segoe UI", 8), foreground="#888").grid(
            row=4, column=0, columnspan=3, sticky=W, padx=6, pady=(2, 4))

        self._food = money_field("Food:")
        self._other = money_field("Other:")

        # ── Who's going + student charge ────────────────────────────────────
        ttk.Separator(m, orient=HORIZONTAL).pack(fill=X, pady=(10, 6))
        ttk.Label(m, text="Attending ensemble(s):", font=("Segoe UI", 9, "bold")).pack(anchor=W)
        ens_grid = ttk.Frame(m); ens_grid.pack(fill=X, pady=(2, 4))
        for i, opt in enumerate(ensembles_for(self.program_type)):
            v = tk.BooleanVar(value=False)
            self._trip_ens_vars[opt] = v
            ttk.Checkbutton(ens_grid, text=opt, variable=v, bootstyle=SUCCESS,
                            command=self._recompute).grid(row=i // 3, column=i % 3,
                                                          sticky=W, padx=6, pady=2)

        charge = ttk.Labelframe(m, text=" Charge students a trip fee ")
        charge.pack(fill=X, pady=(6, 4))
        self._charge_mode = tk.StringVar(value="none")
        ttk.Radiobutton(charge, text="Don't charge students", value="none",
                        variable=self._charge_mode, command=self._recompute).pack(anchor=W, padx=6, pady=1)
        r2 = ttk.Frame(charge); r2.pack(fill=X, padx=6, pady=1)
        ttk.Radiobutton(r2, text="Flat fee per student:  $", value="flat",
                        variable=self._charge_mode, command=self._recompute).pack(side=LEFT)
        self._per_student = tk.StringVar(value="")
        e = ttk.Entry(r2, textvariable=self._per_student, width=8)
        e.pack(side=LEFT)
        self._per_student.trace_add("write", lambda *_: self._recompute())
        ttk.Radiobutton(charge, text="Split total trip cost evenly among attendees",
                        value="split", variable=self._charge_mode,
                        command=self._recompute).pack(anchor=W, padx=6, pady=1)
        self._charge_info = ttk.Label(charge, text="", font=("Segoe UI", 8), foreground="#555")
        self._charge_info.pack(anchor=W, padx=6, pady=(2, 4))

    def _num(self, var):
        try:
            return float(str(var.get()).replace("$", "").replace(",", "") or 0)
        except ValueError:
            return 0.0

    def _sub_cost(self):
        c = self._sub_choice.get()
        return self._num(self._rate_vars[c]) if c in self._rate_vars else 0.0

    def _tagged_ensembles(self):
        return [e for e, v in self._trip_ens_vars.items() if v.get()]

    def _attendees(self):
        """Current active students in any tagged ensemble."""
        tagged = self._tagged_ensembles()
        if not tagged:
            return []
        out = []
        for s in self.db.get_current_roster():
            ens = [x.strip() for x in ((s.get("ensembles") or "").split(",")) if x.strip()]
            if any(t in ens for t in tagged):
                out.append(s)
        return out

    def _per_student_fee(self, total, n):
        mode = self._charge_mode.get()
        if mode == "flat":
            return round(self._num(self._per_student), 2)
        if mode == "split" and n > 0:
            return round(total / n, 2)
        return 0.0

    def _recompute(self):
        total = (self._num(self._entry_fee) + self._num(self._transport) +
                 self._sub_cost() + self._num(self._food) + self._num(self._other))
        self._total_lbl.config(text=_money(total))
        n = len(self._attendees())
        mode = self._charge_mode.get()
        if mode == "none":
            self._charge_info.config(text=f"{n} student(s) in the selected ensemble(s).")
        else:
            per = self._per_student_fee(total, n)
            self._charge_info.config(
                text=f"{n} student(s) → {_money(per)} each  "
                     f"(total collected ≈ {_money(per * n)}).")
        return total

    def _save(self):
        if not self._trip.get().strip():
            Messagebox.show_warning("Enter the trip / event name.", title="Required", parent=self)
            return
        total = self._recompute()
        if total <= 0:
            Messagebox.show_warning("Enter at least one cost.", title="No Costs", parent=self)
            return
        self._save_rates()
        tagged = self._tagged_ensembles()
        parts = []
        for label, var in [("Entry/participation", self._entry_fee),
                           ("Transportation", self._transport),
                           ("Food", self._food), ("Other", self._other)]:
            if self._num(var):
                parts.append(f"{label}: {_money(self._num(var))}")
        if self._sub_choice.get() != "none":
            parts.append(f"Substitute ({self._sub_choice.get()}): {_money(self._sub_cost())}")
        if tagged:
            parts.append("Ensembles: " + ", ".join(tagged))
        self.db.add_budget_transaction({
            "txn_date": self._date.get().strip(),
            "description": f"Field Trip: {self._trip.get().strip()}",
            "category": "Field Trip",
            "kind": "expense",
            "amount": total,
            "funding_source": self._src.get().strip(),
            "student_id": None,
            "notes": "; ".join(parts),
        })

        # Assign per-student trip fees to the attending ensembles
        msg = "Field trip expense saved."
        if self._charge_mode.get() != "none":
            attendees = self._attendees()
            per = self._per_student_fee(total, len(attendees))
            if not tagged:
                Messagebox.show_warning(
                    "Tick at least one attending ensemble to charge students.",
                    title="No Ensemble", parent=self)
                return
            if per > 0 and attendees:
                fee_name = f"Trip: {self._trip.get().strip()}"
                years = self.db.get_school_years()
                roster_year = years[0] if years else self.db.current_school_year()
                self.db.ensure_fee_type(fee_name, per)
                for s in attendees:
                    self.db.ensure_student_fee(s["id"], fee_name, roster_year, per)
                msg += (f"\n\nCharged {_money(per)} to {len(attendees)} student(s) as "
                        f"'{fee_name}'.\nWaive or remove non-attendees in Budget ▸ "
                        f"Student Fees (choose that fee).")
        Messagebox.show_info(msg, title="Saved", parent=self)
        if self.on_done:
            self.on_done()
        self.destroy()


# ── Manage categories ──────────────────────────────────────────────────────────

class _CategoryDialog(ttk.Toplevel):
    def __init__(self, parent, db, on_done=None):
        super().__init__(parent)
        self.db = db
        self.on_done = on_done
        self.title("Budget Categories")
        self.grab_set()
        self.lift()
        self._build()
        self._reload()
        from ui.theme import fit_window
        fit_window(self, 420, 460)

    def _build(self):
        ttk.Label(self, text="Categories", font=("Segoe UI", 12, "bold"),
                  bootstyle=SUCCESS).pack(anchor=W, padx=14, pady=(12, 4))
        add = ttk.Frame(self); add.pack(fill=X, padx=14, pady=4)
        self._name = tk.StringVar()
        ttk.Entry(add, textvariable=self._name, width=22).pack(side=LEFT)
        self._kind = tk.StringVar(value="expense")
        ttk.Combobox(add, textvariable=self._kind, values=["expense", "income"],
                     state="readonly", width=9).pack(side=LEFT, padx=4)
        ttk.Button(add, text="Add", bootstyle=SUCCESS, command=self._add).pack(side=LEFT, padx=2)

        frame = ttk.Frame(self); frame.pack(fill=BOTH, expand=True, padx=14, pady=6)
        sb = ttk.Scrollbar(frame, orient=VERTICAL)
        self.tree = ttk.Treeview(frame, columns=("name", "kind"), show="headings",
                                 yscrollcommand=sb.set, height=12)
        sb.config(command=self.tree.yview); sb.pack(side=RIGHT, fill=Y)
        self.tree.pack(fill=BOTH, expand=True)
        for c, t, w in [("name", "Category", 240), ("kind", "Type", 100)]:
            self.tree.heading(c, text=t); self.tree.column(c, width=w)
        ttk.Button(self, text="Delete Selected", bootstyle=(DANGER, OUTLINE),
                   command=self._del).pack(pady=(0, 4))
        ttk.Button(self, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=self._close).pack(pady=(0, 12))

    def _reload(self):
        self.tree.delete(*self.tree.get_children())
        for c in self.db.get_budget_categories():
            self.tree.insert("", "end", iid=str(c["id"]), values=(c["name"], c["kind"]))

    def _add(self):
        name = self._name.get().strip()
        if name:
            self.db.add_budget_category(name, self._kind.get())
            self._name.set("")
            self._reload()

    def _del(self):
        sel = self.tree.selection()
        if sel:
            self.db.delete_budget_category(int(sel[0]))
            self._reload()

    def _close(self):
        if self.on_done:
            self.on_done()
        self.destroy()


# ── Student fees ────────────────────────────────────────────────────────────────

class _FeesDialog(ttk.Toplevel):
    def __init__(self, parent, db, program_type, school_year):
        super().__init__(parent)
        self.db = db
        self.program_type = program_type
        # Fees attach to the STUDENT roster year (academic), which may differ
        # from the budget's fiscal year — default to a year that has students.
        years = db.get_school_years()
        default = school_year if school_year in years else (years[0] if years else
                                                            db.current_school_year())
        self.school_year = default
        self._year_var = tk.StringVar(value=default)
        self._fee_var = tk.StringVar()
        self._checked = set()   # fee-row ids ticked for bulk actions
        self.title("Student Fees — Roka's Resonance")
        self.resizable(True, True)
        self.grab_set()
        self.lift()
        self._build()
        self._reload()
        from ui.theme import fit_window
        fit_window(self, 900, 600)

    def _build(self):
        hdr = ttk.Frame(self, bootstyle=INFO)
        hdr.pack(fill=X)
        ttk.Label(hdr, text="🎽  Student Fees",
                  font=("Segoe UI", 13, "bold"), bootstyle=(INVERSE, INFO)).pack(
            pady=10, padx=16, anchor=W)

        bar = ttk.Frame(self); bar.pack(fill=X, padx=12, pady=(8, 2))
        ttk.Label(bar, text="School Year:", font=("Segoe UI", 9, "bold")).pack(side=LEFT, padx=(0, 4))
        years = self.db.get_school_years() or [self.school_year]
        if self.school_year not in years:
            years.insert(0, self.school_year)
        yc = ttk.Combobox(bar, textvariable=self._year_var, state="readonly", width=12, values=years)
        yc.pack(side=LEFT, padx=(0, 12))
        yc.bind("<<ComboboxSelected>>", lambda e: self._on_year())
        ttk.Label(bar, text="Fee:", font=("Segoe UI", 9, "bold")).pack(side=LEFT, padx=(0, 4))
        self._fee_combo = ttk.Combobox(bar, textvariable=self._fee_var, state="readonly", width=22)
        self._fee_combo.pack(side=LEFT)
        self._fee_combo.bind("<<ComboboxSelected>>", lambda e: self._reload_list())
        ttk.Button(bar, text="Manage Fee Types", bootstyle=(SECONDARY, OUTLINE),
                   command=self._manage_types).pack(side=LEFT, padx=8)

        tb = ttk.Frame(self); tb.pack(fill=X, padx=12, pady=4)
        ttk.Button(tb, text="➕ Add Students…", bootstyle=SUCCESS,
                   command=self._add_students).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="🎓 Add all Entry students", bootstyle=(SUCCESS, OUTLINE),
                   command=self._add_entry).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="⧉ Duplicate", bootstyle=(SUCCESS, OUTLINE),
                   command=self._duplicate).pack(side=LEFT, padx=2)
        ttk.Separator(tb, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=2)
        ttk.Button(tb, text="Select All", bootstyle=(SECONDARY, OUTLINE),
                   command=self._check_all).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="Clear", bootstyle=(SECONDARY, OUTLINE),
                   command=self._clear_checks).pack(side=LEFT, padx=2)
        ttk.Separator(tb, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=2)
        ttk.Button(tb, text="✅ Paid", bootstyle=(SUCCESS, OUTLINE),
                   command=lambda: self._set_status("paid")).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="⛔ Unpaid", bootstyle=(WARNING, OUTLINE),
                   command=lambda: self._set_status("unpaid")).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="🚫 Waive", bootstyle=(SECONDARY, OUTLINE),
                   command=lambda: self._set_status("waived")).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="🗑️ Remove", bootstyle=(DANGER, OUTLINE),
                   command=self._remove).pack(side=LEFT, padx=2)
        ttk.Separator(tb, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=2)
        ttk.Button(tb, text="✉️ Email Unpaid", bootstyle=INFO,
                   command=self._email_unpaid).pack(side=LEFT, padx=2)
        ttk.Button(tb, text="📊 Export", bootstyle=(SECONDARY, OUTLINE),
                   command=self._export).pack(side=LEFT, padx=2)

        ttk.Label(self, text="Tick the ☐ boxes (or Ctrl/Shift-click rows) to select several, "
                             "then Paid / Waive.", font=("Segoe UI", 8),
                  foreground="#888").pack(anchor=W, padx=14)

        frame = ttk.Frame(self); frame.pack(fill=BOTH, expand=True, padx=12, pady=8)
        cols = ("chk", "name", "grade", "ensembles", "amount", "status", "inst")
        sb = ttk.Scrollbar(frame, orient=VERTICAL)
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="extended",
                                 yscrollcommand=sb.set, bootstyle=INFO)
        sb.config(command=self.tree.yview); sb.pack(side=RIGHT, fill=Y)
        self.tree.pack(fill=BOTH, expand=True)
        heads = {"chk": "✓", "name": "Student", "grade": "Grade", "ensembles": "Ensembles",
                 "amount": "Amount", "status": "Status", "inst": "Instrument Out?"}
        widths = {"chk": 34, "name": 200, "grade": 55, "ensembles": 220, "amount": 90,
                  "status": 90, "inst": 110}
        for c in cols:
            self.tree.heading(c, text=heads[c], anchor=W,
                              command=(self._toggle_all_header if c == "chk" else ""))
            self.tree.column(c, width=widths[c], anchor=(CENTER if c == "chk" else W))
        self.tree.tag_configure("paid", foreground="#1a7a1a")
        self.tree.tag_configure("waived", foreground="#888")
        self.tree.tag_configure("unpaid", foreground="#b00000")
        self.tree.bind("<Button-1>", self._on_click, add="+")
        self._count = ttk.Label(self, text="", foreground="#666")
        self._count.pack(anchor=W, padx=14)

        ttk.Button(self, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(pady=(0, 12))

    def _on_year(self):
        self.school_year = self._year_var.get()
        self._reload_list()

    def _reload(self):
        types = [t["name"] for t in self.db.get_fee_types()]
        self._fee_combo["values"] = types
        if types and not self._fee_var.get():
            self._fee_var.set(types[0])
        self._reload_list()

    def _fee_amount(self):
        for t in self.db.get_fee_types():
            if t["name"] == self._fee_var.get():
                return float(t["default_amount"] or 0)
        return 0.0

    def _reload_list(self):
        self.tree.delete(*self.tree.get_children())
        fee = self._fee_var.get()
        if not fee:
            return
        rows = self.db.get_student_fees(fee, self.school_year)
        present = {r["id"] for r in rows}
        self._checked &= present    # drop checks for rows no longer shown
        n_unpaid = 0
        with self.db._connect() as conn:
            for r in rows:
                out = conn.execute(
                    "SELECT COUNT(*) FROM checkouts WHERE student_id=? AND date_returned IS NULL",
                    (r["student_id"],)).fetchone()[0]
                if r["status"] == "unpaid":
                    n_unpaid += 1
                self.tree.insert("", "end", iid=str(r["id"]), tags=(r["status"],), values=(
                    "☑" if r["id"] in self._checked else "☐",
                    display_last_first(r),
                    r["grade"] or "",
                    r["ensembles"] or "",
                    _money(r["amount"]),
                    r["status"].title(),
                    "✓ Yes" if out else "",
                ))
        self._count.config(text=f"{len(rows)} student(s) • {n_unpaid} unpaid "
                                f"• {len(self._checked)} ticked")

    def _on_click(self, event):
        if self.tree.identify("region", event.x, event.y) != "cell":
            return
        if self.tree.identify_column(event.x) != "#1":
            return
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        fid = int(iid)
        self._checked.symmetric_difference_update({fid})
        self.tree.set(iid, "chk", "☑" if fid in self._checked else "☐")
        n = len([1 for i in self.tree.get_children()
                 if self.tree.set(i, "status") == "Unpaid"])
        self._count.config(text=f"{len(self.tree.get_children())} student(s) • {n} unpaid "
                                f"• {len(self._checked)} ticked")

    def _check_all(self):
        for iid in self.tree.get_children():
            self._checked.add(int(iid)); self.tree.set(iid, "chk", "☑")
        self._reload_list()

    def _clear_checks(self):
        self._checked.clear()
        self._reload_list()

    def _toggle_all_header(self):
        shown = [int(i) for i in self.tree.get_children()]
        if shown and all(s in self._checked for s in shown):
            self._clear_checks()
        else:
            self._check_all()

    def _sel_ids(self):
        """Ticked rows take priority; otherwise the highlighted selection."""
        if self._checked:
            return list(self._checked)
        return [int(i) for i in self.tree.selection()]

    def _set_status(self, status):
        ids = self._sel_ids()
        if not ids:
            Messagebox.show_warning("Tick or select student(s) first.", title="No Selection",
                                    parent=self)
            return
        today = datetime.today().strftime("%Y-%m-%d") if status == "paid" else None
        for fid in ids:
            self.db.set_student_fee_status(fid, status, today)
        self._checked.clear()
        self._reload_list()

    def _remove(self):
        ids = self._sel_ids()
        for fid in ids:
            self.db.delete_student_fee(fid)
        self._checked.clear()
        self._reload_list()

    def _duplicate(self):
        """Add a second (unpaid) copy of the selected fee(s) — for students who
        owe this fee more than once, e.g. renting several instruments (Grace's
        3 summer rentals = 3 × $20).  Each click adds one more copy per row."""
        ids = self._sel_ids()
        if not ids:
            Messagebox.show_warning("Tick or select the student fee(s) to duplicate.",
                                    title="No Selection", parent=self)
            return
        rows = {r["id"]: r for r in self.db.get_student_fees(self._fee_var.get(),
                                                             self.school_year)}
        made = 0
        for fid in ids:
            r = rows.get(fid)
            if not r:
                continue
            self.db.add_student_fee(r["student_id"], r["fee_type"],
                                    self.school_year, float(r["amount"] or 0))
            made += 1
        self._checked.clear()
        self._reload_list()
        if made:
            Messagebox.show_info(
                f"Added {made} more unpaid {self._fee_var.get()} fee(s). Mark each "
                "Paid as the student pays.", title="Duplicated", parent=self)

    def _add_students(self):
        fee = self._fee_var.get()
        if not fee:
            Messagebox.show_warning("Add a fee type first.", title="No Fee", parent=self)
            return
        dlg = _StudentPickerDialog(self, self.db, self.program_type, self.school_year)
        self.wait_window(dlg)
        if dlg.chosen_ids:
            amt = self._fee_amount()
            for sid in dlg.chosen_ids:
                self.db.ensure_student_fee(sid, fee, self.school_year, amt)
            self._reload_list()

    def _add_entry(self):
        """Polo default: everyone in an Entry-level class."""
        fee = self._fee_var.get()
        if not fee:
            return
        entry_names = [e for e in ensembles_for(self.program_type) if e.lower().startswith("entry")]
        amt = self._fee_amount()
        added = 0
        for s in self.db.get_all_students(school_year=self.school_year):
            ens = [x.strip() for x in (self.db_sval(s, "ensembles")).split(",") if x.strip()]
            if any(e in ens for e in entry_names):
                self.db.ensure_student_fee(s["id"], fee, self.school_year, amt)
                added += 1
        Messagebox.show_info(f"Added/kept {fee} for {added} Entry student(s).",
                             title="Done", parent=self)
        self._reload_list()

    @staticmethod
    def db_sval(row, key):
        try:
            return row[key] or "" if key in row.keys() else ""
        except Exception:
            return ""

    def _email_unpaid(self):
        fee = self._fee_var.get()
        if not fee:
            Messagebox.show_warning("Pick a fee first.", title="No Fee", parent=self)
            return
        # School name from teacher settings (falls back to "the main")
        school = "the"
        try:
            from ui.settings_dialog import load_settings
            school = (load_settings(self.base_dir).get("teacher") or {}).get(
                "school_name") or "the"
        except Exception:
            pass

        win = ttk.Toplevel(self)
        win.title(f"Email — Unpaid: {fee}")
        win.grab_set()
        only_out = tk.BooleanVar(value=False)
        txt = tk.Text(win, height=4, width=72, wrap=WORD, relief="solid", bd=1)
        info = ttk.Label(win, text="", font=("Segoe UI", 9), justify=LEFT)

        def _refresh(*_):
            if only_out.get():
                rows = self.db.get_unpaid_fee_with_checkout(fee, self.school_year)
            else:
                rows = self.db.get_unpaid_fee(fee, self.school_year)
            emails, seen = [], set()
            for r in rows:
                for key in ("parent1_email", "parent2_email", "student_email"):
                    a = (r.get(key) or "").strip()
                    if a and "@" in a and a.lower() not in seen:
                        seen.add(a.lower()); emails.append(a)
            info.config(text=f"{len(rows)} student(s) owe '{fee}'"
                             f"{' and have an instrument out' if only_out.get() else ''}. "
                             f"{len(emails)} email(s). Paste into Outlook (semicolon-separated):")
            txt.delete("1.0", "end"); txt.insert("1.0", "; ".join(emails))

        info.pack(anchor=W, padx=14, pady=(12, 4))
        ttk.Checkbutton(win, text="Only those who also have an instrument checked out",
                        variable=only_out, bootstyle=INFO, command=_refresh).pack(anchor=W, padx=14)
        txt.pack(fill=X, padx=14, pady=(4, 0))
        _refresh()

        def _copy():
            win.clipboard_clear(); win.clipboard_append(txt.get("1.0", "end").strip())
            Messagebox.show_info("Recipient addresses copied.", title="Copied", parent=win)

        b = ttk.Frame(win); b.pack(fill=X, padx=14, pady=(6, 4))
        ttk.Button(b, text="📋 Copy Addresses", bootstyle=INFO,
                   command=_copy).pack(side=RIGHT, padx=4)

        # ---- Pre-made message template to paste into Outlook ----
        ttk.Separator(win).pack(fill=X, padx=14, pady=(4, 0))
        ttk.Label(win, text="Message template (edit as needed, then Copy):",
                  font=("Segoe UI", 9, "bold")).pack(anchor=W, padx=14, pady=(8, 2))
        body = (
            "Good morning,\n\n"
            f"If you're receiving this email, we have not yet received your payment "
            f"for {fee}. Payments can be made in person at {school} main office, or "
            "online at https://wa-bellevue.intouchreceipting.com/. Please talk to the "
            "main office right away if you have any questions or if you need financial "
            "assistance.\n\n"
            "Thank you,"
        )
        msg = tk.Text(win, height=9, width=72, wrap=WORD, relief="solid", bd=1)
        msg.pack(fill=BOTH, expand=True, padx=14, pady=(0, 0))
        msg.insert("1.0", body)

        def _copy_msg():
            win.clipboard_clear(); win.clipboard_append(msg.get("1.0", "end").strip())
            Messagebox.show_info("Message template copied. Paste it into Outlook.",
                                 title="Copied", parent=win)

        b2 = ttk.Frame(win); b2.pack(fill=X, padx=14, pady=12)
        ttk.Button(b2, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(b2, text="📋 Copy Message", bootstyle=SUCCESS,
                   command=_copy_msg).pack(side=RIGHT, padx=4)
        from ui.theme import fit_window
        fit_window(win, 600, 560)

    def _manage_types(self):
        win = ttk.Toplevel(self)
        win.title("Fee Types")
        win.grab_set()
        ttk.Label(win, text="Fee Types", font=("Segoe UI", 12, "bold"),
                  bootstyle=INFO).pack(anchor=W, padx=14, pady=(12, 4))
        add = ttk.Frame(win); add.pack(fill=X, padx=14)
        name = tk.StringVar(); amt = tk.StringVar()
        ttk.Entry(add, textvariable=name, width=20).pack(side=LEFT)
        ttk.Label(add, text="$").pack(side=LEFT, padx=(6, 0))
        ttk.Entry(add, textvariable=amt, width=8).pack(side=LEFT, padx=(0, 4))
        lst = tk.Listbox(win, height=8, width=40)
        lst.pack(fill=BOTH, expand=True, padx=14, pady=6)

        def fill():
            lst.delete(0, END)
            self._types = self.db.get_fee_types()
            for t in self._types:
                lst.insert(END, f"{t['name']}  —  {_money(t['default_amount'])}")

        def add_type():
            if name.get().strip():
                try:
                    a = float(amt.get() or 0)
                except ValueError:
                    a = 0
                self.db.add_fee_type(name.get().strip(), a)
                name.set(""); amt.set(""); fill(); self._reload()

        def del_type():
            sel = lst.curselection()
            if sel:
                self.db.delete_fee_type(self._types[sel[0]]["id"]); fill(); self._reload()

        ttk.Button(add, text="Add", bootstyle=SUCCESS, command=add_type).pack(side=LEFT, padx=2)
        bb = ttk.Frame(win); bb.pack(fill=X, padx=14, pady=8)
        ttk.Button(bb, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(bb, text="Delete Selected", bootstyle=(DANGER, OUTLINE),
                   command=del_type).pack(side=RIGHT, padx=4)
        fill()
        from ui.theme import fit_window
        fit_window(win, 380, 340)

    def _export(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            Messagebox.show_error("openpyxl is required.", title="Missing Dependency", parent=self)
            return
        fee = self._fee_var.get()
        rows = self.db.get_student_fees(fee, self.school_year)
        if not rows:
            Messagebox.show_info("No students on this fee.", title="Nothing to Export", parent=self)
            return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fees"
        ws.append(["Last Name", "First Name", "Grade", "Ensembles", "Amount", "Status",
                   "Date Paid"])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2E7D32")
        for r in rows:
            ws.append([r["last_name"], r["first_name"], r["grade"] or "", r["ensembles"] or "",
                       float(r["amount"] or 0), r["status"], r["date_paid"] or ""])
        from tkinter import filedialog
        import datetime as _d
        path = filedialog.asksaveasfilename(
            title="Save Fee Export", parent=self, defaultextension=".xlsx",
            initialfile=f"{fee}_{self.school_year}_{_d.date.today().isoformat()}.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        wb.save(path)
        if Messagebox.yesno("Exported. Open it now?", title="Done", parent=self) == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)


class _StudentPickerDialog(ttk.Toplevel):
    """Tick students (optionally filtered by ensemble) to add to a fee."""
    def __init__(self, parent, db, program_type, school_year):
        super().__init__(parent)
        self.db = db
        self.program_type = program_type
        self.school_year = school_year
        self.chosen_ids = []
        self._checked = set()
        self._filter = tk.StringVar(value="All")
        self.title("Add Students to Fee")
        self.grab_set()
        self.lift()
        self._build()
        self._reload()
        from ui.theme import fit_window
        fit_window(self, 520, 520)

    def _build(self):
        bar = ttk.Frame(self); bar.pack(fill=X, padx=12, pady=(12, 4))
        ttk.Label(bar, text="Ensemble filter:").pack(side=LEFT, padx=(0, 4))
        ttk.Combobox(bar, textvariable=self._filter, state="readonly", width=20,
                     values=["All"] + ensembles_for(self.program_type)).pack(side=LEFT)
        self._filter.trace_add("write", lambda *_: self._reload())
        ttk.Button(bar, text="Select All", bootstyle=(SECONDARY, OUTLINE),
                   command=self._all).pack(side=RIGHT, padx=2)

        frame = ttk.Frame(self); frame.pack(fill=BOTH, expand=True, padx=12, pady=6)
        sb = ttk.Scrollbar(frame, orient=VERTICAL)
        self.tree = ttk.Treeview(frame, columns=("chk", "name", "grade", "ens"),
                                 show="headings", yscrollcommand=sb.set, selectmode="browse")
        sb.config(command=self.tree.yview); sb.pack(side=RIGHT, fill=Y)
        self.tree.pack(fill=BOTH, expand=True)
        for c, t, w in [("chk", "✓", 34), ("name", "Student", 200), ("grade", "Grade", 55),
                        ("ens", "Ensembles", 200)]:
            self.tree.heading(c, text=t)
            self.tree.column(c, width=w, anchor=(CENTER if c == "chk" else W))
        self.tree.bind("<Button-1>", self._click, add="+")

        b = ttk.Frame(self); b.pack(fill=X, padx=14, pady=12)
        ttk.Button(b, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=self.destroy).pack(side=RIGHT, padx=4)
        ttk.Button(b, text="Add Selected", bootstyle=SUCCESS,
                   command=self._ok).pack(side=RIGHT, padx=4)

    def _reload(self):
        self.tree.delete(*self.tree.get_children())
        filt = self._filter.get()
        for s in self.db.get_all_students(school_year=self.school_year):
            ens = _FeesDialog.db_sval(s, "ensembles")
            if filt != "All" and filt not in [x.strip() for x in ens.split(",")]:
                continue
            self.tree.insert("", "end", iid=str(s["id"]),
                             values=("☑" if s["id"] in self._checked else "☐",
                                     display_last_first(s), s["grade"] or "", ens))

    def _click(self, e):
        if self.tree.identify("region", e.x, e.y) != "cell" or self.tree.identify_column(e.x) != "#1":
            return
        iid = self.tree.identify_row(e.y)
        if not iid:
            return
        sid = int(iid)
        self._checked.symmetric_difference_update({sid})
        self.tree.set(iid, "chk", "☑" if sid in self._checked else "☐")

    def _all(self):
        for iid in self.tree.get_children():
            self._checked.add(int(iid)); self.tree.set(iid, "chk", "☑")

    def _ok(self):
        self.chosen_ids = list(self._checked)
        self.destroy()
