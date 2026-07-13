"""
ui/inventory_manager.py - Main inventory management screen
"""

import os
import json
import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from ui.theme import muted_fg, fs, bind_copy_menu


TREEVIEW_COLS = (
    "status", "category", "description", "brand", "model",
    "barcode", "serial_no", "condition",
    "checked_out_to", "est_value"
)
COL_HEADERS = {
    "status":        "●",
    "category":      "Category",
    "description":   "Instrument",
    "brand":         "Brand",
    "model":         "Model",
    "barcode":       "Barcode",
    "serial_no":     "Serial #",
    "condition":     "Condition",
    "checked_out_to": "Assigned To",
    "est_value":     "Est. Value",
}
COL_WIDTHS = {
    "status":        40,
    "category":      95,
    "description":   190,
    "brand":         115,
    "model":         115,
    "barcode":       90,
    "serial_no":     100,
    "condition":     95,     # fits "Needs Repair" without truncating
    "checked_out_to": 150,
    "est_value":     80,
}


class InventoryManager(ttk.Frame):
    def __init__(self, parent, db, base_dir: str, on_checkouts=None):
        super().__init__(parent)
        self.db = db
        self.base_dir = base_dir
        self._on_checkouts = on_checkouts
        self._all_rows = []
        self._selected_id = None
        self._sort_col = "category"
        self._sort_asc = True
        self._search_var = tk.StringVar()
        self._filter_status = tk.StringVar(value="All")
        self._filter_category = tk.StringVar(value="All")
        self._show_inactive = tk.BooleanVar(value=False)
        self._col_vars = {c: tk.BooleanVar(value=True)
                          for c in TREEVIEW_COLS if c != "status"}
        self._col_popup = None
        self._chat_window = None

        self._build()
        self._load_col_prefs()
        self._apply_col_visibility()
        self.refresh()
        self.after(400, self._maybe_recover_repairs)

    def _maybe_recover_repairs(self):
        """One-time: rescue repair info that older check-ins buried in returned
        checkout notes and turn it into visible pending repair records."""
        flag = os.path.join(self.base_dir, ".repairs_recovered")
        if os.path.exists(flag):
            return
        try:
            n = self.db.recover_repair_notes_from_checkins()
        except Exception:
            n = 0
        try:
            with open(flag, "w") as f:
                f.write("done")
        except Exception:
            pass
        if n:
            self.refresh()
            Messagebox.show_info(
                f"Recovered {n} repair note(s) that earlier check-ins had saved but "
                f"never displayed.\n\nThey are now listed as pending repairs — see the "
                f"Repairs tab, or use '🧾 Repairs Needed' to print them for your technician.",
                title="Repair Notes Recovered", parent=self.winfo_toplevel())

    # ─────────────────────────────────────────────────────────────── Build UI ──

    def _build(self):
        # ── Toolbar ───────────────────────────────────────────────────────────
        toolbar = ttk.Frame(self, bootstyle=LIGHT)
        toolbar.pack(fill=X, padx=0, pady=0)

        # ── Group 1: core actions ─────────────────────────────────────────
        ttk.Button(toolbar, text="➕", bootstyle=SUCCESS, width=3,
                   command=self._add_instrument).pack(side=LEFT, padx=6, pady=6)
        ttk.Button(toolbar, text="✏️ Edit", bootstyle=PRIMARY,
                   command=self._edit_instrument).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="📤 Check Out", bootstyle=WARNING,
                   command=self._open_checkout_chooser).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="📥 Check In", bootstyle=INFO,
                   command=self._open_checkin_chooser).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="📄 Generate Form", bootstyle=PRIMARY,
                   command=self._generate_form).pack(side=LEFT, padx=2, pady=6)

        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)

        # ── Group 2: repair hub + exports ─────────────────────────────────
        ttk.Button(toolbar, text="📋 Checkouts", bootstyle=(WARNING, OUTLINE),
                   command=self._show_checkouts).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="🔧 Repair", bootstyle=SECONDARY,
                   command=self._open_repair_hub).pack(side=LEFT, padx=2, pady=6)
        ttk.Button(toolbar, text="📊 Exports ▾", bootstyle=SECONDARY,
                   command=self._open_exports_menu).pack(side=LEFT, padx=2, pady=6)

        ttk.Separator(toolbar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8, pady=4)

        ttk.Checkbutton(toolbar, text="Show Inactive",
                        variable=self._show_inactive,
                        command=self.refresh,
                        bootstyle=(SECONDARY, ROUND, TOGGLE)
                        ).pack(side=LEFT, padx=8, pady=6)

        # ── Search / Filter Bar ───────────────────────────────────────────────
        filter_bar = ttk.Frame(self)
        filter_bar.pack(fill=X, padx=8, pady=(4, 2))

        ttk.Label(filter_bar, text="Search:").pack(side=LEFT, padx=(0, 4))
        search_entry = ttk.Entry(filter_bar, textvariable=self._search_var, width=28)
        search_entry.pack(side=LEFT, padx=(0, 10))
        self._search_var.trace_add("write", lambda *_: self._apply_filters())

        ttk.Label(filter_bar, text="Status:").pack(side=LEFT, padx=(0, 4))
        ttk.Combobox(filter_bar, textvariable=self._filter_status,
                     values=["All", "Available", "Checked Out", "On Loan"],
                     state="readonly", width=14
                     ).pack(side=LEFT, padx=(0, 10))
        self._filter_status.trace_add("write", lambda *_: self._apply_filters())

        ttk.Label(filter_bar, text="Category:").pack(side=LEFT, padx=(0, 4))
        self._cat_combo = ttk.Combobox(filter_bar, textvariable=self._filter_category,
                                        state="readonly", width=18)
        self._cat_combo.pack(side=LEFT, padx=(0, 10))
        self._filter_category.trace_add("write", lambda *_: self._apply_filters())

        self._count_label = ttk.Label(filter_bar, text="", foreground=muted_fg())
        self._count_label.pack(side=RIGHT, padx=6)

        self._col_btn = ttk.Button(
            filter_bar, text="Columns ▼",
            bootstyle=(SECONDARY, OUTLINE), width=10,
            command=lambda: self._show_col_chooser(self._col_btn)
        )
        self._col_btn.pack(side=RIGHT, padx=(0, 4))

        # ── Main Content: Tree + Detail Panel ─────────────────────────────────
        paned = ttk.Panedwindow(self, orient=HORIZONTAL)
        paned.pack(fill=BOTH, expand=True, padx=6, pady=4)

        # Left: Treeview
        tree_frame = ttk.Frame(paned)
        paned.add(tree_frame, weight=3)

        scrollbar_y = ttk.Scrollbar(tree_frame, orient=VERTICAL)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient=HORIZONTAL)

        self.tree = ttk.Treeview(
            tree_frame,
            columns=TREEVIEW_COLS,
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            selectmode="browse",
            bootstyle=PRIMARY,
        )
        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)

        scrollbar_y.pack(side=RIGHT, fill=Y)
        scrollbar_x.pack(side=BOTTOM, fill=X)
        self.tree.pack(fill=BOTH, expand=True)

        # Let every TEXT column share the leftover width evenly (ttk spreads
        # extra space equally across stretchable columns), so widening the window
        # no longer dumps all the slack into two columns and leaves odd gaps.
        # Fixed-width, right/center-aligned columns (status dot, barcode number,
        # est. value) stay put.
        _STRETCH = {"category", "description", "brand", "model", "serial_no",
                    "condition", "checked_out_to"}
        for col in TREEVIEW_COLS:
            self.tree.heading(
                col,
                text=COL_HEADERS[col],
                anchor=W,
                command=lambda c=col: self._sort_by(c)
            )
            self.tree.column(col, width=COL_WIDTHS[col], anchor=W,
                             minwidth=40, stretch=col in _STRETCH)

        self.tree.column("status", anchor=CENTER, stretch=False)
        self.tree.heading("status", anchor=CENTER)
        self.tree.column("est_value", anchor=E, stretch=False)
        self.tree.heading("est_value", anchor=E)

        self.tree.tag_configure("available", foreground="#1a7a1a")
        self.tree.tag_configure("checkedout", foreground="#8B4000")
        self.tree.tag_configure("onloan", foreground="#1f5fbf")
        self.tree.tag_configure("inactive", foreground="#AAAAAA")
        # Bright red for instruments that are NOT ok — unrepairable, missing,
        # lost, stolen — so they don't read as fine/available (green) or get
        # confused with the brick-red "checked out".
        self.tree.tag_configure("notok", foreground="#E60000")

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>", lambda e: self._edit_instrument())
        self.tree.bind("<Button-3>", self._on_right_click)

        # Right: Detail Panel
        detail_frame = ttk.Frame(paned, width=320)
        paned.add(detail_frame, weight=1)

        self._detail_nb = ttk.Notebook(detail_frame, bootstyle=PRIMARY)
        self._detail_nb.pack(fill=BOTH, expand=True)

        self._detail_tab = ttk.Frame(self._detail_nb)
        self._history_tab = ttk.Frame(self._detail_nb)
        self._repair_tab = ttk.Frame(self._detail_nb)

        self._detail_nb.add(self._detail_tab, text="Details")
        self._detail_nb.add(self._history_tab, text="Checkout History")
        self._detail_nb.add(self._repair_tab, text="Repairs")

        self._build_detail_tab()
        self._build_history_tab()
        self._build_repair_tab()

        # ── AI Chat Footer ─────────────────────────────────────────────────────
        footer = ttk.Frame(self)
        footer.pack(fill=X, side=BOTTOM)
        ttk.Separator(footer).pack(fill=X)
        ttk.Button(
            footer,
            text="🎩 Ask Reginald",
            bootstyle=(DARK, OUTLINE),
            command=self._open_chat,
        ).pack(side=RIGHT, padx=10, pady=5)

    def _build_detail_tab(self):
        outer = ttk.Frame(self._detail_tab)
        outer.pack(fill=BOTH, expand=True, padx=8, pady=8)

        self._detail_labels = {}
        fields = [
            ("Description", "description"),
            ("Category", "category"),
            ("Brand", "brand"),
            ("Model", "model"),
            ("Barcode", "barcode"),
            ("Serial #", "serial_no"),
            ("Condition", "condition"),
            ("Status", "_status"),
            ("Assigned To", "checked_out_to"),
            ("Checkout Date", "checkout_date"),
            ("Est. Value", "est_value"),
            ("Amount Paid", "amount_paid"),
            ("Date Purchased", "date_purchased"),
            ("PO Number", "po_number"),
            ("Last Service", "last_service"),
            ("Case #", "case_no"),
            ("Locker", "locker"),
            ("Lock #", "lock_no"),
            ("Combo", "combo"),
            ("Accessories", "accessories"),
            ("Comments", "comments"),
        ]

        for label, key in fields:
            row = ttk.Frame(outer)
            row.pack(fill=X, pady=1)
            ttk.Label(row, text=f"{label}:", font=("Segoe UI", fs(8), "bold"),
                      width=14, anchor=W).pack(side=LEFT)
            val_lbl = ttk.Label(row, text="", font=("Segoe UI", fs(8)),
                                 anchor=W, wraplength=180, justify=LEFT)
            val_lbl.pack(side=LEFT, fill=X, expand=True)
            bind_copy_menu(val_lbl)
            self._detail_labels[key] = val_lbl

    def _build_history_tab(self):
        frame = ttk.Frame(self._history_tab)
        frame.pack(fill=BOTH, expand=True)

        cols = ("Student", "Date Out", "Date In")
        sb = ttk.Scrollbar(frame, orient=VERTICAL)
        self._history_tree = ttk.Treeview(frame, columns=cols, show="headings",
                                           yscrollcommand=sb.set, bootstyle=INFO,
                                           height=12)
        sb.config(command=self._history_tree.yview)
        sb.pack(side=RIGHT, fill=Y)
        self._history_tree.pack(fill=BOTH, expand=True)

        _stretch_h = {"Student"}
        for col in cols:
            self._history_tree.heading(col, text=col, anchor=W)
            self._history_tree.column(col, width=100, anchor=W,
                                      minwidth=40, stretch=col in _stretch_h)

    def _build_repair_tab(self):
        frame = ttk.Frame(self._repair_tab)
        frame.pack(fill=BOTH, expand=True)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=X, padx=6, pady=4)
        ttk.Button(btn_frame, text="Add Repair", bootstyle=(SECONDARY, OUTLINE),
                   command=self._add_repair).pack(side=LEFT, padx=2)
        ttk.Button(btn_frame, text="Edit", bootstyle=(PRIMARY, OUTLINE),
                   command=self._edit_repair).pack(side=LEFT, padx=2)
        ttk.Button(btn_frame, text="Delete", bootstyle=(DANGER, OUTLINE),
                   command=self._delete_repair).pack(side=LEFT, padx=2)

        cols = ("Date", "Description", "Est $", "Act $", "Repaired")
        sb = ttk.Scrollbar(frame, orient=VERTICAL)
        self._repair_tree = ttk.Treeview(frame, columns=cols, show="headings",
                                          yscrollcommand=sb.set, bootstyle=WARNING,
                                          height=10, selectmode="extended")
        sb.config(command=self._repair_tree.yview)
        sb.pack(side=RIGHT, fill=Y)
        self._repair_tree.pack(fill=BOTH, expand=True)

        widths = [80, 160, 55, 55, 80]
        _stretch_r = {"Description"}
        for col, w in zip(cols, widths):
            self._repair_tree.heading(col, text=col, anchor=W)
            self._repair_tree.column(col, width=w, anchor=W,
                                     minwidth=40, stretch=col in _stretch_r)

        self._repair_total_var = tk.StringVar(value="")
        self._repair_costs = {}  # iid -> act_cost or est_cost
        self._repair_total_label = ttk.Label(
            frame, textvariable=self._repair_total_var,
            font=("Segoe UI", fs(9)), foreground=muted_fg(), anchor=E,
        )
        self._repair_total_label.pack(fill=X, padx=8, pady=(2, 4))
        self._repair_tree.bind("<<TreeviewSelect>>", self._on_repair_select)

    # ──────────────────────────────────────────────────── Column Visibility ────

    def _load_col_prefs(self):
        path = os.path.join(self.base_dir, "column_prefs.json")
        try:
            with open(path) as f:
                prefs = json.load(f)
            for col, var in self._col_vars.items():
                if col in prefs:
                    var.set(bool(prefs[col]))
        except (FileNotFoundError, json.JSONDecodeError):
            pass  # Use defaults (all visible)

    def _save_col_prefs(self):
        path = os.path.join(self.base_dir, "column_prefs.json")
        try:
            with open(path, "w") as f:
                json.dump({c: v.get() for c, v in self._col_vars.items()}, f, indent=2)
        except Exception:
            pass

    def _apply_col_visibility(self):
        visible = ["status"] + [
            c for c in TREEVIEW_COLS if c != "status" and self._col_vars[c].get()
        ]
        self.tree["displaycolumns"] = visible

    def _show_col_chooser(self, btn):
        # Toggle: close if already open
        if self._col_popup and self._col_popup.winfo_exists():
            self._col_popup.destroy()
            self._col_popup = None
            return

        popup = tk.Toplevel(self)
        popup.wm_overrideredirect(True)
        popup.wm_attributes("-topmost", True)
        self._col_popup = popup

        # Position below the button
        bx = btn.winfo_rootx()
        by = btn.winfo_rooty() + btn.winfo_height()
        popup.geometry(f"+{bx}+{by}")

        outer = ttk.Frame(popup, relief="solid", borderwidth=1)
        outer.pack(fill=BOTH, expand=True)

        ttk.Label(outer, text="Show Columns",
                  font=("Segoe UI", 9, "bold")).pack(anchor=W, padx=10, pady=(8, 2))
        ttk.Separator(outer, orient=HORIZONTAL).pack(fill=X, padx=6, pady=2)

        for col in TREEVIEW_COLS:
            if col == "status":
                continue
            ttk.Checkbutton(
                outer,
                text=COL_HEADERS[col],
                variable=self._col_vars[col],
                command=lambda: (self._apply_col_visibility(), self._save_col_prefs())
            ).pack(anchor=W, padx=10, pady=1)

        ttk.Separator(outer, orient=HORIZONTAL).pack(fill=X, padx=6, pady=(4, 2))

        btn_row = ttk.Frame(outer)
        btn_row.pack(fill=X, padx=10, pady=(2, 8))

        def _show_all():
            for var in self._col_vars.values():
                var.set(True)
            self._apply_col_visibility()
            self._save_col_prefs()

        ttk.Button(btn_row, text="Show All", bootstyle=(SECONDARY, OUTLINE),
                   command=_show_all).pack(side=LEFT)
        ttk.Button(btn_row, text="Done", bootstyle=PRIMARY,
                   command=popup.destroy).pack(side=RIGHT)

        # Close when clicking outside the popup
        root = self.winfo_toplevel()
        _bid = [None]

        def _check_outside(event):
            try:
                if not popup.winfo_exists():
                    if _bid[0]:
                        root.unbind("<Button-1>", _bid[0])
                    return
                px, py = popup.winfo_rootx(), popup.winfo_rooty()
                pw, ph = popup.winfo_width(), popup.winfo_height()
                if not (px <= event.x_root <= px + pw and
                        py <= event.y_root <= py + ph):
                    popup.destroy()
                    self._col_popup = None
                    if _bid[0]:
                        root.unbind("<Button-1>", _bid[0])
            except Exception:
                pass

        def _setup_binding():
            _bid[0] = root.bind("<Button-1>", _check_outside, add="+")

        popup.after(150, _setup_binding)

    # ─────────────────────────────────────────────────────────── Data Loading ──

    def refresh(self):
        """Reload all data from DB."""
        self._all_rows = list(self.db.get_instruments_with_status(
            include_inactive=self._show_inactive.get()
        ))
        self._update_category_filter()
        self._apply_filters()
        if self._selected_id:
            self._restore_selection()

    def _update_category_filter(self):
        cats = sorted(set(r["category"] or "Unknown" for r in self._all_rows))
        self._cat_combo["values"] = ["All"] + cats
        if self._filter_category.get() not in ["All"] + cats:
            self._filter_category.set("All")

    def _apply_filters(self):
        search = self._search_var.get().lower()
        status_filter = self._filter_status.get()
        cat_filter = self._filter_category.get()

        visible = []
        for row in self._all_rows:
            status = row["status"]
            cat = row["category"] or ""

            if status_filter != "All" and status != status_filter:
                continue
            if cat_filter != "All" and cat != cat_filter:
                continue
            if search:
                haystack = " ".join([
                    str(row["description"] or ""),
                    str(row["brand"] or ""),
                    str(row["model"] or ""),
                    str(row["barcode"] or ""),
                    str(row["district_no"] or ""),
                    str(row["serial_no"] or ""),
                    str(row["checked_out_to"] or ""),
                    str(row["category"] or ""),
                ]).lower()
                if search not in haystack:
                    continue
            visible.append(row)

        self._populate_tree(visible)
        n = len(visible)
        total = len(self._all_rows)
        self._count_label.config(text=f"{n} of {total} instruments")

    def _populate_tree(self, rows):
        # Sort
        reverse = not self._sort_asc
        try:
            rows = sorted(rows, key=lambda r: (r[self._sort_col] or "").lower(), reverse=reverse)
        except Exception:
            pass

        self.tree.delete(*self.tree.get_children())
        for row in rows:
            status = row["status"]
            is_active = row["is_active"] if "is_active" in row.keys() else 1
            cond = (row["condition"] or "").strip().lower()
            not_ok = any(w in cond for w in
                         ("unrepairable", "missing", "lost", "stolen"))
            if not is_active:
                tag = "inactive"
            elif not_ok:
                tag = "notok"          # bright red — not usable / not present
            elif status == "Available":
                tag = "available"
            elif status == "On Loan":
                tag = "onloan"
            else:
                tag = "checkedout"

            est = row["est_value"]
            try:
                est_str = f"${float(est):,.2f}" if est else ""
            except (TypeError, ValueError):
                est_str = str(est or "")

            values = (
                "●" if status == "Available" else ("⇄" if status == "On Loan" else "◉"),
                row["category"] or "",
                row["description"] or "",
                row["brand"] or "",
                row["model"] or "",
                row["barcode"] or row["district_no"] or "",
                row["serial_no"] or "",
                row["condition"] or "",
                row["checked_out_to"] or "",
                est_str,
            )
            iid = str(row["id"])
            self.tree.insert("", "end", iid=iid, values=values, tags=(tag,))

    def _sort_by(self, col: str):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._apply_filters()

    def _restore_selection(self):
        iid = str(self._selected_id)
        if self.tree.exists(iid):
            self.tree.selection_set(iid)
            self.tree.see(iid)

    # ─────────────────────────────────────────────────────── Detail Panel ──────

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        self._selected_id = int(iid)
        self._load_detail(self._selected_id)
        # Keep open chat window in sync with selected instrument
        if self._chat_window and self._chat_window.winfo_exists():
            instrument = self.db.get_instrument(self._selected_id)
            if instrument:
                self._chat_window.update_selected_instrument(dict(instrument))

    def _open_chat(self):
        from ui.chat_dialog import ChatDialog
        # Raise existing window if already open
        if self._chat_window and self._chat_window.winfo_exists():
            self._chat_window.lift()
            self._chat_window.focus_force()
            return
        instrument = None
        if self._selected_id:
            row = self.db.get_instrument(self._selected_id)
            if row:
                instrument = dict(row)
        self._chat_window = ChatDialog(
            self.winfo_toplevel(), self.db, self.base_dir, instrument
        )

    def _load_detail(self, instrument_id: int):
        instrument = self.db.get_instrument(instrument_id)
        if not instrument:
            return

        # Current status: on loan takes precedence, then (possibly several) checkouts
        loan = self.db.get_active_loan(instrument_id)
        actives = [dict(a) for a in self.db.get_active_checkouts_for_instrument(instrument_id)]
        if loan:
            status = "On Loan"
            parts = [loan["school"] or "Another school"]
            if loan["contact_name"]:
                parts.append(loan["contact_name"])
            checked_out_to = " — ".join(parts)
            contact_bits = [b for b in (loan["contact_email"], loan["contact_phone"]) if b]
            if contact_bits:
                checked_out_to += "  (" + ", ".join(contact_bits) + ")"
            checkout_date = loan["date_out"] or ""
        elif actives:
            status = "Checked Out"
            checked_out_to = ", ".join(a["student_name"] or "?" for a in actives)
            checkout_date = actives[0]["date_assigned"] or ""
        else:
            status = "Available"
            checked_out_to = ""
            checkout_date = ""

        row = dict(instrument)
        row["_status"] = status
        row["checked_out_to"] = checked_out_to
        row["checkout_date"] = checkout_date
        # Barcode and District # are the same identifier — show whichever is set.
        row["barcode"] = row.get("barcode") or row.get("district_no") or ""

        for key, lbl in self._detail_labels.items():
            val = row.get(key, "")
            if val is None:
                val = ""
            # Format currency fields
            if key in ("est_value", "amount_paid") and val:
                try:
                    val = f"${float(val):,.2f}"
                except (ValueError, TypeError):
                    val = str(val)
            lbl.config(text=str(val))

        # Color status label
        status_lbl = self._detail_labels.get("_status")
        if status_lbl:
            color = {"Available": "#1a7a1a", "On Loan": "#1f5fbf"}.get(status, "#8B4000")
            status_lbl.config(foreground=color, font=("Segoe UI", 8, "bold"))

        self._load_history(instrument_id)
        self._load_repairs(instrument_id)

    def _load_history(self, instrument_id: int):
        self._history_tree.delete(*self._history_tree.get_children())
        history = self.db.get_checkout_history(instrument_id)
        for h in history:
            self._history_tree.insert("", "end", values=(
                h["student_name"] or "",
                h["date_assigned"] or "",
                h["date_returned"] or "Active",
            ))

    def _load_repairs(self, instrument_id: int):
        self._repair_tree.delete(*self._repair_tree.get_children())
        self._repair_costs = {}
        repairs = self.db.get_repairs(instrument_id)
        for r in repairs:
            est = f"${r['est_cost']:,.2f}" if r["est_cost"] else ""
            act = f"${r['act_cost']:,.2f}" if r["act_cost"] else ""
            iid = self._repair_tree.insert("", "end", values=(
                r["date_added"] or "",
                r["description"] or "",
                est,
                act,
                r["date_repaired"] or "Pending",
            ))
            # Store repair id in item
            self._repair_tree.item(iid, tags=(str(r["id"]),))
            self._repair_costs[iid] = float(r["act_cost"] or r["est_cost"] or 0)
        total = sum(self._repair_costs.values())
        n = len(self._repair_costs)
        if n:
            self._repair_total_var.set(f"Total spent: ${total:,.2f}  ({n} repair{'s' if n != 1 else ''})")
        else:
            self._repair_total_var.set("")

    def _on_repair_select(self, _event=None):
        sel = self._repair_tree.selection()
        total = sum(self._repair_costs.values())
        if not sel:
            n = len(self._repair_costs)
            if n:
                self._repair_total_var.set(f"Total spent: ${total:,.2f}  ({n} repair{'s' if n != 1 else ''})")
            else:
                self._repair_total_var.set("")
            return
        sel_total = sum(self._repair_costs.get(iid, 0) for iid in sel)
        self._repair_total_var.set(
            f"Selected: ${sel_total:,.2f}  •  Total: ${total:,.2f}"
        )

    # ─────────────────────────────────────────────────────── Action Handlers ──

    def _get_selected_instrument(self):
        sel = self.tree.selection()
        if not sel:
            Messagebox.show_warning("Please select an instrument first.", title="No Selection",
                                    parent=self.winfo_toplevel())
            return None
        return int(sel[0])

    def _on_right_click(self, event):
        """Right-click a row → act on that instrument directly.  Keeps the less
        common actions (Loan to another school, Check In) one-click now that the
        Check Out button skips straight to checkout on a selection."""
        rowid = self.tree.identify_row(event.y)
        if not rowid:
            return
        self.tree.selection_set(rowid)
        self.tree.focus(rowid)
        iid = int(rowid)
        menu = tk.Menu(self.tree, tearoff=0)
        menu.add_command(label="📤  Check Out",
                         command=lambda: self._do_single_checkout(iid))
        menu.add_command(label="🏫  Loan to Another School…",
                         command=lambda: self._loan_instrument(iid))
        menu.add_command(label="📥  Check In",
                         command=lambda: self._do_single_checkin(iid))
        menu.add_separator()
        menu.add_command(label="✏️  Edit", command=self._edit_instrument)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _add_instrument(self):
        from ui.instrument_dialog import InstrumentDialog
        dlg = InstrumentDialog(self.winfo_toplevel(), self.db, instrument_id=None)
        self.wait_window(dlg)
        self.refresh()

    def _edit_instrument(self):
        iid = self._get_selected_instrument()
        if iid is None:
            return
        from ui.instrument_dialog import InstrumentDialog
        dlg = InstrumentDialog(self.winfo_toplevel(), self.db, instrument_id=iid)
        self.wait_window(dlg)
        self.refresh()
        self._load_detail(iid)

    # ── Check Out / Check In (chooser → single scan, bulk, or item) ────────────

    def _do_single_checkout(self, iid):
        # Blocked entirely while the instrument is out on loan to another school.
        loan = self.db.get_active_loan(iid)
        if loan:
            Messagebox.show_warning(
                f"This instrument is on loan to {loan['school']} and is not available "
                f"for local checkout.\n\nReturn it from loan first.",
                title="On Loan", parent=self.winfo_toplevel()
            )
            return
        # An instrument can be checked out to more than one person; confirm first.
        active = self.db.get_active_checkouts_for_instrument(iid)
        if active:
            names = ", ".join(a["student_name"] or "?" for a in active)
            if Messagebox.yesno(
                f"This instrument is already checked out to: {names}.\n\n"
                "Check it out to an additional person as well?",
                title="Already Checked Out", parent=self.winfo_toplevel()
            ) != "Yes":
                return
        from ui.checkout_dialog import CheckoutDialog
        dlg = CheckoutDialog(self.winfo_toplevel(), self.db, instrument_id=iid, mode="checkout")
        self.wait_window(dlg)
        self.refresh()
        self._load_detail(iid)

    def _do_single_checkin(self, iid):
        # On loan → offer to return the loan instead.
        loan = self.db.get_active_loan(iid)
        if loan:
            if Messagebox.yesno(
                f"This instrument is on loan to {loan['school']}.\n\n"
                "Mark it as returned from loan?",
                title="Return from Loan", parent=self.winfo_toplevel()) == "Yes":
                from datetime import datetime as _dt
                self.db.return_loan(loan["id"], _dt.today().strftime("%Y-%m-%d"))
                self.refresh()
                self._load_detail(iid)
            return

        active = [dict(a) for a in self.db.get_active_checkouts_for_instrument(iid)]
        if not active:
            Messagebox.show_warning(
                "This instrument is not currently checked out.",
                title="Not Checked Out", parent=self.winfo_toplevel()
            )
            return
        # If several people have it, choose which checkout to return.
        checkout = active[0] if len(active) == 1 else self._pick_active_checkout(active, "check in")
        if not checkout:
            return
        from ui.checkout_dialog import CheckoutDialog
        dlg = CheckoutDialog(self.winfo_toplevel(), self.db, instrument_id=iid,
                             mode="checkin", checkout_data=checkout)
        self.wait_window(dlg)
        self.refresh()
        self._load_detail(iid)

    def _pick_active_checkout(self, active, verb):
        """Choose one of several active checkouts. Returns a checkout dict or None."""
        win = ttk.Toplevel(self.winfo_toplevel())
        win.title(f"Which checkout to {verb}?")
        win.grab_set()
        ttk.Label(win, text=f"This instrument is checked out to {len(active)} people.\n"
                            f"Choose which one to {verb}:", font=("Segoe UI", 9),
                  justify=LEFT).pack(anchor=W, padx=16, pady=(14, 6))
        lb = tk.Listbox(win, font=("Segoe UI", 9), height=min(len(active), 10), width=48)
        lb.pack(fill=BOTH, expand=True, padx=16)
        for a in active:
            lb.insert(END, f"{a.get('student_name') or '?'}   (out {a.get('date_assigned') or '—'})")
        lb.selection_set(0)
        result = {"c": None}

        def _ok():
            sel = lb.curselection()
            if sel:
                result["c"] = active[sel[0]]
            win.destroy()
        btns = ttk.Frame(win); btns.pack(pady=12)
        ttk.Button(btns, text="OK", bootstyle=PRIMARY, command=_ok).pack(side=LEFT, padx=4)
        ttk.Button(btns, text="Cancel", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(side=LEFT, padx=4)
        from ui.theme import fit_window
        fit_window(win, 420, 300)
        self.wait_window(win)
        return result["c"]

    def _loan_instrument(self, iid):
        loan = self.db.get_active_loan(iid)
        if loan:
            Messagebox.show_info(f"Already on loan to {loan['school']}.",
                                 title="On Loan", parent=self.winfo_toplevel())
            return
        active = self.db.get_active_checkouts_for_instrument(iid)
        if active:
            if Messagebox.yesno(
                "This instrument is currently checked out locally.\n\n"
                "Loan it to another school anyway? (Local checkouts remain recorded.)",
                title="Currently Checked Out", parent=self.winfo_toplevel()) != "Yes":
                return
        from ui.checkout_dialog import LoanDialog
        dlg = LoanDialog(self.winfo_toplevel(), self.db, instrument_id=iid)
        self.wait_window(dlg)
        self.refresh()
        self._load_detail(iid)

    def _lookup_by_scan(self, text):
        """Resolve a typed/scanned barcode or serial to an instrument id."""
        text = (text or "").strip()
        if not text:
            return None
        inst = self.db.get_instrument_by_barcode(text) or self.db.get_instrument_by_serial(text)
        return inst["id"] if inst else None

    def _open_checkout_chooser(self):
        # Already selected an instrument in the grid? Go straight to checking THAT
        # one out — no need to re-find what you just clicked. With nothing
        # selected, open the scan / bulk / item / loan chooser.
        sel = self.tree.selection()
        if sel:
            self._do_single_checkout(int(sel[0]))
        else:
            self._open_scan_chooser("checkout")

    def _open_checkin_chooser(self):
        sel = self.tree.selection()
        if sel:
            self._do_single_checkin(int(sel[0]))
        else:
            self._open_scan_chooser("checkin")

    def _open_scan_chooser(self, mode):
        is_out = mode == "checkout"
        win = ttk.Toplevel(self.winfo_toplevel())
        win.title("Check Out" if is_out else "Check In")
        win.resizable(False, False)
        win.grab_set()

        style = WARNING if is_out else INFO
        hdr = ttk.Frame(win, bootstyle=style)
        hdr.pack(fill=X)
        ttk.Label(hdr, text=("📤  Check Out" if is_out else "📥  Check In"),
                  font=("Segoe UI", 13, "bold"),
                  bootstyle=(INVERSE, style)).pack(pady=10, padx=16, anchor=W)

        body = ttk.Frame(win)
        body.pack(fill=BOTH, expand=True, padx=20, pady=14)

        ttk.Label(body, text="Scan or type a barcode / serial number:",
                  font=("Segoe UI", 9, "bold")).pack(anchor=W)
        scan_var = tk.StringVar()
        # Pre-fill from the currently selected grid row, if any.
        sel = self.tree.selection()
        if sel:
            inst = self.db.get_instrument(int(sel[0]))
            if inst:
                scan_var.set(inst["barcode"] or inst["district_no"] or inst["serial_no"] or "")
        entry = ttk.Entry(body, textvariable=scan_var, width=34)
        entry.pack(anchor=W, pady=(3, 2))
        entry.focus_set()
        entry.select_range(0, END)

        status_lbl = ttk.Label(body, text="", font=("Segoe UI", 8), foreground="#CC0000")
        status_lbl.pack(anchor=W)

        def _go(_e=None):
            iid = self._lookup_by_scan(scan_var.get())
            if iid is None:
                status_lbl.config(text="No active instrument found for that barcode / serial #.")
                entry.select_range(0, END)
                return
            win.destroy()
            if is_out:
                self._do_single_checkout(iid)
            else:
                self._do_single_checkin(iid)

        entry.bind("<Return>", _go)
        ttk.Button(body, text=("Find & Check Out" if is_out else "Find & Check In"),
                   bootstyle=style, command=_go).pack(anchor=W, pady=(6, 0))

        ttk.Separator(body, orient=HORIZONTAL).pack(fill=X, pady=12)

        def _bulk():
            win.destroy()
            self._open_bulk("checkout" if is_out else "checkin")
        ttk.Button(body, text=("📦 Bulk Check Out (scan many)" if is_out
                               else "📦 Bulk Check In (scan many)"),
                   bootstyle=(style, OUTLINE), command=_bulk).pack(fill=X, pady=2)

        if is_out:
            def _item():
                win.destroy()
                self._checkout_item()
            ttk.Button(body, text="🎒 Check Out Item (mute, book, lyre…)",
                       bootstyle=(SECONDARY, OUTLINE), command=_item).pack(fill=X, pady=2)

            def _loan():
                iid = self._lookup_by_scan(scan_var.get())
                if iid is None:
                    status_lbl.config(text="Scan/type the instrument's barcode above first, "
                                           "then choose Loan.")
                    entry.focus_set()
                    return
                win.destroy()
                self._loan_instrument(iid)
            ttk.Button(body, text="🏫 Loan to Another School",
                       bootstyle=(PRIMARY, OUTLINE), command=_loan).pack(fill=X, pady=2)

        ttk.Button(win, text="Close", bootstyle=(SECONDARY, OUTLINE),
                   command=win.destroy).pack(pady=(0, 12))

        from ui.theme import fit_window
        fit_window(win, 380, 320)

    def _open_bulk(self, initial_tab):
        from ui.bulk_checkout_dialog import BulkCheckoutDialog
        dlg = BulkCheckoutDialog(self.winfo_toplevel(), self.db, self.base_dir,
                                 refresh_callback=self.refresh, initial_tab=initial_tab)
        self.wait_window(dlg)
        self.refresh()

    def _checkout_item(self):
        """Check out a free-text item (mute, book, lyre) not in inventory."""
        from ui.checkout_dialog import ItemCheckoutDialog
        dlg = ItemCheckoutDialog(self.winfo_toplevel(), self.db)
        self.wait_window(dlg)
        self.refresh()

    def _show_checkouts(self):
        """Open the Active Checkouts window (moved here from the main menu)."""
        if self._on_checkouts:
            self._on_checkouts()
        else:
            Messagebox.show_info("Active checkouts are available from the main menu.",
                                 title="Checkouts", parent=self.winfo_toplevel())

    def _open_repair_hub(self):
        from ui.repair_hub import RepairHub
        dlg = RepairHub(self.winfo_toplevel(), self)
        self.wait_window(dlg)
        self.refresh()
        if self._selected_id:
            self._load_detail(self._selected_id)

    def _open_exports_menu(self):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="📊  Entire Inventory (Excel)…",
                         command=self._export_inventory)
        menu.add_command(label="🧾  Instruments Needing Repair…",
                         command=self._export_repairs_needed)
        menu.add_command(label="📤  Currently Checked Out…",
                         command=self._export_checked_out)
        menu.add_separator()
        menu.add_command(label="🏷️  Barcode Sheet — Brass & Woodwind (PDF)…",
                         command=self._export_barcodes)
        menu.add_separator()
        menu.add_command(label="💰  Repair Cost Report (by $ spent)…",
                         command=self._export_cost_report)
        menu.add_command(label="🔧  Full Repair Details…",
                         command=self._export_repairs)
        try:
            x = self.winfo_pointerx()
            y = self.winfo_pointery()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _add_repair(self):
        iid = self._get_selected_instrument()
        if iid is None:
            return
        from ui.repair_dialog import RepairDialog
        dlg = RepairDialog(self.winfo_toplevel(), self.db, instrument_id=iid, repair_id=None)
        self.wait_window(dlg)
        if self._selected_id:
            self._load_repairs(self._selected_id)

    def _edit_repair(self):
        sel = self._repair_tree.selection()
        if not sel:
            return
        tags = self._repair_tree.item(sel[0], "tags")
        if not tags:
            return
        repair_id = int(tags[0])
        iid = self._get_selected_instrument()
        from ui.repair_dialog import RepairDialog
        dlg = RepairDialog(self.winfo_toplevel(), self.db, instrument_id=iid, repair_id=repair_id)
        self.wait_window(dlg)
        if self._selected_id:
            self._load_repairs(self._selected_id)

    def _delete_repair(self):
        sel = self._repair_tree.selection()
        if not sel:
            return
        tags = self._repair_tree.item(sel[0], "tags")
        if not tags:
            return
        repair_id = int(tags[0])
        if Messagebox.yesno("Delete this repair record?", title="Confirm Delete",
                            parent=self.winfo_toplevel()) == "Yes":
            self.db.delete_repair(repair_id)
            if self._selected_id:
                self._load_repairs(self._selected_id)

    def _generate_form(self):
        iid = self._get_selected_instrument()
        if iid is None:
            return
        active = [dict(a) for a in self.db.get_active_checkouts_for_instrument(iid)]
        if not active:
            Messagebox.show_warning(
                "This instrument must be checked out to generate a loan form.",
                title="Not Checked Out", parent=self.winfo_toplevel()
            )
            return
        # Several borrowers → let the user pick which loan form to generate.
        chosen = active[0] if len(active) == 1 else self._pick_active_checkout(active, "make a form for")
        if not chosen:
            return
        try:
            from pdf_generator import generate_form_for_checkout
            path = generate_form_for_checkout(self.db, chosen["id"], self.base_dir)
            self.db.mark_form_generated(chosen["id"])
            Messagebox.show_info(
                f"Loan form generated!\n\n{path}\n\nOpening now...",
                title="Form Generated", parent=self.winfo_toplevel()
            )
            os.startfile(path)
        except Exception as e:
            Messagebox.show_error(f"Failed to generate form:\n{e}", title="Error",
                                  parent=self.winfo_toplevel())

    def _upload_invoice(self):
        from tkinter import filedialog
        try:
            from invoice_parser import parse_invoices
        except ImportError as e:
            Messagebox.show_error(str(e), title="Missing Dependency",
                                  parent=self.winfo_toplevel())
            return

        paths = filedialog.askopenfilenames(
            title="Select Invoice PDF(s)",
            parent=self.winfo_toplevel(),
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not paths:
            return

        instruments = [dict(r) for r in self.db.get_all_instruments(include_inactive=False)]

        try:
            results = parse_invoices(list(paths), instruments)
        except ImportError as e:
            Messagebox.show_error(str(e), title="Missing Dependency",
                                  parent=self.winfo_toplevel())
            return
        except Exception as e:
            Messagebox.show_error(f"Error parsing invoice(s):\n{e}", title="Parse Error",
                                  parent=self.winfo_toplevel())
            return

        errors    = [r for r in results if "error" in r]
        summaries = [r for r in results if r.get("summary")]
        matches   = [r for r in results
                     if "error" not in r and not r.get("summary")]

        if not matches:
            msg = "No instrument matches found in the selected invoice(s)."
            if errors:
                msg += "\n\nErrors:\n" + "\n".join(e["error"] for e in errors)
            Messagebox.show_info(msg, title="No Matches Found",
                                 parent=self.winfo_toplevel())
            return

        # Reconciliation: matched records vs each invoice's grand total
        recon_lines = []
        for s in summaries:
            name = s["source_file"]
            if s["invoice_total"] <= 0:
                recon_lines.append(f"• {name}: couldn't read the invoice total "
                                   "to double-check the costs.")
            elif s["balanced"]:
                recon_lines.append(f"• {name}: records add up to the invoice "
                                   f"total (${s['invoice_total']:,.2f}) ✓")
            else:
                diff = s["invoice_total"] - s["matched_total"]
                if diff > 0:
                    recon_lines.append(
                        f"• {name}: ⚠ records total ${s['matched_total']:,.2f} of the "
                        f"${s['invoice_total']:,.2f} invoice — ${diff:,.2f} unaccounted "
                        "for (line items that didn't match an inventory instrument?)")
                else:
                    recon_lines.append(
                        f"• {name}: ⚠ records total ${s['matched_total']:,.2f}, MORE than "
                        f"the ${s['invoice_total']:,.2f} invoice — a cost may be "
                        "double-counted; check each amount while reviewing.")
        recon_note = ("\n\nCost check:\n" + "\n".join(recon_lines)) if recon_lines else ""

        # Confirm before opening dialogs
        error_note = f"\n\n({len(errors)} file(s) could not be parsed)" if errors else ""
        answer = Messagebox.yesno(
            f"Found {len(matches)} repair record(s) across "
            f"{len(set(m['source_file'] for m in matches))} invoice file(s)."
            f"{recon_note}{error_note}\n\n"
            "Review each repair record now?",
            title="Invoice Parsed", parent=self.winfo_toplevel()
        )
        if answer != "Yes":
            return

        # Duplicate guard: skip repairs already entered from this same invoice
        # (same instrument + invoice number) so re-scanning doesn't double them.
        dups = [m for m in matches
                if self.db.find_duplicate_repair(
                    m["instrument_id"],
                    m["prefill"].get("invoice_number"),
                    m["prefill"].get("description"))]
        if dups:
            skip = Messagebox.yesno(
                f"{len(dups)} of these repair(s) look already entered from this "
                "invoice (same instrument + invoice number).\n\n"
                "Skip the duplicates?  Choose No to enter them again anyway.",
                title="Possible Duplicates", parent=self.winfo_toplevel()) == "Yes"
            if skip:
                dup_ids = {id(m) for m in dups}
                matches = [m for m in matches if id(m) not in dup_ids]
            if not matches:
                Messagebox.show_info("Nothing new to enter — all matched repairs "
                                     "were already recorded.", title="All Duplicates",
                                     parent=self.winfo_toplevel())
                return

        saved_count = 0
        from ui.repair_dialog import RepairDialog
        for i, match in enumerate(matches, 1):
            suffix = f"{i} of {len(matches)}  —  {match['instrument_label']}"
            dlg = RepairDialog(
                self.winfo_toplevel(), self.db,
                instrument_id=match["instrument_id"],
                repair_id=None,
                prefill_data=match["prefill"],
                title_suffix=suffix,
            )
            self.wait_window(dlg)
            if dlg.saved:
                saved_count += 1

        self.refresh()
        if self._selected_id:
            self._load_repairs(self._selected_id)

        Messagebox.show_info(
            f"Saved {saved_count} of {len(matches)} repair record(s).",
            title="Invoice Review Complete", parent=self.winfo_toplevel()
        )

    def _export_repairs(self):
        # ── Scope selection dialog ─────────────────────────────────────────
        selected_ids = {int(iid) for iid in self.tree.selection()}
        n_sel = len(selected_ids)

        dlg = ttk.Toplevel(self.winfo_toplevel())
        dlg.title("Export Repairs")
        dlg.resizable(False, False)
        dlg.grab_set()

        ttk.Label(dlg, text="Export repair records for:",
                  font=("Segoe UI", fs(10), "bold")).pack(anchor=W, padx=20, pady=(18, 8))

        scope_var = tk.StringVar(value="all" if n_sel == 0 else "selected")

        sel_text = (f"Selected instruments only  ({n_sel} selected)"
                    if n_sel else "Selected instruments only  (none selected)")
        ttk.Radiobutton(dlg, text=sel_text, variable=scope_var,
                        value="selected",
                        state="normal" if n_sel else "disabled").pack(anchor=W, padx=32, pady=2)
        ttk.Radiobutton(dlg, text="All instruments", variable=scope_var,
                        value="all").pack(anchor=W, padx=32, pady=2)

        result = {"ok": False}

        def _ok():
            result["ok"] = True
            dlg.destroy()

        btn_row = ttk.Frame(dlg)
        btn_row.pack(pady=(14, 18))
        ttk.Button(btn_row, text="Export", bootstyle=PRIMARY,
                   command=_ok).pack(side=LEFT, padx=6)
        ttk.Button(btn_row, text="Cancel", bootstyle=SECONDARY,
                   command=dlg.destroy).pack(side=LEFT, padx=6)

        from ui.theme import fit_window
        fit_window(dlg, 340, 185)
        dlg.wait_window()
        if not result["ok"]:
            return

        # ── Gather instrument + repair data ───────────────────────────────
        all_instruments = [dict(r) for r in self.db.get_all_instruments(include_inactive=True)]
        if scope_var.get() == "selected" and selected_ids:
            all_instruments = [i for i in all_instruments if i["id"] in selected_ids]

        data = []
        for inst in all_instruments:
            repairs = [dict(r) for r in self.db.get_repairs(inst["id"])]
            if repairs:
                data.append((inst, repairs))

        if not data:
            Messagebox.show_info("No repair records found for the selected instruments.",
                                 title="Nothing to Export", parent=self.winfo_toplevel())
            return

        # ── Build Excel workbook ───────────────────────────────────────────
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required. Run: pip install openpyxl",
                                  title="Missing Dependency", parent=self.winfo_toplevel())
            return

        wb = openpyxl.Workbook()

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        hdr_font   = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill   = _fill("2E5E8E")   # dark blue
        inst_font  = Font(bold=True, size=9)
        inst_fill  = _fill("D9E1F2")   # light blue
        total_font = Font(bold=True, size=9)
        total_fill = _fill("E2EFDA")   # light green
        alt_fill   = _fill("F5F5F5")   # light gray
        money_fmt  = '"$"#,##0.00'

        def _row(ws, r, values, font=None, fill=None, fmts=None):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = _border()
                cell.alignment = Alignment(vertical="center")
                if font:  cell.font = font
                if fill:  cell.fill = fill
                if fmts and fmts.get(c):
                    cell.number_format = fmts[c]

        # ── Sheet 1: Summary ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.freeze_panes = "A2"

        s1_hdrs = ["Category", "Instrument", "Brand", "Model",
                   "Barcode", "District #", "Serial #", "Condition",
                   "# Repairs", "Total Spent", "Last Repaired"]
        _row(ws1, 1, s1_hdrs, font=hdr_font, fill=hdr_fill)
        ws1.row_dimensions[1].height = 18

        summary = []
        for inst, repairs in data:
            total = sum(float(r.get("act_cost") or r.get("est_cost") or 0) for r in repairs)
            last  = max((r.get("date_repaired") or r.get("date_added") or "") for r in repairs)
            summary.append((inst, repairs, total, last))
        summary.sort(key=lambda x: -x[2])

        grand_total = 0.0
        for ri, (inst, repairs, total, last) in enumerate(summary, 2):
            grand_total += total
            fill = alt_fill if ri % 2 == 0 else None
            _row(ws1, ri, [
                inst.get("category") or "",
                inst.get("description") or "",
                inst.get("brand") or "",
                inst.get("model") or "",
                inst.get("barcode") or "",
                inst.get("district_no") or "",
                inst.get("serial_no") or "",
                inst.get("condition") or "",
                len(repairs),
                total,
                last or "",
            ], fill=fill, fmts={10: money_fmt})

        gt_row = len(summary) + 2
        _row(ws1, gt_row,
             ["", "", "", "", "", "", "", "GRAND TOTAL", "", grand_total, ""],
             font=total_font, fill=total_fill, fmts={10: money_fmt})

        for col, w in zip(range(1, 12), [14, 26, 16, 16, 10, 10, 14, 12, 9, 13, 13]):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ── Sheet 2: Repair Details ───────────────────────────────────────
        ws2 = wb.create_sheet("Repair Details")
        ws2.freeze_panes = "A2"

        s2_hdrs = ["Category", "Instrument", "Brand", "Barcode",
                   "Date Added", "Description", "Shop / Location",
                   "Est. Cost", "Actual Cost", "Date Repaired", "Status"]
        _row(ws2, 1, s2_hdrs, font=hdr_font, fill=hdr_fill)
        ws2.row_dimensions[1].height = 18

        dr = 2
        for inst, repairs, total, _ in summary:
            # Instrument header spanning all columns
            label = "  ".join(filter(None, [
                inst.get("category") or "",
                inst.get("description") or "",
                inst.get("brand") or "",
                f"Barcode: {inst.get('barcode')}" if inst.get("barcode") else "",
                f"District #: {inst.get('district_no')}" if inst.get("district_no") else "",
            ]))
            _row(ws2, dr, [label] + [""] * (len(s2_hdrs) - 1),
                 font=inst_font, fill=inst_fill)
            ws2.merge_cells(start_row=dr, start_column=1,
                            end_row=dr, end_column=len(s2_hdrs))
            dr += 1

            for ridx, r in enumerate(
                sorted(repairs, key=lambda x: x.get("date_added") or "")
            ):
                fill = alt_fill if ridx % 2 == 1 else None
                status = "Completed" if r.get("date_repaired") else "Pending"
                _row(ws2, dr, [
                    inst.get("category") or "",
                    inst.get("description") or "",
                    inst.get("brand") or "",
                    inst.get("barcode") or "",
                    r.get("date_added") or "",
                    r.get("description") or "",
                    r.get("assigned_to") or r.get("location") or "",
                    float(r.get("est_cost") or 0),
                    float(r.get("act_cost") or 0),
                    r.get("date_repaired") or "",
                    status,
                ], fill=fill, fmts={8: money_fmt, 9: money_fmt})
                dr += 1

            # Subtotal row
            _row(ws2, dr,
                 ["", "", "", "", "", "", "Subtotal", "", total, "", ""],
                 font=total_font, fill=total_fill, fmts={8: money_fmt, 9: money_fmt})
            dr += 2  # blank row between instruments

        for col, w in zip(range(1, 12), [12, 22, 14, 10, 12, 34, 22, 10, 12, 14, 10]):
            ws2.column_dimensions[get_column_letter(col)].width = w

        # ── Save ──────────────────────────────────────────────────────────
        from tkinter import filedialog
        import datetime
        path = filedialog.asksaveasfilename(
            title="Save Repair Export",
            parent=self.winfo_toplevel(),
            defaultextension=".xlsx",
            initialfile=f"Repairs_{datetime.date.today().isoformat()}.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return

        try:
            wb.save(path)
        except Exception as e:
            Messagebox.show_error(f"Could not save file:\n{e}", title="Save Error",
                                  parent=self.winfo_toplevel())
            return

        answer = Messagebox.yesno(
            "Repair data exported successfully.\n\nOpen the file now?",
            title="Export Complete", parent=self.winfo_toplevel()
        )
        if answer == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)

    # ── Excel helpers (shared by the inventory / repairs-needed exports) ────────

    def _excel_styles(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        return {
            "hdr_font":  Font(bold=True, color="FFFFFF", size=10),
            "hdr_fill":  _fill("2E5E8E"),
            "total_font": Font(bold=True, size=9),
            "total_fill": _fill("E2EFDA"),
            "alt_fill":  _fill("F5F5F5"),
            "border":    _border,
            "align":     Alignment,
            "money_fmt": '"$"#,##0.00',
        }

    def _export_barcodes(self):
        """A printable Code128 sheet for the handheld scanner — brass & woodwind
        instruments (percussion excluded), sorted by type, 3 columns."""
        try:
            import barcode_labels as bl
        except Exception as e:
            Messagebox.show_error(f"Could not load the barcode tool:\n{e}",
                                  title="Error", parent=self.winfo_toplevel())
            return
        instruments = self.db.get_all_instruments()
        printable = [r for r in instruments
                     if bl.instrument_family(r["category"]) is not None
                     and ((r["barcode"] or "").strip()
                          or (r["district_no"] or "").strip())]
        if not printable:
            Messagebox.show_info(
                "No brass or woodwind instruments with a barcode were found.",
                title="Nothing to Print", parent=self.winfo_toplevel())
            return
        from tkinter import filedialog
        import datetime as _d
        path = filedialog.asksaveasfilename(
            title="Save Barcode Sheet", parent=self.winfo_toplevel(),
            defaultextension=".pdf",
            initialfile=f"Instrument_Barcodes_{_d.date.today().isoformat()}.pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if not path:
            return
        try:
            n = bl.export_instrument_barcodes(instruments, path)
        except Exception as e:
            Messagebox.show_error(f"Could not create the PDF:\n{e}",
                                  title="Export Error", parent=self.winfo_toplevel())
            return
        if Messagebox.yesno(f"Created a barcode sheet with {n} instrument(s).\n\n"
                            "Open it now?", title="Barcode Sheet",
                            parent=self.winfo_toplevel()) == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)

    def _save_and_open(self, wb, default_name):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            title="Save Export", parent=self.winfo_toplevel(),
            defaultextension=".xlsx", initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            wb.save(path)
        except Exception as e:
            Messagebox.show_error(f"Could not save file:\n{e}", title="Save Error",
                                  parent=self.winfo_toplevel())
            return
        if Messagebox.yesno("Export complete.\n\nOpen the file now?",
                            title="Export Complete", parent=self.winfo_toplevel()) == "Yes":
            import subprocess
            subprocess.Popen(["start", "", path], shell=True)

    # ── #6: Full inventory export for district personnel ────────────────────────

    def _export_inventory(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required. Run: pip install openpyxl",
                                  title="Missing Dependency", parent=self.winfo_toplevel())
            return

        include_inactive = bool(self._show_inactive.get())
        instruments = [dict(r) for r in self.db.get_all_instruments(include_inactive=include_inactive)]
        if not instruments:
            Messagebox.show_info("No instruments to export.", title="Nothing to Export",
                                 parent=self.winfo_toplevel())
            return

        import datetime
        this_year = datetime.date.today().year

        def _age(inst):
            # Age is based on the PRODUCTION year (serial-dated) when known,
            # falling back to purchase year.
            yr = str(inst.get("year_manufactured") or inst.get("year_purchased")
                     or "").strip()[:4]
            if yr.isdigit():
                a = this_year - int(yr)
                return a if a >= 0 else ""
            return ""

        st = self._excel_styles()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instrument Inventory"
        ws.freeze_panes = "A2"

        headers = ["Category", "Instrument", "Brand", "Model", "Serial #",
                   "Barcode", "Condition", "Status",
                   "Assigned To", "Year Purchased", "Year Made", "Age (yrs)",
                   "Amount Paid", "Est. Value", "Repair $ Spent", "Location", "Comments"]

        def _row(r, values, font=None, fill=None, money_cols=()):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = st["border"]()
                cell.alignment = Alignment(vertical="center")
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if c in money_cols:
                    cell.number_format = st["money_fmt"]

        _row(1, headers, font=st["hdr_font"], fill=st["hdr_fill"])
        ws.row_dimensions[1].height = 18

        money_cols = (13, 14, 15)
        grand_paid = grand_value = grand_repair = 0.0

        # An instrument is "unavailable" (still district property, but not usable
        # by students) if it's been deactivated, marked Unrepairable, or its
        # condition/comments note it as lost/missing/etc.  These are moved to a
        # section at the very bottom so the usable list isn't cluttered.
        _UNAVAIL_WORDS = ("lost", "missing", "stolen", "retired", "disposed",
                          "out of service", "unavailable", "scrap", "written off",
                          "write-off", "beyond repair")

        def _unavailable(inst):
            if "is_active" in inst and not inst.get("is_active", 1):
                return True
            cond = (inst.get("condition") or "").strip().lower()
            if cond in ("unrepairable", "unrepairable / lost"):
                return True
            blob = cond + " " + (inst.get("comments") or "").lower()
            return any(w in blob for w in _UNAVAIL_WORDS)

        def _emit(inst, r):
            nonlocal grand_paid, grand_value, grand_repair
            repairs = [dict(x) for x in self.db.get_repairs(inst["id"])]
            # Only actual recorded costs count as money spent — estimates are not
            # reliable enough to report to the district.
            repair_spent = sum(float(x.get("act_cost") or 0) for x in repairs)
            loan = self.db.get_active_loan(inst["id"])
            actives = self.db.get_active_checkouts_for_instrument(inst["id"])
            if loan:
                status = "On Loan"
                assigned = loan["school"] or ""
            elif actives:
                status = "Checked Out"
                assigned = ", ".join(a["student_name"] or "?" for a in actives)
            else:
                status = "Available"
                assigned = ""
            paid = float(inst.get("amount_paid") or 0)
            value = float(inst.get("est_value") or 0)
            grand_paid += paid; grand_value += value; grand_repair += repair_spent
            fill = st["alt_fill"] if r % 2 == 0 else None
            _row(r, [
                inst.get("category") or "",
                inst.get("description") or "",
                inst.get("brand") or "",
                inst.get("model") or "",
                inst.get("serial_no") or "",
                inst.get("barcode") or inst.get("district_no") or "",
                inst.get("condition") or "",
                status,
                assigned,
                inst.get("year_purchased") or inst.get("date_purchased") or "",
                inst.get("year_manufactured") or "",
                _age(inst),
                paid,
                value,
                repair_spent,
                inst.get("locker") or "",
                inst.get("comments") or "",
            ], fill=fill, money_cols=money_cols)
            return r + 1

        skey = lambda x: (x.get("category") or "", x.get("description") or "")
        available = [i for i in instruments if not _unavailable(i)]
        unavailable = [i for i in instruments if _unavailable(i)]

        r = 2
        for inst in sorted(available, key=skey):
            r = _emit(inst, r)

        if unavailable:
            banner = ("Unavailable — still district property, not usable by "
                      "students (unrepairable / lost / out of service)")
            _row(r, [banner] + [""] * (len(headers) - 1),
                 font=st["total_font"], fill=st["total_fill"])
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=len(headers))
            r += 1
            for inst in sorted(unavailable, key=skey):
                r = _emit(inst, r)

        _row(r, ["", "", "", "", "", "", "", "", "", "", "", "TOTALS",
                 grand_paid, grand_value, grand_repair, "", ""],
             font=st["total_font"], fill=st["total_fill"], money_cols=money_cols)

        widths = [14, 24, 14, 14, 14, 14, 12, 12, 18, 12, 12, 9, 12, 12, 13, 12, 30]
        for col, w in zip(range(1, len(widths) + 1), widths):
            ws.column_dimensions[get_column_letter(col)].width = w

        import datetime as _d
        self._save_and_open(wb, f"Instrument_Inventory_{_d.date.today().isoformat()}.xlsx")

    # ── #8: Repairs-needed printout for the technician ──────────────────────────

    def _export_repairs_needed(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required. Run: pip install openpyxl",
                                  title="Missing Dependency", parent=self.winfo_toplevel())
            return

        pending = [dict(r) for r in self.db.get_instruments_needing_repair()]
        if not pending:
            Messagebox.show_info(
                "No outstanding repairs. Everything is marked repaired!",
                title="Nothing to Export", parent=self.winfo_toplevel())
            return

        PRI = {3: "Urgent", 2: "High", 1: "Normal", 0: "Low"}
        st = self._excel_styles()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Repairs Needed"
        ws.freeze_panes = "A2"

        headers = ["Priority", "Category", "Instrument", "Brand", "Model",
                   "Serial #", "Barcode", "Repair(s) Needed",
                   "Reported", "Shop / Location"]

        def _row(r, values, font=None, fill=None):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = st["border"]()
                cell.alignment = Alignment(vertical="center", wrap_text=(c == 8))
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill

        _row(1, headers, font=st["hdr_font"], fill=st["hdr_fill"])
        ws.row_dimensions[1].height = 18

        r = 2
        for rep in pending:
            fill = st["alt_fill"] if r % 2 == 0 else None
            _row(r, [
                PRI.get(int(rep.get("max_priority") or 0), ""),
                rep.get("category") or "",
                rep.get("instrument_desc") or "",
                rep.get("brand") or "",
                rep.get("model") or "",
                rep.get("serial_no") or "",
                rep.get("barcode") or rep.get("district_no") or "",
                rep.get("needs") or "",
                rep.get("last_reported") or "",
                rep.get("shop") or "",
            ], fill=fill)
            r += 1

        widths = [9, 14, 22, 14, 14, 14, 14, 46, 13, 18]
        for col, w in zip(range(1, len(widths) + 1), widths):
            ws.column_dimensions[get_column_letter(col)].width = w

        import datetime as _d
        self._save_and_open(wb, f"Repairs_Needed_{_d.date.today().isoformat()}.xlsx")

    # ── Currently-checked-out export ────────────────────────────────────────────

    def _export_checked_out(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required. Run: pip install openpyxl",
                                  title="Missing Dependency", parent=self.winfo_toplevel())
            return

        rows = [dict(r) for r in self.db.get_all_active_checkouts()]
        if not rows:
            Messagebox.show_info("Nothing is currently checked out.",
                                 title="Nothing to Export", parent=self.winfo_toplevel())
            return

        st = self._excel_styles()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Checked Out"
        ws.freeze_panes = "A2"
        headers = ["Checked Out To", "Item / Instrument", "Category",
                   "Barcode", "Serial #", "Date Out", "Due Date", "Type"]

        def _row(r, values, font=None, fill=None):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = st["border"]()
                cell.alignment = Alignment(vertical="center")
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill

        _row(1, headers, font=st["hdr_font"], fill=st["hdr_fill"])
        ws.row_dimensions[1].height = 18
        r = 2
        for c in rows:
            is_item = not c.get("instrument_id")
            fill = st["alt_fill"] if r % 2 == 0 else None
            _row(r, [
                c.get("student_name") or "",
                c.get("description") or "",
                c.get("category") or "",
                c.get("barcode") or c.get("district_no") or "",
                c.get("serial_no") or "",
                c.get("date_assigned") or "",
                c.get("due_date") or "",
                "Item" if is_item else "Instrument",
            ], fill=fill)
            r += 1
        # Instruments out on loan to other schools
        for l in (dict(x) for x in self.db.get_all_active_loans()):
            who = l.get("school") or ""
            if l.get("contact_name"):
                who += f" — {l['contact_name']}"
            fill = st["alt_fill"] if r % 2 == 0 else None
            _row(r, [
                who,
                l.get("description") or "",
                l.get("category") or "",
                l.get("barcode") or l.get("district_no") or "",
                l.get("serial_no") or "",
                l.get("date_out") or "",
                l.get("date_due") or "",
                "On Loan",
            ], fill=fill)
            r += 1

        widths = [22, 26, 14, 14, 14, 13, 13, 12]
        for col, w in zip(range(1, len(widths) + 1), widths):
            ws.column_dimensions[get_column_letter(col)].width = w

        import datetime as _d
        self._save_and_open(wb, f"Checked_Out_{_d.date.today().isoformat()}.xlsx")

    # ── Repair cost report (for district / PTSA replacement decisions) ──────────

    def _export_cost_report(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            Messagebox.show_error("openpyxl is required. Run: pip install openpyxl",
                                  title="Missing Dependency", parent=self.winfo_toplevel())
            return

        summary = [dict(r) for r in self.db.get_repair_cost_summary()]
        if not summary:
            Messagebox.show_info("No repair spending recorded yet.",
                                 title="Nothing to Export", parent=self.winfo_toplevel())
            return

        import datetime
        this_year = datetime.date.today().year
        st = self._excel_styles()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Repair Cost Analysis"
        ws.freeze_panes = "A2"
        headers = ["Rank", "Category", "Instrument", "Brand", "Serial #", "Barcode",
                   "# Repairs", "Total Repair $", "Est. Value", "Amount Paid",
                   "Age (yrs)", "Last Repaired"]

        def _row(r, values, font=None, fill=None, money_cols=()):
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = st["border"]()
                cell.alignment = Alignment(vertical="center")
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if c in money_cols:
                    cell.number_format = st["money_fmt"]

        _row(1, headers, font=st["hdr_font"], fill=st["hdr_fill"])
        ws.row_dimensions[1].height = 18
        money_cols = (8, 9, 10)
        grand = 0.0
        r = 2
        for rank, s in enumerate(summary, 1):
            total = float(s.get("total_spent") or 0)
            grand += total
            yr = str(s.get("year_purchased") or "").strip()[:4]
            age = (this_year - int(yr)) if yr.isdigit() else ""
            fill = st["alt_fill"] if r % 2 == 0 else None
            _row(r, [
                rank,
                s.get("category") or "",
                s.get("instrument_desc") or "",
                s.get("brand") or "",
                s.get("serial_no") or "",
                s.get("barcode") or s.get("district_no") or "",
                s.get("repair_count") or 0,
                total,
                float(s.get("est_value") or 0),
                float(s.get("amount_paid") or 0),
                age,
                s.get("last_repair") or "",
            ], fill=fill, money_cols=money_cols)
            r += 1

        _row(r, ["", "", "", "", "", "", "GRAND TOTAL", grand, "", "", "", ""],
             font=st["total_font"], fill=st["total_fill"], money_cols=money_cols)

        widths = [6, 14, 24, 14, 14, 14, 10, 14, 12, 12, 9, 13]
        for col, w in zip(range(1, len(widths) + 1), widths):
            ws.column_dimensions[get_column_letter(col)].width = w

        import datetime as _d
        self._save_and_open(wb, f"Repair_Cost_Report_{_d.date.today().isoformat()}.xlsx")
